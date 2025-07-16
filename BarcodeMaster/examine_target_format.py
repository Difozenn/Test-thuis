#!/usr/bin/env python3
"""
Examine the target 1.xlsx file to understand the expected format
"""

import openpyxl
import pandas as pd

def examine_target_excel(file_path: str):
    """Examine the target Excel file structure"""
    
    print(f"📊 Examining target file: {file_path}")
    
    try:
        # Load with openpyxl for detailed inspection
        workbook = openpyxl.load_workbook(file_path)
        
        print(f"📄 Found {len(workbook.sheetnames)} sheets:")
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            print(f"   Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            
            # Show first few rows
            print("   First 10 rows:")
            for row_num in range(1, min(11, sheet.max_row + 1)):
                row = sheet[row_num]
                row_values = [str(cell.value or '') for cell in row]
                # Only show non-empty rows
                if any(val.strip() for val in row_values):
                    print(f"     Row {row_num}: {row_values[:8]}...")  # First 8 columns
            
            # Show structure analysis
            print(f"   Analysis:")
            
            # Check for headers
            if sheet.max_row > 0:
                first_row = [str(cell.value or '') for cell in sheet[1]]
                header_keywords = ['N°', 'Onderdeel', 'Materiaal', 'Lengte', 'Breedte', 'Dikte', 'L1', 'L2', 'B1', 'B2', 'Pro.methode']
                found_headers = [kw for kw in header_keywords if any(kw.lower() in val.lower() for val in first_row)]
                print(f"     Headers found: {found_headers}")
                
                # Count data rows
                data_rows = 0
                for row_num in range(2, sheet.max_row + 1):
                    row = sheet[row_num]
                    if any(str(cell.value or '').strip() for cell in row):
                        data_rows += 1
                
                print(f"     Data rows: {data_rows}")
        
        workbook.close()
        
        # Also try pandas for different view
        print(f"\n📊 Pandas view:")
        try:
            xl_file = pd.ExcelFile(file_path)
            print(f"   Sheet names: {xl_file.sheet_names}")
            
            for sheet_name in xl_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                print(f"\n   Sheet {sheet_name}:")
                print(f"     Shape: {df.shape}")
                print(f"     Columns: {list(df.columns)}")
                if not df.empty:
                    print(f"     First few rows:")
                    print(df.head(3).to_string())
                    
        except Exception as e:
            print(f"   Pandas read failed: {e}")
        
    except Exception as e:
        print(f"❌ Error examining file: {e}")

if __name__ == "__main__":
    examine_target_excel("1.xlsx")