#!/usr/bin/env python3
"""
PDF to Excel Converter
Converts manufacturing PDFs to Excel format matching the exact structure of 1.xlsx
"""

import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
from typing import Dict, List, Tuple, Any
import os


class ManufacturingPDFConverter:
    """Converts manufacturing PDFs to Excel with exact 1.xlsx structure"""
    
    def __init__(self, template_path: str = "1.xlsx"):
        """Initialize with template Excel file"""
        self.template_path = template_path
        self.template_structure = self._analyze_template()
        
    def _analyze_template(self) -> Dict[str, Any]:
        """Analyze the template Excel file structure"""
        if not os.path.exists(self.template_path):
            print(f"Warning: Template {self.template_path} not found")
            return {}
            
        wb = openpyxl.load_workbook(self.template_path)
        structure = {
            'sheets': [],
            'column_widths': {},
            'merged_cells': {},
            'styles': {}
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            structure['sheets'].append(sheet_name)
            
            # Capture column widths
            widths = {}
            for col in sheet.column_dimensions:
                widths[col] = sheet.column_dimensions[col].width
            structure['column_widths'][sheet_name] = widths
            
            # Capture merged cells
            structure['merged_cells'][sheet_name] = list(sheet.merged_cells.ranges)
            
        return structure
    
    def convert_pdf(self, pdf_path: str, output_path: str = None) -> str:
        """Convert PDF to Excel matching template structure"""
        if not output_path:
            output_path = pdf_path.replace('.pdf', '_converted.xlsx').replace('.PDF', '_converted.xlsx')
        
        # Extract data from PDF
        pdf_data = self._extract_pdf_data(pdf_path)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create cover sheet (Table 1)
        self._create_cover_sheet(wb, pdf_data['metadata'])
        
        # Create data sheets for each table section
        sheet_num = 2
        for section in pdf_data['sections']:
            self._create_data_sheet(wb, section, pdf_data['metadata'], sheet_num)
            sheet_num += 1
        
        # Save workbook
        wb.save(output_path)
        print(f"✅ Excel file created: {output_path}")
        return output_path
    
    def _extract_pdf_data(self, pdf_path: str) -> Dict[str, Any]:
        """Extract all data from PDF"""
        print(f"📄 Extracting data from: {pdf_path}")
        
        all_text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text += text + "\n"
        
        # Parse metadata
        metadata = self._parse_metadata(all_text)
        
        # Parse sections
        sections = self._parse_sections(all_text)
        
        return {
            'metadata': metadata,
            'sections': sections
        }
    
    def _parse_metadata(self, text: str) -> Dict[str, str]:
        """Extract project metadata from PDF text"""
        lines = text.split('\n')[:50]  # Check first 50 lines
        text_block = ' '.join(lines)
        
        metadata = {
            'project_code': '',
            'sales_number': '',
            'project_name': '',
            'client': '',
            'designer': '',
            'department': ''
        }
        
        # Extract patterns
        mo_match = re.search(r'(MO\d+(?:-\d+)?)', text_block)
        if mo_match:
            metadata['project_code'] = mo_match.group(1)
        
        s_match = re.search(r'(S\d+)', text_block)
        if s_match:
            metadata['sales_number'] = s_match.group(1)
            metadata['department'] = s_match.group(1)
        
        # Project name
        project_match = re.search(r'0411_MO\d+[-\d]*_([^"\n]+?)(?:\s*\()', text_block)
        if project_match:
            metadata['project_name'] = project_match.group(1).strip()
        
        # Client
        client_match = re.search(r'(?:Klant:|Client:)\s*([A-Za-z\s]+?)(?:\n|Tekenaar|JW)', text_block)
        if client_match:
            metadata['client'] = client_match.group(1).strip()
        
        # Designer
        designer_match = re.search(r'(?:Tekenaar:|Designer:)\s*([A-Z]+)', text_block)
        if designer_match:
            metadata['designer'] = designer_match.group(1)
        elif 'JW' in text_block:
            metadata['designer'] = 'JW'
        
        return metadata
    
    def _parse_sections(self, text: str) -> List[Dict[str, Any]]:
        """Parse all table sections from PDF"""
        sections = []
        lines = text.split('\n')
        
        section_types = ['Nesting', 'Opdeelzaag', 'Massief', 'Controle', 'Magazijn']
        current_section = None
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Check for section start
            for section_type in section_types:
                if section_type.lower() in line.lower():
                    # Save previous section
                    if current_section and current_section['data']:
                        sections.append(current_section)
                    
                    # Start new section
                    current_section = {
                        'type': section_type.lower(),
                        'headers': [],
                        'data': []
                    }
                    break
            
            # Look for headers (line with N°)
            if current_section and 'N°' in line and not current_section['headers']:
                headers = self._parse_headers(line)
                current_section['headers'] = headers
            
            # Parse data rows
            elif current_section and current_section['headers'] and re.match(r'^\s*\d+\s+', line):
                row_data = self._parse_data_row(line, current_section['type'], lines, i)
                if row_data:
                    current_section['data'].append(row_data)
            
            # Check for section end
            if 'Aantal onderdelen:' in line:
                match = re.search(r'Aantal onderdelen:\s*(\d+)', line)
                if match and current_section:
                    current_section['total'] = int(match.group(1))
                    sections.append(current_section)
                    current_section = None
            
            i += 1
        
        # Don't forget last section
        if current_section and current_section['data']:
            sections.append(current_section)
        
        return sections
    
    def _parse_headers(self, line: str) -> List[str]:
        """Parse header line"""
        # Common headers for different sections
        if 'Beschrijving' in line:
            return ['N°', 'Beschrijving', 'Aantal stuks', 'GB nummer']
        elif 'L1' in line and 'L2' in line:
            return ['N°', 'Onderdeel', 'Materiaal', 'Lengte', 'Breedte', 'Dikte', 'L1', 'L2', 'B1', 'B2', 'ProductieM.', 'Opmerkingen']
        else:
            return ['N°', 'Onderdeel', 'Materiaal', 'Lengte', 'Breedte', 'Dikte', 'Opmerkingen']
    
    def _parse_data_row(self, line: str, section_type: str, all_lines: List[str], line_idx: int) -> Dict[str, Any]:
        """Parse a data row based on section type"""
        # Get the full row content (might span multiple lines)
        full_line = line
        
        # Check if "1mm" values are on next line
        if line_idx + 1 < len(all_lines):
            next_line = all_lines[line_idx + 1].strip()
            if '1mm' in next_line and not re.match(r'^\s*\d+\s+', next_line):
                full_line += ' ' + next_line
        
        # Extract row number
        match = re.match(r'^\s*(\d+)\s+(.+)$', line)
        if not match:
            return None
        
        row_num = match.group(1)
        rest = match.group(2)
        
        if section_type in ['nesting', 'opdeelzaag']:
            return self._parse_production_row(row_num, rest, full_line)
        elif section_type == 'magazijn':
            return self._parse_magazijn_row(row_num, rest)
        else:
            return self._parse_generic_row(row_num, rest)
    
    def _parse_production_row(self, row_num: str, rest: str, full_line: str) -> Dict[str, Any]:
        """Parse nesting/opdeelzaag row with L1/L2/B1/B2"""
        parts = rest.split()
        
        data = {'N°': row_num}
        
        # Parse components
        idx = 0
        
        # Onderdeel (part code)
        data['Onderdeel'] = parts[idx] if idx < len(parts) else ''
        idx += 1
        
        # Material (collect until number)
        material_parts = []
        while idx < len(parts) and not re.match(r'^\d+\.?\d*$', parts[idx]):
            material_parts.append(parts[idx])
            idx += 1
        data['Materiaal'] = ' '.join(material_parts)
        
        # Dimensions
        data['Lengte'] = parts[idx] if idx < len(parts) else ''
        idx += 1
        data['Breedte'] = parts[idx] if idx < len(parts) else ''
        idx += 1
        data['Dikte'] = parts[idx] if idx < len(parts) else ''
        idx += 1
        
        # Parse Fineer/edge data
        # Count "Fineer eik" occurrences (they might be concatenated)
        fineer_count = len(re.findall(r'Fineer\s*eik', full_line))
        
        data['L1'] = 'Fineer eik 1mm' if fineer_count >= 1 else ''
        data['L2'] = 'Fineer eik 1mm' if fineer_count >= 2 else ''
        data['B1'] = 'Fineer eik 1mm' if fineer_count >= 3 else ''
        data['B2'] = 'Fineer eik 1mm' if fineer_count >= 4 else ''
        
        # ProductieM
        if 'Standaard' in full_line:
            data['ProductieM.'] = 'Standaard'
            if 'Dik' in full_line:
                data['ProductieM.'] += ' Dik'
        else:
            data['ProductieM.'] = ''
        
        # Opmerkingen
        remarks = []
        if 'overmaat' in full_line:
            overmaat_match = re.search(r'([LB+]=\d+mm overmaat|[\w\s]*\d+mm overmaat)', full_line)
            if overmaat_match:
                remarks.append(overmaat_match.group(1))
        if 'frezen' in full_line.lower():
            frezen_match = re.search(r'([\w\s]*frezen[\w\s,]*)', full_line, re.IGNORECASE)
            if frezen_match:
                remarks.append(frezen_match.group(1))
        
        data['Opmerkingen'] = ', '.join(remarks)
        
        return data
    
    def _parse_magazijn_row(self, row_num: str, rest: str) -> Dict[str, Any]:
        """Parse magazijn row"""
        # Try to extract pattern: description, number, GB code
        match = re.match(r'^(.+?)\s+(\d+)\s*(.*)$', rest)
        
        if match:
            return {
                'N°': row_num,
                'Beschrijving': match.group(1).strip(),
                'Aantal stuks': match.group(2),
                'GB nummer': match.group(3).strip()
            }
        else:
            return {
                'N°': row_num,
                'Beschrijving': rest,
                'Aantal stuks': '',
                'GB nummer': ''
            }
    
    def _parse_generic_row(self, row_num: str, rest: str) -> Dict[str, Any]:
        """Parse generic row"""
        parts = rest.split()
        return {
            'N°': row_num,
            'Data': ' '.join(parts)
        }
    
    def _create_cover_sheet(self, wb: openpyxl.Workbook, metadata: Dict[str, str]):
        """Create cover sheet matching 1.xlsx Table 1"""
        ws = wb.create_sheet("Table 1")
        
        # Set content
        ws['A1'] = 'Project:\nKlant:\nTekenaar:'
        ws['D1'] = f"{metadata['project_code']}\n{metadata['sales_number']}\n{metadata['project_name']} {metadata['client']}\n{metadata['designer']}"
        ws['A2'] = metadata['department']
        ws['A3'] = 'info:\nSchuren'
        ws['A4'] = 'Totaal aantal onderdelen:'
        ws['A5'] = 'Afwerking: Lakstraat'
        ws['A6'] = 'Enkel als aangevinkt.                  Handwerk voor het schuren.\nKasten monteren! onderdelen sorteren per object Vlakstraat: gekleurde sjang gebruiken.'
        ws['A7'] = 'Datum:'
        ws['B7'] = 'kopie: terugbezorgen na schuren!'
        
        # Signature sections
        sections = ['Cel Holzer:', 'Accura:', 'Reichenbacher:', 'Kl Gannomat:', 'Cel Massief:', 'Cel schuren:']
        for i, section in enumerate(sections):
            ws[f'A{8+i}'] = f'{section}\nNaam:                          .../...'
        
        ws['C8'] = 'Opmerkingen:'
        
        # Apply merges
        merges = [
            'A1:C1', 'D1:F1', 'A2:B2', 'A3:B3', 'A4:B4', 'A5:F5',
            'A6:D6', 'B7:F7', 'A8:B8', 'C8:E13', 'A9:B9', 'A10:B10',
            'A11:B11', 'A12:B12', 'A13:B13', 'A14:F14'
        ]
        for merge in merges:
            ws.merge_cells(merge)
        
        # Set column widths
        ws.column_dimensions['A'].width = 37.56
        ws.column_dimensions['B'].width = 2
        ws.column_dimensions['C'].width = 11.56
        ws.column_dimensions['D'].width = 31.78
        ws.column_dimensions['E'].width = 42.22
        ws.column_dimensions['F'].width = 2.89
    
    def _create_data_sheet(self, wb: openpyxl.Workbook, section: Dict[str, Any], 
                          metadata: Dict[str, str], sheet_num: int):
        """Create data sheet for a section"""
        ws = wb.create_sheet(f"Table {sheet_num}")
        
        # Header section
        ws['A1'] = metadata['department']
        ws['A2'] = 'Klant:'
        ws['G2'] = metadata['client']
        ws['O2'] = metadata['project_code']
        ws['A3'] = f"Tekenaar:  {metadata['designer']}\nSales nr:    {metadata['sales_number']}"
        ws['A4'] = 'Schuren'
        ws['E4'] = 'Project:'
        ws['H4'] = f"{metadata['project_name']}                                                       {section['type'].capitalize()}"
        
        # Column headers based on section type
        if section['type'] in ['nesting', 'opdeelzaag']:
            ws['A5'] = 'N°'
            ws['B5'] = 'Onderdeel'
            ws['D5'] = 'Materiaal'
            ws['F5'] = 'Lengte'
            ws['I5'] = 'Breedte'
            ws['J5'] = 'Dikte'
            ws['K5'] = 'L1'
            ws['L5'] = 'L2'
            ws['M5'] = 'B1'
            ws['N5'] = 'B2'
            if section['type'] == 'nesting':
                ws['P5'] = 'ProductieM.'
                ws['Q5'] = 'Opmerkingen'
            else:
                ws['P5'] = 'Opmerkingen'
        elif section['type'] == 'magazijn':
            ws['A5'] = 'N°'
            ws['B5'] = 'Beschrijving'
            ws['I5'] = 'Aantal stuks'
            ws['P5'] = 'GB nummer'
        
        # Data rows
        row = 6
        for item in section['data']:
            if section['type'] in ['nesting', 'opdeelzaag']:
                ws[f'A{row}'] = item.get('N°', '')
                ws[f'B{row}'] = item.get('Onderdeel', '')
                ws[f'D{row}'] = item.get('Materiaal', '')
                ws[f'F{row}'] = item.get('Lengte', '')
                ws[f'I{row}'] = item.get('Breedte', '')
                ws[f'J{row}'] = item.get('Dikte', '')
                ws[f'K{row}'] = item.get('L1', '')
                ws[f'L{row}'] = item.get('L2', '')
                ws[f'M{row}'] = item.get('B1', '')
                ws[f'N{row}'] = item.get('B2', '')
                if section['type'] == 'nesting':
                    ws[f'P{row}'] = item.get('ProductieM.', '')
                    ws[f'Q{row}'] = item.get('Opmerkingen', '')
                else:
                    ws[f'P{row}'] = item.get('Opmerkingen', '')
            elif section['type'] == 'magazijn':
                ws[f'A{row}'] = item.get('N°', '')
                ws[f'B{row}'] = item.get('Beschrijving', '')
                ws[f'I{row}'] = item.get('Aantal stuks', '')
                ws[f'P{row}'] = item.get('GB nummer', '')
            row += 1
        
        # Total row
        total = section.get('total', len(section['data']))
        ws[f'A{row}'] = f'Aantal onderdelen: {total}'
        
        # Apply merges
        self._apply_data_sheet_merges(ws, len(section['data']))
        
        # Set column widths
        self._set_column_widths(ws)
    
    def _apply_data_sheet_merges(self, ws, data_row_count: int):
        """Apply merges for data sheets"""
        # Header merges
        merges = [
            'A1:B1', 'A2:F2', 'G2:N2', 'O2:R2', 'A3:R3',
            'A4:D4', 'E4:G4', 'H4:R4',
            # Column header merges
            'B5:C5', 'D5:E5', 'F5:H5', 'N5:O5'
        ]
        
        # Data row merges
        for i in range(6, 6 + data_row_count):
            merges.extend([
                f'B{i}:C{i}', f'D{i}:E{i}', f'F{i}:H{i}', f'N{i}:O{i}'
            ])
        
        # Total row merge
        total_row = 6 + data_row_count
        merges.append(f'A{total_row}:R{total_row}')
        
        for merge in merges:
            try:
                ws.merge_cells(merge)
            except:
                pass  # Skip if merge fails
    
    def _set_column_widths(self, ws):
        """Set standard column widths"""
        widths = {
            'A': 5.11, 'B': 9.33, 'C': 14.44, 'D': 26.44, 'E': 2.44,
            'F': 7.33, 'G': 0.67, 'H': 5.11, 'I': 11.78, 'J': 7.78,
            'K': 6.44, 'L': 6.44, 'M': 6.67, 'N': 3.11, 'O': 3.11,
            'P': 7.78, 'Q': 57.56, 'R': 3.11
        }
        
        for col, width in widths.items():
            ws.column_dimensions[col].width = width


def main():
    """Test the converter"""
    converter = ManufacturingPDFConverter(template_path="1.xlsx")
    
    pdf_file = "S04479_RAPPORT_Rudi Matterne_0411_MO07199_Hoekdressing - opklapbed (4-7).PDF"
    output_file = "hoekdressing_perfect_conversion.xlsx"
    
    print(f"🔄 Converting {pdf_file}...")
    print(f"📋 Using template structure from 1.xlsx")
    
    converter.convert_pdf(pdf_file, output_file)
    
    # Verify the conversion
    print("\n📊 Verifying conversion...")
    wb = openpyxl.load_workbook(output_file)
    print(f"Created {len(wb.sheetnames)} sheets: {wb.sheetnames}")
    
    # Check L1/L2/B1/B2 content
    for sheet_name in wb.sheetnames:
        if sheet_name != "Table 1":
            sheet = wb[sheet_name]
            if sheet['K5'].value == 'L1':  # Has L1/L2/B1/B2 columns
                print(f"\n{sheet_name} - Edge processing data:")
                fineer_count = 0
                for row in range(6, 15):  # Check first few rows
                    l1 = sheet[f'K{row}'].value
                    l2 = sheet[f'L{row}'].value
                    b1 = sheet[f'M{row}'].value
                    b2 = sheet[f'N{row}'].value
                    if any([l1, l2, b1, b2]):
                        sides = sum(1 for x in [l1, l2, b1, b2] if x and 'Fineer' in str(x))
                        fineer_count += sides
                        if sides > 0:
                            print(f"  Row {row}: {sides} sides")
                print(f"  Total Fineer sides: {fineer_count}")


if __name__ == "__main__":
    main()