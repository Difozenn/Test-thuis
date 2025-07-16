#!/usr/bin/env python3
"""
PDF to Excel Parser - Automated PDF conversion and robust Excel parsing
"""

import os
import subprocess
import tempfile
from pathlib import Path
import pandas as pd
import openpyxl
from typing import List, Dict, Optional
import re

class PDFToExcelParser:
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp()
    
    def parse_pdf_via_excel(self, pdf_path: str) -> Dict[str, int]:
        """Main method: Convert PDF to Excel and extract data"""
        
        print("=== PDF TO EXCEL CONVERSION ===")
        
        # Step 1: Convert PDF to Excel
        excel_path = self._convert_pdf_to_excel(pdf_path)
        
        if not excel_path:
            raise Exception("PDF to Excel conversion failed")
        
        # Step 2: Parse Excel file with 100% robust extraction
        results = self._parse_excel_file(excel_path)
        
        # Clean up temp files
        self._cleanup_temp_files()
        
        return results
    
    def _convert_pdf_to_excel(self, pdf_path: str) -> Optional[str]:
        """Convert PDF to Excel using multiple robust methods"""
        
        excel_path = os.path.join(self.temp_dir, "converted_tables.xlsx")
        
        # Method 1: Try tabula-py (Java-based, very robust)
        if self._try_tabula_conversion(pdf_path, excel_path):
            return excel_path
        
        # Method 2: Try camelot (if available)
        if self._try_camelot_conversion(pdf_path, excel_path):
            return excel_path
        
        # Method 3: Try pdfplumber + pandas Excel export
        if self._try_pdfplumber_excel_export(pdf_path, excel_path):
            return excel_path
        
        print("❌ All PDF to Excel conversion methods failed")
        return None
    
    def _try_tabula_conversion(self, pdf_path: str, excel_path: str) -> bool:
        """Try conversion using tabula-py"""
        try:
            import tabula
            print("🔄 Trying tabula-py conversion...")
            
            # Extract all tables from PDF (pages 11-25 for BOERE)
            dfs = tabula.read_pdf(
                pdf_path, 
                pages="11-25",  # Controle to Magazijn pages
                multiple_tables=True,
                pandas_options={'header': None}  # Don't assume headers
            )
            
            if dfs:
                # Combine all tables into one Excel file with multiple sheets
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    for i, df in enumerate(dfs):
                        if not df.empty:
                            df.to_excel(writer, sheet_name=f'Table_{i}', index=False, header=False)
                
                print(f"✅ Tabula conversion successful: {len(dfs)} tables extracted")
                return True
                
        except Exception as e:
            print(f"❌ Tabula conversion failed: {e}")
        
        return False
    
    def _try_camelot_conversion(self, pdf_path: str, excel_path: str) -> bool:
        """Try conversion using camelot"""
        try:
            import camelot
            print("🔄 Trying camelot conversion...")
            
            # Extract tables
            tables = camelot.read_pdf(pdf_path, pages='11-25', flavor='lattice')
            
            if tables:
                # Export to Excel
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    for i, table in enumerate(tables):
                        table.df.to_excel(writer, sheet_name=f'Table_{i}', index=False, header=False)
                
                print(f"✅ Camelot conversion successful: {len(tables)} tables extracted")
                return True
                
        except Exception as e:
            print(f"❌ Camelot conversion failed: {e}")
        
        return False
    
    def _try_pdfplumber_excel_export(self, pdf_path: str, excel_path: str) -> bool:
        """Fallback: Use pdfplumber + pandas Excel export"""
        try:
            import pdfplumber
            print("🔄 Trying pdfplumber + pandas conversion...")
            
            all_tables = []
            
            with pdfplumber.open(pdf_path) as pdf:
                for page_num in range(11, 26):  # Pages 11-25
                    page = pdf.pages[page_num-1]
                    tables = page.extract_tables()
                    
                    for table in tables:
                        if table and len(table) > 0:
                            # Convert to DataFrame
                            df = pd.DataFrame(table)
                            all_tables.append((f'Page_{page_num}_Table_{len(all_tables)}', df))
            
            if all_tables:
                # Export to Excel
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    for sheet_name, df in all_tables:
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                print(f"✅ PDFplumber conversion successful: {len(all_tables)} tables extracted")
                return True
                
        except Exception as e:
            print(f"❌ PDFplumber conversion failed: {e}")
        
        return False
    
    def _parse_excel_file(self, excel_path: str) -> Dict[str, int]:
        """Parse Excel file with 100% robust extraction"""
        
        print("📊 Parsing Excel file for BOERE data...")
        
        # Load Excel file
        workbook = openpyxl.load_workbook(excel_path)
        
        total_boere_count = 0
        total_te_bestellen = 0
        
        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            print(f"\n--- Processing {sheet_name} ---")
            
            # Find column positions dynamically
            columns = self._find_columns_in_sheet(sheet)
            
            if columns:
                # Extract data rows
                sheet_count, sheet_te_bestellen = self._extract_sheet_data(sheet, columns)
                total_boere_count += sheet_count
                total_te_bestellen += sheet_te_bestellen
                
                print(f"  {sheet_count} BOERE items, {sheet_te_bestellen} Te bestellen")
            else:
                print("  No valid columns found")
        
        workbook.close()
        
        print(f"\n🎯 Total BOERE count: {total_boere_count}")
        print(f"📝 Total Te bestellen excluded: {total_te_bestellen}")
        
        return {
            'boere': total_boere_count,
            'te_bestellen_excluded': total_te_bestellen
        }
    
    def _find_columns_in_sheet(self, sheet) -> Dict[str, int]:
        """Find column positions in Excel sheet"""
        
        columns = {}
        
        # Look for header row (first few rows) OR assume standard positions
        for row_num in range(1, min(6, sheet.max_row + 1)):
            row = sheet[row_num]
            
            for col_num, cell in enumerate(row, 1):
                if cell.value:
                    cell_text = str(cell.value).strip().upper()
                    
                    # Map header names to column numbers
                    if 'N°' in cell_text or cell_text == 'NO' or cell_text == 'N':
                        columns['item_number'] = col_num
                    elif 'ONDERDEEL' in cell_text:
                        columns['onderdeel'] = col_num
                    elif 'MATERIAAL' in cell_text:
                        columns['materiaal'] = col_num
                    elif 'PRO.METHODE' in cell_text or 'METHODE' in cell_text or 'PRODUCTIEMETHODE' in cell_text:
                        columns['pro_methode'] = col_num
        
        # If no columns found, try to infer from data patterns
        if not columns:
            columns = self._infer_columns_from_data(sheet)
        
        return columns
    
    def _infer_columns_from_data(self, sheet) -> Dict[str, int]:
        """Infer column positions from data patterns when headers aren't clear"""
        
        columns = {}
        
        # Analyze first few data rows to find patterns
        for row_num in range(1, min(10, sheet.max_row + 1)):
            row = sheet[row_num]
            
            for col_num, cell in enumerate(row, 1):
                if cell.value:
                    cell_text = str(cell.value).strip()
                    
                    # Look for N° column (digits only)
                    if cell_text.isdigit() and 'item_number' not in columns:
                        columns['item_number'] = col_num
                    
                    # Look for Pro.methode column (only check for "TE BESTELLEN")
                    elif 'TE BESTELLEN' in cell_text.upper():
                        if 'pro_methode' not in columns:
                            columns['pro_methode'] = col_num
        
        return columns
    
    def _extract_sheet_data(self, sheet, columns: Dict[str, int]) -> tuple:
        """Extract data from Excel sheet using column positions"""
        
        boere_count = 0
        te_bestellen_count = 0
        
        # Process all data rows
        for row_num in range(1, sheet.max_row + 1):
            row = sheet[row_num]
            
            # Get cell values
            item_number = ""
            pro_methode = ""
            
            if 'item_number' in columns:
                cell = row[columns['item_number'] - 1]  # Convert to 0-based
                if cell.value:
                    item_number = str(cell.value).strip()
            
            if 'pro_methode' in columns:
                cell = row[columns['pro_methode'] - 1]  # Convert to 0-based
                if cell.value:
                    pro_methode = str(cell.value).strip()
            
            # Count if valid BOERE item
            if item_number.isdigit():
                if 'TE BESTELLEN' in pro_methode.upper():
                    te_bestellen_count += 1
                else:
                    boere_count += 1
        
        return boere_count, te_bestellen_count
    
    def _cleanup_temp_files(self):
        """Clean up temporary files"""
        try:
            import shutil
            shutil.rmtree(self.temp_dir)
        except:
            pass

# Test the PDF to Excel parser
if __name__ == "__main__":
    parser = PDFToExcelParser()
    
    try:
        results = parser.parse_pdf_via_excel('S04479_RAPPORT_Rudi Matterne_0411_MO07202-7203_TV-wand (7-7).PDF')
        
        print("\n" + "="*50)
        print("🎯 FINAL RESULTS")
        print("="*50)
        print(f"BOERE count: {results['boere']}")
        print(f"Te bestellen excluded: {results['te_bestellen_excluded']}")
        print(f"Expected: 139")
        print(f"Accuracy: {results['boere']/139*100:.1f}%" if results['boere'] <= 139 else f"Over-count: +{results['boere']-139}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()