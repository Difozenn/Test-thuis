#!/usr/bin/env python3
"""
Precision PDF to Excel Converter
Uses manual text parsing with high precision + exact Excel template matching
Based on proven patterns from your R extraction success
"""

import pdfplumber
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import re
import os
from typing import Dict, List, Any, Tuple


class PrecisionPDFConverter:
    """High-precision PDF converter matching your successful R approach"""
    
    def convert(self, pdf_path: str, output_excel: str = None, template_excel: str = "1.xlsx"):
        """Convert PDF to Excel with precision extraction"""
        if not output_excel:
            output_excel = pdf_path.replace('.pdf', '_precision_converted.xlsx').replace('.PDF', '_precision_converted.xlsx')
        
        print(f"🎯 Precision converting {pdf_path}...")
        
        # Extract with high precision
        extracted_data = self._precision_extract(pdf_path)
        
        # Create Excel with exact template matching
        self._create_precision_excel(extracted_data, output_excel)
        
        print(f"✅ Precision conversion complete: {output_excel}")
        self._verify_conversion(output_excel, extracted_data)
        
        return output_excel
    
    def _precision_extract(self, pdf_path: str) -> Dict[str, Any]:
        """Extract PDF data with maximum precision"""
        print("📊 Extracting with precision parsing...")
        
        # Get all text
        all_text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text += text + "\n"
        
        lines = all_text.split('\n')
        lines = [line.strip() for line in lines if line.strip()]
        
        # Extract metadata with precision
        metadata = self._extract_metadata_precision(lines)
        
        # Extract sections with precision  
        sections = self._extract_sections_precision(lines)
        
        return {
            'metadata': metadata,
            'sections': sections
        }
    
    def _extract_metadata_precision(self, lines: List[str]) -> Dict[str, str]:
        """Extract metadata with high precision"""
        first_text = ' '.join(lines[:50])
        
        metadata = {
            'project_code': '',
            'sales_number': '',
            'project_name': '',
            'client': '',
            'designer': '',
            'department': ''
        }
        
        # Project code (MO pattern)
        mo_match = re.search(r'(MO\d+(?:-\d+)?)', first_text)
        if mo_match:
            metadata['project_code'] = mo_match.group(1)
        
        # Sales number (S pattern)
        s_match = re.search(r'(S\d+)', first_text)
        if s_match:
            metadata['sales_number'] = s_match.group(1)
            metadata['department'] = s_match.group(1)
        
        # Project name (complex pattern)
        project_patterns = [
            r'0411_MO\d+[-\d]*_([^"\n(]+?)(?:\s*\()',
            r'Project:\s*([^"\n]+?)(?:\s*(?:Nesting|Klant))',
            r'MO\d+[-\d]*[_\s]+([^"\n(]+?)(?:\s*\()'
        ]
        
        for pattern in project_patterns:
            match = re.search(pattern, first_text)
            if match:
                metadata['project_name'] = match.group(1).strip()
                break
        
        # Client (Klant pattern)
        client_patterns = [
            r'Klant:\s*([A-Za-z\s]+?)(?:\n|Tekenaar|JW|$)',
            r'Client:\s*([A-Za-z\s]+?)(?:\n|Tekenaar|JW|$)'
        ]
        
        for pattern in client_patterns:
            match = re.search(pattern, first_text)
            if match:
                metadata['client'] = match.group(1).strip()
                break
        
        # Designer
        designer_match = re.search(r'Tekenaar:\s*([A-Z]+)', first_text)
        if designer_match:
            metadata['designer'] = designer_match.group(1)
        elif 'JW' in first_text:
            metadata['designer'] = 'JW'
        
        return metadata
    
    def _extract_sections_precision(self, lines: List[str]) -> List[Dict[str, Any]]:
        """Extract sections with maximum precision"""
        sections = []
        current_section = None
        
        i = 0
        while i < len(lines):
            line = lines[i]
            line_lower = line.lower()
            
            # Detect section headers with precision
            section_type = self._detect_section_type_precision(line_lower)
            
            if section_type:
                # Save previous section
                if current_section and current_section['data']:
                    sections.append(current_section)
                
                # Start new section
                current_section = {
                    'type': section_type,
                    'headers': [],
                    'data': [],
                    'total_items': 0
                }
                print(f"  Found section: {section_type}")
            
            # Detect headers (N° pattern)
            if current_section and 'N°' in line and not current_section['headers']:
                current_section['headers'] = self._parse_headers_precision(line, current_section['type'])
                print(f"    Headers: {current_section['headers']}")
            
            # Parse data rows with precision
            if (current_section and current_section['headers'] and 
                re.match(r'^\s*\d+\s+', line)):
                
                # Get next line for potential "1mm" continuation
                next_line = lines[i + 1] if i + 1 < len(lines) else ""
                
                row_data = self._parse_row_precision(line, next_line, current_section['type'])
                if row_data:
                    current_section['data'].append(row_data)
            
            # Check for section end
            if 'Aantal onderdelen:' in line:
                if current_section:
                    total_match = re.search(r'Aantal onderdelen:\s*(\d+)', line)
                    if total_match:
                        current_section['total_items'] = int(total_match.group(1))
                    else:
                        current_section['total_items'] = len(current_section['data'])
                    
                    print(f"    Section complete: {current_section['total_items']} items")
                    sections.append(current_section)
                    current_section = None
            
            i += 1
        
        # Handle last section
        if current_section and current_section['data']:
            current_section['total_items'] = len(current_section['data'])
            sections.append(current_section)
        
        print(f"📋 Extracted {len(sections)} sections total")
        return sections
    
    def _detect_section_type_precision(self, line_lower: str) -> str:
        """Detect section type with high precision"""
        # Priority order matters
        patterns = [
            ('nesting', r'\bnesting\b'),
            ('opdeelzaag', r'\bopdeelzaag\b'),
            ('magazijn', r'\bmagazijn\b'),
            ('massief', r'\bmassief\b'),
            ('controle', r'\bcontrole\b'),
            ('te_bestellen', r'\bte\s+bestellen\b')
        ]
        
        for section_type, pattern in patterns:
            if re.search(pattern, line_lower):
                return section_type
        
        return None
    
    def _parse_headers_precision(self, line: str, section_type: str) -> List[str]:
        """Parse headers based on section type"""
        if section_type == 'magazijn':
            return ['N°', 'Beschrijving', 'Aantal stuks', 'GB nummer']
        elif section_type in ['nesting', 'opdeelzaag']:
            return ['N°', 'Onderdeel', 'Materiaal', 'Lengte', 'Breedte', 'Dikte', 
                   'L1', 'L2', 'B1', 'B2', 'ProductieM.', 'Opmerkingen']
        else:
            return ['N°', 'Onderdeel', 'Materiaal', 'Lengte', 'Breedte', 'Dikte', 'Opmerkingen']
    
    def _parse_row_precision(self, line: str, next_line: str, section_type: str) -> Dict[str, Any]:
        """Parse data row with maximum precision"""
        # Extract row number
        row_match = re.match(r'^\s*(\d+)\s+(.+)$', line)
        if not row_match:
            return None
        
        row_num = row_match.group(1)
        rest = row_match.group(2)
        
        # Combine with next line if it contains "1mm" and isn't a new row
        full_line = line
        if ('1mm' in next_line and not re.match(r'^\s*\d+\s+', next_line)):
            full_line = line + ' ' + next_line
        
        if section_type in ['nesting', 'opdeelzaag']:
            return self._parse_production_row_precision(row_num, rest, full_line, section_type)
        elif section_type == 'magazijn':
            return self._parse_magazijn_row_precision(row_num, rest)
        else:
            return {'N°': row_num, 'Data': rest}
    
    def _parse_production_row_precision(self, row_num: str, rest: str, full_line: str, section_type: str) -> Dict[str, Any]:
        """Parse production row (nesting/opdeelzaag) with precision"""
        parts = rest.split()
        idx = 0
        
        data = {'N°': row_num}
        
        # Onderdeel (part name)
        data['Onderdeel'] = parts[idx] if idx < len(parts) else ''
        idx += 1
        
        # Materiaal (collect until we hit a number)
        material_parts = []
        while idx < len(parts) and not re.match(r'^\d+\.?\d*$', parts[idx]):
            material_parts.append(parts[idx])
            idx += 1
        data['Materiaal'] = ' '.join(material_parts)
        
        # Dimensions (length, width, thickness)
        data['Lengte'] = parts[idx] if idx < len(parts) else ''
        idx += 1
        data['Breedte'] = parts[idx] if idx < len(parts) else ''
        idx += 1  
        data['Dikte'] = parts[idx] if idx < len(parts) else ''
        
        # Fineer data with precision (this is the critical part!)
        fineer_count = self._count_fineer_precision(full_line)
        
        data['L1'] = 'Fineer eik 1mm' if fineer_count >= 1 else ''
        data['L2'] = 'Fineer eik 1mm' if fineer_count >= 2 else ''
        data['B1'] = 'Fineer eik 1mm' if fineer_count >= 3 else ''
        data['B2'] = 'Fineer eik 1mm' if fineer_count >= 4 else ''
        
        # ProductieM
        productie_parts = []
        if 'Standaard' in full_line:
            productie_parts.append('Standaard')
        if 'Dik' in full_line:
            productie_parts.append('Dik')
        data['ProductieM.'] = ' '.join(productie_parts)
        
        # Opmerkingen (extract specific patterns)
        remarks = []
        
        # Overmaat patterns
        overmaat_patterns = [
            r'[LB]\+?=\d+mm overmaat',
            r'[A-Za-z\s]*\d+mm overmaat'
        ]
        for pattern in overmaat_patterns:
            matches = re.findall(pattern, full_line)
            remarks.extend(matches)
        
        # Frezen patterns
        frezen_patterns = [
            r'[A-Za-z\s]*frezen[A-Za-z\s,]*',
            r'Hoeken? recht frezen[^,\n]*'
        ]
        for pattern in frezen_patterns:
            matches = re.findall(pattern, full_line, re.IGNORECASE)
            remarks.extend(matches)
        
        data['Opmerkingen'] = ', '.join(remarks)
        
        return data
    
    def _count_fineer_precision(self, text: str) -> int:
        """Count Fineer occurrences with maximum precision"""
        # Multiple strategies for counting Fineer
        
        # Strategy 1: Count "Fineer eik" patterns
        fineer_eik_count = len(re.findall(r'Fineer\s*eik', text, re.IGNORECASE))
        
        # Strategy 2: Count "1mm" occurrences (they usually correspond to Fineer)
        mm_count = len(re.findall(r'1mm', text))
        
        # Strategy 3: Look for concatenated patterns like "FineereikaFineer eik"
        concatenated_count = len(re.findall(r'Fineer\s*eik.*?Fineer\s*eik', text, re.IGNORECASE))
        
        # Use the most reliable count
        # If we have explicit "Fineer eik" mentions, use that
        if fineer_eik_count > 0:
            return fineer_eik_count
        
        # Otherwise use 1mm count as proxy
        return mm_count
    
    def _parse_magazijn_row_precision(self, row_num: str, rest: str) -> Dict[str, Any]:
        """Parse magazijn row with precision"""
        # Try to split into: description, number, GB code
        # Pattern: "description ... number GB_CODE"
        
        # Look for the last number in the string (usually the quantity)
        parts = rest.split()
        number_idx = -1
        
        # Find rightmost pure number
        for i in range(len(parts) - 1, -1, -1):
            if re.match(r'^\d+$', parts[i]):
                number_idx = i
                break
        
        if number_idx > 0:
            description = ' '.join(parts[:number_idx])
            aantal = parts[number_idx]
            gb_nummer = ' '.join(parts[number_idx + 1:]) if number_idx + 1 < len(parts) else ''
        else:
            # Fallback: treat as description only
            description = rest
            aantal = ''
            gb_nummer = ''
        
        return {
            'N°': row_num,
            'Beschrijving': description,
            'Aantal stuks': aantal,
            'GB nummer': gb_nummer
        }
    
    def _create_precision_excel(self, extracted_data: Dict[str, Any], output_path: str):
        """Create Excel file with precision formatting"""
        print("📝 Creating precision Excel file...")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create cover sheet
        self._create_cover_sheet_precision(wb, extracted_data['metadata'])
        
        # Create data sheets
        sheet_num = 2
        for section in extracted_data['sections']:
            self._create_data_sheet_precision(wb, section, extracted_data['metadata'], sheet_num)
            sheet_num += 1
        
        wb.save(output_path)
    
    def _create_cover_sheet_precision(self, wb: openpyxl.Workbook, metadata: Dict[str, str]):
        """Create cover sheet with precision"""
        ws = wb.create_sheet("Table 1")
        
        # Exact content
        ws['A1'] = 'Project:\nKlant:\nTekenaar:'
        ws['D1'] = f"{metadata['project_code']}\n{metadata['sales_number']}\n{metadata['project_name']} {metadata['client']}\n{metadata['designer']}"
        ws['A2'] = metadata['department']
        ws['A3'] = 'info:\nSchuren'
        ws['A4'] = 'Totaal aantal onderdelen:'
        ws['A5'] = 'Afwerking: Lakstraat'
        ws['A6'] = 'Enkel als aangevinkt.                  Handwerk voor het schuren.\nKasten monteren! onderdelen sorteren per object Vlakstraat: gekleurde sjang gebruiken.'
        ws['A7'] = 'Datum:'
        ws['B7'] = 'kopie: terugbezorgen na schuren!'
        
        # Signature sections
        sections = ['Cel Holzer:', 'Accura:', 'Reichenbacher:', 'Kl Gannomat:', 'Cel Massief:', 'Cel schuren:']
        for i, section in enumerate(sections):
            ws[f'A{8+i}'] = f'{section}\nNaam:                          .../...'
        
        ws['C8'] = 'Opmerkingen:'
        
        # Merges
        merges = [
            'A1:C1', 'D1:F1', 'A2:B2', 'A3:B3', 'A4:B4', 'A5:F5',
            'A6:D6', 'B7:F7', 'A8:B8', 'C8:E13', 'A9:B9', 'A10:B10',
            'A11:B11', 'A12:B12', 'A13:B13', 'A14:F14'
        ]
        for merge in merges:
            ws.merge_cells(merge)
        
        # Column widths
        ws.column_dimensions['A'].width = 37.555556
        ws.column_dimensions['B'].width = 2
        ws.column_dimensions['C'].width = 11.555556
        ws.column_dimensions['D'].width = 31.777778
        ws.column_dimensions['E'].width = 42.222222
        ws.column_dimensions['F'].width = 2.888889
        
        # Text wrapping
        for cell in ['A1', 'D1', 'A3', 'A6', 'A8', 'A9', 'A10', 'A11', 'A12', 'A13']:
            ws[cell].alignment = Alignment(wrap_text=True, vertical='top')
    
    def _create_data_sheet_precision(self, wb: openpyxl.Workbook, section: Dict[str, Any], 
                                   metadata: Dict[str, str], sheet_num: int):
        """Create data sheet with precision"""
        ws = wb.create_sheet(f"Table {sheet_num}")
        
        # Header section
        ws['A1'] = metadata['department']
        ws['A2'] = 'Klant:'
        ws['G2'] = metadata['client']
        ws['O2'] = metadata['project_code']
        ws['A3'] = f"Tekenaar:  {metadata['designer']}\nSales nr:    {metadata['sales_number']}"
        ws['A4'] = 'Schuren'
        ws['E4'] = 'Project:'
        ws['H4'] = f"{metadata['project_name']}                                                       {section['type'].capitalize()}"
        
        # Column headers
        if section['type'] in ['nesting', 'opdeelzaag']:
            headers = [
                ('A5', 'N°'), ('B5', 'Onderdeel'), ('D5', 'Materiaal'), ('F5', 'Lengte'),
                ('I5', 'Breedte'), ('J5', 'Dikte'), ('K5', 'L1'), ('L5', 'L2'),
                ('M5', 'B1'), ('N5', 'B2')
            ]
            if section['type'] == 'nesting':
                headers.extend([('P5', 'ProductieM.'), ('Q5', 'Opmerkingen')])
            else:
                headers.append(('P5', 'Opmerkingen'))
        elif section['type'] == 'magazijn':
            headers = [
                ('A5', 'N°'), ('B5', 'Beschrijving'), ('I5', 'Aantal stuks'), ('P5', 'GB nummer')
            ]
        
        for cell, value in headers:
            ws[cell] = value
        
        # Data rows
        row = 6
        for item in section['data']:
            if section['type'] in ['nesting', 'opdeelzaag']:
                ws[f'A{row}'] = item.get('N°', '')
                ws[f'B{row}'] = item.get('Onderdeel', '')
                ws[f'D{row}'] = item.get('Materiaal', '')
                ws[f'F{row}'] = item.get('Lengte', '')
                ws[f'I{row}'] = item.get('Breedte', '')
                ws[f'J{row}'] = item.get('Dikte', '')
                ws[f'K{row}'] = item.get('L1', '')
                ws[f'L{row}'] = item.get('L2', '')
                ws[f'M{row}'] = item.get('B1', '')
                ws[f'N{row}'] = item.get('B2', '')
                if section['type'] == 'nesting':
                    ws[f'P{row}'] = item.get('ProductieM.', '')
                    ws[f'Q{row}'] = item.get('Opmerkingen', '')
                else:
                    ws[f'P{row}'] = item.get('Opmerkingen', '')
            elif section['type'] == 'magazijn':
                ws[f'A{row}'] = item.get('N°', '')
                ws[f'B{row}'] = item.get('Beschrijving', '')
                ws[f'I{row}'] = item.get('Aantal stuks', '')
                ws[f'P{row}'] = item.get('GB nummer', '')
            row += 1
        
        # Total row
        ws[f'A{row}'] = f'Aantal onderdelen: {section["total_items"]}'
        
        # Merges
        self._apply_data_sheet_merges_precision(ws, len(section['data']))
        
        # Column widths
        self._set_data_sheet_widths_precision(ws)
    
    def _apply_data_sheet_merges_precision(self, ws, data_row_count: int):
        """Apply merges with precision"""
        merges = [
            'A1:B1', 'A2:F2', 'G2:N2', 'O2:R2', 'A3:R3',
            'A4:D4', 'E4:G4', 'H4:R4',
            'B5:C5', 'D5:E5', 'F5:H5', 'N5:O5'
        ]
        
        for i in range(6, 6 + data_row_count):
            merges.extend([f'B{i}:C{i}', f'D{i}:E{i}', f'F{i}:H{i}', f'N{i}:O{i}'])
        
        total_row = 6 + data_row_count
        merges.append(f'A{total_row}:R{total_row}')
        
        for merge in merges:
            try:
                ws.merge_cells(merge)
            except:
                pass
    
    def _set_data_sheet_widths_precision(self, ws):
        """Set column widths with precision"""
        widths = {
            'A': 5.111111, 'B': 9.333333, 'C': 14.444444, 'D': 26.444444,
            'E': 2.444444, 'F': 7.333333, 'G': 0.666667, 'H': 5.111111,
            'I': 11.777778, 'J': 7.777778, 'K': 6.444444, 'L': 6.444444,
            'M': 6.666667, 'N': 3.111111, 'O': 3.111111, 'P': 7.777778,
            'Q': 57.555556, 'R': 3.111111
        }
        
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def _verify_conversion(self, output_path: str, extracted_data: Dict[str, Any]):
        """Verify the conversion quality"""
        print("\n🔍 Verifying conversion...")
        
        wb = openpyxl.load_workbook(output_path)
        print(f"✅ Created {len(wb.sheetnames)} sheets: {wb.sheetnames}")
        
        # Check sections
        for i, section in enumerate(extracted_data['sections']):
            sheet_name = f"Table {i + 2}"
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print(f"  {sheet_name} ({section['type']}): {section['total_items']} items")
                
                # Check L1/L2/B1/B2 if applicable
                if section['type'] in ['nesting', 'opdeelzaag'] and sheet['K5'].value == 'L1':
                    fineer_count = 0
                    for row in range(6, 16):
                        cells = [sheet[f'{col}{row}'].value for col in ['K', 'L', 'M', 'N']]
                        fineer_count += sum(1 for cell in cells if cell and 'Fineer' in str(cell))
                    
                    if fineer_count > 0:
                        print(f"    🎯 Fineer data: {fineer_count} sides detected")


def main():
    """Test the precision converter"""
    converter = PrecisionPDFConverter()
    
    # Test files
    test_files = [
        "S04479_RAPPORT_Rudi Matterne_0411_MO07199_Hoekdressing - opklapbed (4-7).PDF",
        "S04479_RAPPORT_Rudi Matterne_0411_MO07202-7203_TV-wand (7-7).PDF"
    ]
    
    for pdf_file in test_files:
        if os.path.exists(pdf_file):
            print(f"\n{'='*60}")
            print(f"🎯 PRECISION CONVERTING: {pdf_file}")
            print('='*60)
            
            output_file = pdf_file.replace('.PDF', '_precision_converted.xlsx')
            converter.convert(pdf_file, output_file, "1.xlsx")


if __name__ == "__main__":
    main()