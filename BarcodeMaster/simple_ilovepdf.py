#!/usr/bin/env python3
"""
Simple ILovePDF automation - replicate what actually works
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import time
import os
from pathlib import Path

def convert_pdf_with_ilovepdf(pdf_path: str) -> str:
    """Convert PDF using ILovePDF website automation"""
    
    print(f"🔄 Converting {pdf_path} using ILovePDF automation...")
    
    # Setup Chrome options
    chrome_options = Options()
    # chrome_options.add_argument("--headless")  # Comment out to see browser
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    # Setup download directory
    download_dir = os.path.abspath("ilovepdf_downloads")
    os.makedirs(download_dir, exist_ok=True)
    
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    chrome_options.add_experimental_option("prefs", prefs)
    
    try:
        # Setup Chrome driver
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        
        # Execute script to avoid detection
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        try:
            print("🌐 Opening ILovePDF...")
            driver.get("https://www.ilovepdf.com/pdf_to_excel")
            
            # Wait for page to load
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
            )
            
            print("📁 Uploading file...")
            # Find and upload file
            file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
            file_input.send_keys(os.path.abspath(pdf_path))
            
            print("⏳ Waiting for upload to complete...")
            time.sleep(3)
            
            # Look for convert button or automatic processing
            try:
                # Sometimes there's a convert button
                convert_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "[data-action='convert'], .btn-convert, #processTask"))
                )
                print("🔄 Clicking convert button...")
                convert_btn.click()
            except:
                print("🔄 Conversion may be automatic...")
            
            print("⏳ Waiting for conversion to complete...")
            
            # Wait for download button to appear
            download_btn = WebDriverWait(driver, 120).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 
                    "[data-action='download'], .btn-download, #download, .download-btn, .downloadfile"))
            )
            
            print("⬇️ Downloading converted file...")
            download_btn.click()
            
            # Wait for download to complete
            print("⏳ Waiting for download to finish...")
            time.sleep(10)
            
            # Find the downloaded file
            downloaded_files = list(Path(download_dir).glob("*.xlsx"))
            if downloaded_files:
                # Get the most recent file
                latest_file = max(downloaded_files, key=os.path.getctime)
                
                # Move to main directory
                output_path = f"ilovepdf_converted_{int(time.time())}.xlsx"
                os.rename(str(latest_file), output_path)
                
                print(f"✅ Conversion successful! File saved as: {output_path}")
                return output_path
            else:
                print("❌ No Excel file found in downloads")
                return None
            
        except Exception as e:
            print(f"❌ Conversion process failed: {e}")
            
            # Try to get page source for debugging
            try:
                page_source = driver.page_source
                with open("debug_page.html", "w") as f:
                    f.write(page_source)
                print("🔍 Page source saved to debug_page.html for analysis")
            except:
                pass
            
            return None
            
        finally:
            driver.quit()
            
    except Exception as e:
        print(f"❌ Driver setup failed: {e}")
        return None

def test_ilovepdf_automation():
    """Test the ILovePDF automation"""
    
    pdf_file = 'S04479_RAPPORT_Rudi Matterne_0411_MO07202-7203_TV-wand (7-7).PDF'
    
    if not os.path.exists(pdf_file):
        print(f"❌ PDF file not found: {pdf_file}")
        return
    
    print("🚀 Starting ILovePDF automation test...")
    
    result = convert_pdf_with_ilovepdf(pdf_file)
    
    if result:
        print(f"\n🎉 SUCCESS! Converted file: {result}")
        print("🔍 Now we can parse this clean Excel file!")
        
        # Quick test of the converted file
        try:
            import openpyxl
            wb = openpyxl.load_workbook(result)
            print(f"📊 Excel file has {len(wb.sheetnames)} sheets: {wb.sheetnames[:5]}...")
            wb.close()
        except Exception as e:
            print(f"⚠️ Could not verify Excel file: {e}")
    else:
        print("\n❌ Automation failed")
        print("💡 Fallback options:")
        print("1. Run with browser visible (comment out --headless)")
        print("2. Manual conversion at https://www.ilovepdf.com/pdf_to_excel")
        print("3. Try different online converter")

if __name__ == "__main__":
    test_ilovepdf_automation()