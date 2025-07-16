#!/usr/bin/env python3
"""
Unified PDF Parser - Single approach for NESTING, BOERE, and ACCURA
PDF → Excel → Parse → Clean up → Return data for database
"""

import os
import tempfile
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional
import re
import tabula

class UnifiedPDFParser:
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp()
    
    def parse_pdf_for_all_users(self, pdf_path: str) -> Dict[str, any]:
        """
        Unified parsing for all three users: NESTING, BOERE, ACCURA
        Returns structured data ready for database insertion
        """
        
        print("🔄 Starting unified PDF parsing...")
        
        # Step 1: Convert PDF to Excel
        excel_path = self._convert_pdf_to_excel(pdf_path)
        if not excel_path:
            raise Exception("PDF to Excel conversion failed")
        
        # Step 2: Parse Excel for all users
        results = {
            'nesting': self._parse_nesting_data(excel_path),
            'boere': self._parse_boere_data(excel_path), 
            'accura': self._parse_accura_data(excel_path)
        }
        
        # Step 3: Clean up Excel file
        self._cleanup_temp_files()
        
        print("✅ Unified parsing complete")
        return results
    
    def _convert_pdf_to_excel(self, pdf_path: str) -> Optional[str]:
        """Convert PDF to Excel using tabula-py"""
        
        excel_path = os.path.join(self.temp_dir, "unified_extraction.xlsx")
        
        try:
            print("📄 Converting PDF to Excel...")
            
            # Extract all tables from relevant pages
            # Pages 2-7: Nesting/Opdeelzaag
            # Pages 11-25: Controle/Massief (BOERE)
            dfs = tabula.read_pdf(
                pdf_path,
                pages="2-7,11-25",  # All relevant pages
                multiple_tables=True,
                pandas_options={'header': None}
            )
            
            if dfs:
                # Save all tables to Excel
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    for i, df in enumerate(dfs):
                        if not df.empty:
                            df.to_excel(writer, sheet_name=f'Table_{i}', index=False, header=False)
                
                print(f"✅ Converted to Excel: {len(dfs)} tables extracted")
                return excel_path
            
        except Exception as e:
            print(f"❌ PDF to Excel conversion failed: {e}")
        
        return None
    
    def _parse_nesting_data(self, excel_path: str) -> Dict[str, int]:
        """Parse NESTING data from Excel"""
        
        print("🔍 Parsing NESTING data...")
        
        workbook = openpyxl.load_workbook(excel_path)
        nesting_count = 0
        opdeelzaag_count = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Check if this sheet contains Nesting or Opdeelzaag data
            sheet_text = self._get_sheet_text(sheet)
            
            if 'NESTING' in sheet_text.upper() or self._looks_like_nesting_table(sheet):
                count = self._count_data_rows(sheet)
                nesting_count += count
                print(f"  {sheet_name}: {count} Nesting items")
                
            elif 'OPDEELZAAG' in sheet_text.upper() or self._looks_like_opdeelzaag_table(sheet):
                count = self._count_data_rows(sheet)
                opdeelzaag_count += count
                print(f"  {sheet_name}: {count} Opdeelzaag items")
        
        workbook.close()
        
        total = nesting_count + opdeelzaag_count
        print(f"📊 NESTING total: {nesting_count} + {opdeelzaag_count} = {total} items")
        
        return {
            'nesting_count': nesting_count,
            'opdeelzaag_count': opdeelzaag_count,
            'total_count': total
        }
    
    def _parse_boere_data(self, excel_path: str) -> Dict[str, int]:
        """Parse BOERE data from Excel"""
        
        print("🔍 Parsing BOERE data...")
        
        workbook = openpyxl.load_workbook(excel_path)
        boere_count = 0
        te_bestellen_count = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Find N° and Pro.methode columns
            columns = self._find_boere_columns(sheet)
            
            if columns:
                sheet_boere, sheet_te_bestellen = self._count_boere_items(sheet, columns)
                boere_count += sheet_boere
                te_bestellen_count += sheet_te_bestellen
                
                if sheet_boere > 0 or sheet_te_bestellen > 0:
                    print(f"  {sheet_name}: {sheet_boere} BOERE, {sheet_te_bestellen} Te bestellen")
        
        workbook.close()
        
        print(f"📊 BOERE total: {boere_count} items (excluded {te_bestellen_count} Te bestellen)")
        
        return {
            'boere_count': boere_count,
            'te_bestellen_excluded': te_bestellen_count
        }
    
    def _parse_accura_data(self, excel_path: str) -> Dict[str, int]:
        """Parse ACCURA data from Excel"""
        
        print("🔍 Parsing ACCURA data...")
        
        workbook = openpyxl.load_workbook(excel_path)
        accura_items = 0
        total_sides = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Check if this sheet has L1/L2/B1/B2 columns (ACCURA data)
            if self._has_accura_columns(sheet):
                items, sides = self._count_accura_items(sheet)
                accura_items += items
                total_sides += sides
                
                if items > 0:
                    print(f"  {sheet_name}: {items} items, {sides} sides")
        
        workbook.close()
        
        print(f"📊 ACCURA total: {accura_items} items, {total_sides} sides")
        
        return {
            'accura_items': accura_items,
            'total_sides': total_sides
        }
    
    def _get_sheet_text(self, sheet) -> str:
        """Get all text from a sheet for analysis"""
        text_parts = []
        for row in sheet.iter_rows(max_row=10, values_only=True):
            for cell in row:
                if cell:
                    text_parts.append(str(cell))
        return ' '.join(text_parts)
    
    def _looks_like_nesting_table(self, sheet) -> bool:
        """Check if sheet contains Nesting table structure"""
        # Look for L1/L2/B1/B2 columns
        for row in sheet.iter_rows(max_row=3, values_only=True):
            row_text = ' '.join(str(cell) for cell in row if cell).upper()
            if all(col in row_text for col in ['L1', 'L2', 'B1', 'B2']):
                return True
        return False
    
    def _looks_like_opdeelzaag_table(self, sheet) -> bool:
        """Check if sheet contains Opdeelzaag table structure"""
        # Similar to nesting but may have different patterns
        return self._looks_like_nesting_table(sheet)
    
    def _count_data_rows(self, sheet) -> int:
        """Count data rows in a sheet"""
        count = 0
        for row_num in range(1, sheet.max_row + 1):
            row = sheet[row_num]
            
            # Check if row has meaningful content
            has_content = any(cell.value and str(cell.value).strip() for cell in row)
            
            # Skip header rows
            row_text = ' '.join(str(cell.value) for cell in row if cell.value).upper()
            is_header = any(header in row_text for header in ['ONDERDEEL', 'MATERIAAL', 'LENGTE', 'BREEDTE'])
            
            if has_content and not is_header:
                count += 1
        
        return count
    
    def _find_boere_columns(self, sheet) -> Dict[str, int]:
        """Find N° and Pro.methode columns for BOERE parsing"""
        columns = {}
        
        # Look in first few rows for column headers or data patterns
        for row_num in range(1, min(10, sheet.max_row + 1)):
            row = sheet[row_num]
            
            for col_num, cell in enumerate(row, 1):
                if cell.value:
                    cell_text = str(cell.value).strip()
                    
                    # Look for N° column (digits)
                    if cell_text.isdigit() and 'item_number' not in columns:
                        columns['item_number'] = col_num
                    
                    # Look for Pro.methode column
                    elif 'TE BESTELLEN' in cell_text.upper():
                        if 'pro_methode' not in columns:
                            columns['pro_methode'] = col_num
        
        return columns
    
    def _count_boere_items(self, sheet, columns: Dict[str, int]) -> tuple:
        """Count BOERE items in a sheet"""
        boere_count = 0
        te_bestellen_count = 0
        
        for row_num in range(1, sheet.max_row + 1):
            row = sheet[row_num]
            
            # Get item number
            item_number = ""
            if 'item_number' in columns:
                cell = row[columns['item_number'] - 1]
                if cell.value:
                    item_number = str(cell.value).strip()
            
            # Get pro.methode  
            pro_methode = ""
            if 'pro_methode' in columns:
                cell = row[columns['pro_methode'] - 1]
                if cell.value:
                    pro_methode = str(cell.value).strip()
            
            # Count if valid item
            if item_number.isdigit():
                if 'TE BESTELLEN' in pro_methode.upper():
                    te_bestellen_count += 1
                else:
                    boere_count += 1
        
        return boere_count, te_bestellen_count
    
    def _has_accura_columns(self, sheet) -> bool:
        """Check if sheet has L1/L2/B1/B2 columns for ACCURA"""
        sheet_text = self._get_sheet_text(sheet)
        return all(col in sheet_text.upper() for col in ['L1', 'L2', 'B1', 'B2'])
    
    def _count_accura_items(self, sheet) -> tuple:
        """Count ACCURA items with L1/L2/B1/B2 data"""
        items = 0
        total_sides = 0
        
        # Find L1/L2/B1/B2 column positions
        l_columns = {}
        for row_num in range(1, min(5, sheet.max_row + 1)):
            row = sheet[row_num]
            for col_num, cell in enumerate(row, 1):
                if cell.value:
                    cell_text = str(cell.value).strip().upper()
                    if cell_text in ['L1', 'L2', 'B1', 'B2']:
                        l_columns[cell_text] = col_num
        
        # Count items with L1/L2/B1/B2 data
        for row_num in range(1, sheet.max_row + 1):
            row = sheet[row_num]
            
            sides_in_row = 0
            for col_name, col_pos in l_columns.items():
                if col_pos <= len(row):
                    cell = row[col_pos - 1]
                    if cell.value and str(cell.value).strip():
                        content = str(cell.value).strip()
                        if len(content) > 0 and content.upper() not in ['', 'TE BESTELLEN', 'DUMMY']:
                            sides_in_row += 1
            
            if sides_in_row > 0:
                items += 1
                total_sides += sides_in_row
        
        return items, total_sides
    
    def _cleanup_temp_files(self):
        """Clean up temporary Excel files"""
        try:
            import shutil
            shutil.rmtree(self.temp_dir)
            print("🗑️ Cleaned up temporary files")
        except:
            pass

# Test the unified parser
if __name__ == "__main__":
    parser = UnifiedPDFParser()
    
    try:
        results = parser.parse_pdf_for_all_users('S04479_RAPPORT_Rudi Matterne_0411_MO07202-7203_TV-wand (7-7).PDF')
        
        print("\n" + "="*60)
        print("🎯 UNIFIED PARSING RESULTS")
        print("="*60)
        
        print(f"📊 NESTING: {results['nesting']['total_count']} items")
        print(f"   • Nesting: {results['nesting']['nesting_count']}")
        print(f"   • Opdeelzaag: {results['nesting']['opdeelzaag_count']}")
        
        print(f"📊 BOERE: {results['boere']['boere_count']} items")
        print(f"   • Excluded Te bestellen: {results['boere']['te_bestellen_excluded']}")
        
        print(f"📊 ACCURA: {results['accura']['accura_items']} items")
        print(f"   • Total sides: {results['accura']['total_sides']}")
        
        print("\n✅ All data ready for database insertion!")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()