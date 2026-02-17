#!/usr/bin/env python3
"""
Excel Parser - Parse the extracted Excel file for all three users
Simple and clean approach: Excel → Parse → Results
"""

import openpyxl
from typing import Dict

class ExcelParser:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path)
        
    def parse_for_all_users(self) -> Dict[str, any]:
        """Parse Excel file for NESTING, BOERE, and ACCURA"""
        
        print(f"📊 Parsing Excel file: {self.excel_path}")
        print(f"📄 Found {len(self.workbook.sheetnames)} sheets")
        
        # Parse each sheet and collect results
        nesting_items = 0
        boere_items = 0
        accura_items = 0
        te_bestellen = 0
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # Analyze this sheet
            sheet_data = self._analyze_sheet(sheet, sheet_name)
            
            nesting_items += sheet_data['nesting']
            boere_items += sheet_data['boere'] 
            accura_items += sheet_data['accura']
            te_bestellen += sheet_data['te_bestellen']
            
            if sheet_data['nesting'] > 0 or sheet_data['boere'] > 0 or sheet_data['accura'] > 0:
                print(f"  {sheet_name}: N={sheet_data['nesting']}, B={sheet_data['boere']}, A={sheet_data['accura']}, TE={sheet_data['te_bestellen']}")
        
        self.workbook.close()
        
        results = {
            'nesting': {'total_count': nesting_items},
            'boere': {'boere_count': boere_items, 'te_bestellen_excluded': te_bestellen},
            'accura': {'accura_items': accura_items}
        }
        
        print(f"\n🎯 FINAL RESULTS:")
        print(f"📊 NESTING: {nesting_items} items")
        print(f"📊 BOERE: {boere_items} items (excluded {te_bestellen} Te bestellen)")
        print(f"📊 ACCURA: {accura_items} items")
        
        return results
    
    def _analyze_sheet(self, sheet, sheet_name: str) -> Dict[str, int]:
        """Analyze a single sheet and count items for each user type"""
        
        nesting_count = 0
        boere_count = 0
        accura_count = 0
        te_bestellen_count = 0
        
        # Check if sheet has any content
        if sheet.max_row < 2:
            return {'nesting': 0, 'boere': 0, 'accura': 0, 'te_bestellen': 0}
        
        # Look for column headers to identify sheet type
        has_l1_l2_b1_b2 = self._has_accura_columns(sheet)
        has_pro_methode = self._has_pro_methode_column(sheet)
        
        # Count rows with data
        for row_num in range(1, sheet.max_row + 1):
            row = sheet[row_num]
            
            # Get row text for analysis
            row_text = ' '.join(str(cell.value or '') for cell in row)
            
            # Skip empty rows or header rows
            if not row_text.strip() or self._is_header_row(row_text):
                continue
            
            # Check for Te bestellen
            is_te_bestellen = 'Te bestellen' in row_text or 'TE BESTELLEN' in row_text.upper()
            
            # Check if row has meaningful content (item number or component name)
            has_item_number = any(str(cell.value or '').strip().isdigit() for cell in row[:3])
            has_component = any(len(str(cell.value or '').strip()) > 2 for cell in row[:5])
            
            if not has_item_number and not has_component:
                continue
            
            # Count for Te bestellen
            if is_te_bestellen:
                te_bestellen_count += 1
                continue
            
            # Determine user type based on sheet characteristics
            if has_l1_l2_b1_b2:
                # ACCURA (needs L1/L2/B1/B2 columns)
                if self._has_side_data(row):
                    accura_count += 1
                # Also count for NESTING (pages 2-7 have L1/L2/B1/B2)
                nesting_count += 1
            elif has_pro_methode:
                # BOERE (has Pro.methode column, pages 11-25)
                boere_count += 1
            else:
                # Default: could be NESTING or BOERE
                nesting_count += 1
        
        return {
            'nesting': nesting_count,
            'boere': boere_count,
            'accura': accura_count,
            'te_bestellen': te_bestellen_count
        }
    
    def _has_accura_columns(self, sheet) -> bool:
        """Check if sheet has L1/L2/B1/B2 columns"""
        # Check first few rows for these column headers
        for row_num in range(1, min(4, sheet.max_row + 1)):
            row = sheet[row_num]
            row_text = ' '.join(str(cell.value or '') for cell in row).upper()
            if all(col in row_text for col in ['L1', 'L2', 'B1', 'B2']):
                return True
        return False
    
    def _has_pro_methode_column(self, sheet) -> bool:
        """Check if sheet has Pro.methode column"""
        for row_num in range(1, min(4, sheet.max_row + 1)):
            row = sheet[row_num]
            row_text = ' '.join(str(cell.value or '') for cell in row).upper()
            if 'PRO.METHODE' in row_text or 'METHODE' in row_text:
                return True
        return False
    
    def _has_side_data(self, row) -> bool:
        """Check if row has meaningful L1/L2/B1/B2 data"""
        # Look for non-empty cells in positions that might be L1/L2/B1/B2
        side_data_count = 0
        for cell in row[6:10]:  # Approximate positions for L1/L2/B1/B2
            if cell.value and str(cell.value).strip():
                content = str(cell.value).strip()
                if len(content) > 0 and content.upper() not in ['', 'TE BESTELLEN', 'DUMMY']:
                    side_data_count += 1
        
        return side_data_count > 0
    
    def _is_header_row(self, row_text: str) -> bool:
        """Check if row is a header"""
        header_keywords = ['ONDERDEEL', 'MATERIAAL', 'LENGTE', 'BREEDTE', 'DIKTE', 'L1', 'L2', 'B1', 'B2', 'PRO.METHODE']
        return any(keyword in row_text.upper() for keyword in header_keywords)

# Test the Excel parser
if __name__ == "__main__":
    parser = ExcelParser('extracted_tables.xlsx')
    
    try:
        results = parser.parse_for_all_users()
        
        print("\n" + "="*50)
        print("🎯 EXCEL PARSER RESULTS")
        print("="*50)
        
        print(f"📊 NESTING: {results['nesting']['total_count']} items")
        print(f"📊 BOERE: {results['boere']['boere_count']} items")
        print(f"   • Excluded Te bestellen: {results['boere']['te_bestellen_excluded']}")
        print(f"📊 ACCURA: {results['accura']['accura_items']} items")
        
        # Simple validation
        nesting_ok = 90 <= results['nesting']['total_count'] <= 120
        boere_ok = 120 <= results['boere']['boere_count'] <= 160
        accura_ok = 60 <= results['accura']['accura_items'] <= 100
        
        print(f"\n🎯 VALIDATION:")
        print(f"   NESTING: {'✅' if nesting_ok else '❌'} {results['nesting']['total_count']} (expected ~102)")
        print(f"   BOERE: {'✅' if boere_ok else '❌'} {results['boere']['boere_count']} (expected ~144)")
        print(f"   ACCURA: {'✅' if accura_ok else '❌'} {results['accura']['accura_items']} (expected ~84)")
        
        if nesting_ok and boere_ok and accura_ok:
            print("\n🏆 EXCEL PARSING SUCCESSFUL!")
            print("🚀 Ready for database insertion!")
            
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()