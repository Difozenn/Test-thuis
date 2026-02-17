#!/usr/bin/env python3
"""
Parse the clean Excel format (1.xlsx) - this is what we should target
"""

import openpyxl
from typing import Dict, List, Optional

class CleanExcelParser:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        
    def parse_for_all_users(self) -> Dict[str, any]:
        """Parse the clean Excel format for all users"""
        
        print(f"📊 Parsing clean Excel: {self.excel_path}")
        
        workbook = openpyxl.load_workbook(self.excel_path)
        
        nesting_items = []
        boere_items = []
        accura_items = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Identify section type
            section_type = self._identify_section_type(sheet)
            
            if section_type:
                # Extract items from this sheet
                sheet_items = self._extract_sheet_items(sheet, section_type, sheet_name)
                
                print(f"  {sheet_name}: {section_type} - {len(sheet_items)} items")
                
                # Categorize items
                if section_type in ['Nesting', 'Opdeelzaag']:
                    nesting_items.extend(sheet_items)
                    # Also add to ACCURA if has L1/L2/B1/B2 data
                    for item in sheet_items:
                        if self._has_accura_data(item):
                            accura_items.append(item)
                            
                elif section_type in ['Controle', 'Massief']:
                    # Only add to BOERE if not "Te bestellen"
                    for item in sheet_items:
                        if not self._is_te_bestellen(item):
                            boere_items.append(item)
        
        workbook.close()
        
        # Generate results
        results = {
            'nesting': {
                'nesting_count': len([item for item in nesting_items if self._is_nesting_type(item)]),
                'opdeelzaag_count': len([item for item in nesting_items if self._is_opdeelzaag_type(item)]),
                'total_count': len(nesting_items)
            },
            'boere': {
                'boere_count': len(boere_items),
                'te_bestellen_excluded': 0  # Will count separately
            },
            'accura': {
                'accura_items': len(accura_items),
                'total_sides': sum(self._count_sides(item) for item in accura_items)
            }
        }
        
        print(f"\n🎯 CLEAN EXCEL RESULTS:")
        print(f"📊 NESTING: {results['nesting']['total_count']} items")
        print(f"   • Nesting: {results['nesting']['nesting_count']}")
        print(f"   • Opdeelzaag: {results['nesting']['opdeelzaag_count']}")
        print(f"📊 BOERE: {results['boere']['boere_count']} items")
        print(f"📊 ACCURA: {results['accura']['accura_items']} items")
        print(f"   • Total sides: {results['accura']['total_sides']}")
        
        return results
    
    def _identify_section_type(self, sheet) -> Optional[str]:
        """Identify what section type this sheet represents"""
        
        # Look for section identifiers in the sheet
        for row_num in range(1, min(10, sheet.max_row + 1)):
            for cell in sheet[row_num]:
                if cell.value:
                    text = str(cell.value).upper()
                    
                    if 'NESTING' in text:
                        return 'Nesting'
                    elif 'OPDEELZAAG' in text:
                        return 'Opdeelzaag'
                    elif 'CONTROLE' in text:
                        return 'Controle'
                    elif 'MASSIEF' in text:
                        return 'Massief'
                    elif 'MAGAZIJN' in text:
                        return 'Magazijn'
        
        return None
    
    def _extract_sheet_items(self, sheet, section_type: str, sheet_name: str) -> List[Dict]:
        """Extract items from a sheet"""
        
        items = []
        
        # Find the header row (contains N°, Onderdeel, Materiaal, etc.)
        header_row_num = None
        column_map = {}
        
        for row_num in range(1, min(15, sheet.max_row + 1)):
            row = sheet[row_num]
            row_text = ' '.join(str(cell.value or '') for cell in row).upper()
            
            if 'N°' in row_text and 'ONDERDEEL' in row_text:
                header_row_num = row_num
                column_map = self._map_columns_from_row(row)
                break
        
        if not header_row_num:
            return items
        
        print(f"    Found headers at row {header_row_num}: {list(column_map.keys())}")
        
        # Extract data rows
        for row_num in range(header_row_num + 1, sheet.max_row + 1):
            row = sheet[row_num]
            
            # Skip empty rows
            if not any(cell.value for cell in row):
                continue
            
            item = self._extract_item_from_row(row, column_map, section_type, sheet_name, row_num)
            if item:
                items.append(item)
        
        return items
    
    def _map_columns_from_row(self, row) -> Dict[str, int]:
        """Map column headers to indices"""
        
        column_map = {}
        
        for col_idx, cell in enumerate(row):
            if cell.value:
                header = str(cell.value).upper().strip()
                
                if 'N°' in header:
                    column_map['item_number'] = col_idx
                elif 'ONDERDEEL' in header:
                    column_map['onderdeel'] = col_idx
                elif 'MATERIAAL' in header:
                    column_map['materiaal'] = col_idx
                elif 'LENGTE' in header:
                    column_map['lengte'] = col_idx
                elif 'BREEDTE' in header:
                    column_map['breedte'] = col_idx
                elif 'DIKTE' in header:
                    column_map['dikte'] = col_idx
                elif header == 'L1':
                    column_map['l1'] = col_idx
                elif header == 'L2':
                    column_map['l2'] = col_idx
                elif header == 'B1':
                    column_map['b1'] = col_idx
                elif header == 'B2':
                    column_map['b2'] = col_idx
                elif 'PRO.METHODE' in header:
                    column_map['pro_methode'] = col_idx
                elif 'COMMENTAAR' in header:
                    column_map['commentaar'] = col_idx
        
        return column_map
    
    def _extract_item_from_row(self, row, column_map: Dict[str, int], 
                              section_type: str, sheet_name: str, row_num: int) -> Optional[Dict]:
        """Extract item data from a row"""
        
        def get_cell_value(field: str) -> str:
            col_idx = column_map.get(field)
            if col_idx is not None and col_idx < len(row) and row[col_idx].value:
                return str(row[col_idx].value).strip()
            return ""
        
        # Extract all fields
        item_number = get_cell_value('item_number')
        onderdeel = get_cell_value('onderdeel')
        materiaal = get_cell_value('materiaal')
        
        # Skip if no meaningful content
        if not item_number and not onderdeel:
            return None
        
        item = {
            'item_number': item_number,
            'onderdeel': onderdeel,
            'materiaal': materiaal,
            'lengte': get_cell_value('lengte'),
            'breedte': get_cell_value('breedte'),
            'dikte': get_cell_value('dikte'),
            'l1': get_cell_value('l1'),
            'l2': get_cell_value('l2'),
            'b1': get_cell_value('b1'),
            'b2': get_cell_value('b2'),
            'pro_methode': get_cell_value('pro_methode'),
            'commentaar': get_cell_value('commentaar'),
            'section_type': section_type,
            'sheet_name': sheet_name,
            'row_num': row_num
        }
        
        return item
    
    def _is_nesting_type(self, item: Dict) -> bool:
        """Check if item belongs to Nesting section"""
        return item['section_type'] == 'Nesting'
    
    def _is_opdeelzaag_type(self, item: Dict) -> bool:
        """Check if item belongs to Opdeelzaag section"""
        return item['section_type'] == 'Opdeelzaag'
    
    def _is_te_bestellen(self, item: Dict) -> bool:
        """Check if item is 'Te bestellen'"""
        pro_methode = item.get('pro_methode', '').upper()
        return 'TE BESTELLEN' in pro_methode
    
    def _has_accura_data(self, item: Dict) -> bool:
        """Check if item has L1/L2/B1/B2 data for ACCURA"""
        return any(item.get(field, '').strip() for field in ['l1', 'l2', 'b1', 'b2'])
    
    def _count_sides(self, item: Dict) -> int:
        """Count valid sides (L1/L2/B1/B2) for ACCURA"""
        count = 0
        for field in ['l1', 'l2', 'b1', 'b2']:
            value = item.get(field, '').strip()
            if value and len(value) > 0 and value.upper() not in ['', 'TE BESTELLEN', 'DUMMY']:
                count += 1
        return count

# Test the clean Excel parser
if __name__ == "__main__":
    parser = CleanExcelParser("1.xlsx")
    
    try:
        results = parser.parse_for_all_users()
        
        print("\n" + "="*60)
        print("🎯 TARGET FORMAT ANALYSIS")
        print("="*60)
        
        print(f"📊 NESTING: {results['nesting']['total_count']} items")
        print(f"   • Nesting: {results['nesting']['nesting_count']}")
        print(f"   • Opdeelzaag: {results['nesting']['opdeelzaag_count']}")
        
        print(f"📊 BOERE: {results['boere']['boere_count']} items")
        
        print(f"📊 ACCURA: {results['accura']['accura_items']} items")
        print(f"   • Total sides: {results['accura']['total_sides']}")
        
        print(f"\n🎯 This is the TARGET quality we need to achieve!")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()