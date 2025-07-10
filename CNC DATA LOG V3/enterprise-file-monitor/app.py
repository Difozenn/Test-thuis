# app.py - Main Flask Application with Automatic Database Migration
import os
import json
import threading
import socket
import sqlite3
import shutil
import zipfile
import csv
import tempfile
from io import StringIO, BytesIO
from datetime import datetime, timedelta, timezone
from flask import Flask, render_template, redirect, url_for, flash, request, jsonify, send_file, Blueprint
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
from flask_migrate import Migrate
import plotly.graph_objs as go
import plotly.utils
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re
from functools import wraps
from sqlalchemy import func, and_, or_, text
from sqlalchemy.exc import OperationalError
import pytz
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import atexit
# Import our translation module
from translations import get_translation, setup_translations, get_available_languages, format_date_localized

# Initialize Flask extensions
db = SQLAlchemy()
bcrypt = Bcrypt()
login_manager = LoginManager()
migrate = Migrate()

def check_and_migrate_database():
    """Check if database needs migration and perform it automatically"""
    db_path = 'file_monitor.db'
    
    if not os.path.exists(db_path):
        return True  # New database, will be created by SQLAlchemy
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        print("🔄 Checking database schema for required migrations...")
        
        # Migration 1: Check efficiency columns in weekly_work_hours (only if table exists)
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='weekly_work_hours'")
        wwh_table_exists = cursor.fetchone() is not None
        
        if wwh_table_exists:
            cursor.execute("PRAGMA table_info(weekly_work_hours)")
            wwh_columns = [column[1] for column in cursor.fetchall()]
            
            efficiency_high_exists = 'efficiency_high_threshold' in wwh_columns
            efficiency_medium_exists = 'efficiency_medium_threshold' in wwh_columns
            
            if not efficiency_high_exists or not efficiency_medium_exists:
                print("📝 Migration 1: Adding efficiency threshold columns...")
                
                if not efficiency_high_exists:
                    cursor.execute("""
                        ALTER TABLE weekly_work_hours 
                        ADD COLUMN efficiency_high_threshold REAL DEFAULT 5.0
                    """)
                    print("✅ Added efficiency_high_threshold column")
                
                if not efficiency_medium_exists:
                    cursor.execute("""
                        ALTER TABLE weekly_work_hours 
                        ADD COLUMN efficiency_medium_threshold REAL DEFAULT 2.0
                    """)
                    print("✅ Added efficiency_medium_threshold column")
                
                # Update existing records
                cursor.execute("""
                    UPDATE weekly_work_hours 
                    SET efficiency_high_threshold = COALESCE(efficiency_high_threshold, 5.0),
                        efficiency_medium_threshold = COALESCE(efficiency_medium_threshold, 2.0)
                """)
                
                conn.commit()
        else:
            print("📝 Migration 1: weekly_work_hours table doesn't exist yet - will be created by SQLAlchemy")
        
        # Migration 2: Check and create work calendar tables
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='work_calendar'")
        work_calendar_exists = cursor.fetchone() is not None
        
        if not work_calendar_exists:
            print("📝 Migration 2: Creating work calendar tables...")
            
            # Create work_calendar table
            cursor.execute("""
                CREATE TABLE work_calendar (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    date DATE NOT NULL,
                    work_hours REAL DEFAULT 8.0,
                    day_type VARCHAR(20) DEFAULT 'workday',
                    notes VARCHAR(255),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES user (id),
                    UNIQUE (user_id, date)
                )
            """)
            
            # Create index for performance
            cursor.execute("""
                CREATE INDEX idx_work_calendar_user_date ON work_calendar (user_id, date)
            """)
            
            print("✅ Created work_calendar table")
            
            # Create work_schedule_templates table
            cursor.execute("""
                CREATE TABLE work_schedule_templates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name VARCHAR(100) NOT NULL,
                    description VARCHAR(255),
                    monday_hours REAL DEFAULT 8.0,
                    tuesday_hours REAL DEFAULT 8.0,
                    wednesday_hours REAL DEFAULT 8.0,
                    thursday_hours REAL DEFAULT 8.0,
                    friday_hours REAL DEFAULT 8.0,
                    saturday_hours REAL DEFAULT 0.0,
                    sunday_hours REAL DEFAULT 0.0
                )
            """)
            
            print("✅ Created work_schedule_templates table")
            
            # Create holiday_templates table
            cursor.execute("""
                CREATE TABLE holiday_templates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name VARCHAR(100) NOT NULL,
                    country VARCHAR(2) DEFAULT 'BE',
                    date DATE NOT NULL,
                    recurring BOOLEAN DEFAULT 1,
                    description VARCHAR(255)
                )
            """)
            
            print("✅ Created holiday_templates table")
            
            # Insert default schedule template
            cursor.execute("""
                INSERT INTO work_schedule_templates 
                (name, description, monday_hours, tuesday_hours, wednesday_hours, 
                 thursday_hours, friday_hours, saturday_hours, sunday_hours)
                VALUES ('Standard 5-Day Week', 'Monday-Friday 8 hours', 8.0, 8.0, 8.0, 8.0, 8.0, 0.0, 0.0)
            """)
            
            print("✅ Added default work schedule template")
            
            conn.commit()
        
        # Migration 3: Check CNC analysis foreign key fix (only if table exists)
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='cnc_analysis'")
        cnc_table_exists = cursor.fetchone() is not None
        
        if cnc_table_exists:
            # Check if foreign key references correct table
            cursor.execute("PRAGMA foreign_key_list(cnc_analysis)")
            fk_info = cursor.fetchall()
            
            # Look for incorrect file_events reference
            file_events_fk = any(fk[2] == 'file_events' for fk in fk_info)
            
            if file_events_fk:
                print("📝 Migration 3: Fixing CNC analysis foreign key reference...")
                
                # Create temporary table with correct foreign key
                cursor.execute("""
                    CREATE TABLE cnc_analysis_temp (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        event_id INTEGER NOT NULL,
                        file_path TEXT NOT NULL,
                        cycle_time_seconds REAL,
                        machine_time_minutes REAL,
                        tool_changes INTEGER DEFAULT 0,
                        rapid_moves INTEGER DEFAULT 0,
                        feed_moves INTEGER DEFAULT 0,
                        spindle_commands INTEGER DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (event_id) REFERENCES event (id)
                    )
                """)
                
                # Copy data from old table
                cursor.execute("""
                    INSERT INTO cnc_analysis_temp 
                    SELECT * FROM cnc_analysis
                """)
                
                # Drop old table and rename new one
                cursor.execute("DROP TABLE cnc_analysis")
                cursor.execute("ALTER TABLE cnc_analysis_temp RENAME TO cnc_analysis")
                
                print("✅ Fixed CNC analysis foreign key reference")
                conn.commit()
        else:
            print("📝 Migration 3: cnc_analysis table doesn't exist yet - will be created by SQLAlchemy")
        
        print("✅ All database migrations completed successfully!")
        
        conn.close()
        return True
        
    except Exception as e:
        print(f"❌ Database migration failed: {e}")
        if 'conn' in locals():
            conn.close()
        return False

# Create Flask app with application factory pattern
def create_app():
    app = Flask(__name__)
    
    # Configuration
    app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here-change-in-production')
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///file_monitor.db'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['UPLOAD_FOLDER'] = 'reports'
    app.config['LANGUAGES'] = get_available_languages()
    
    # Static files configuration
    app.static_folder = 'static'
    app.static_url_path = '/static'
    
    # Initialize extensions with app
    db.init_app(app)
    bcrypt.init_app(app)
    login_manager.init_app(app)
    migrate.init_app(app, db)
    
    # Setup translations
    setup_translations(app)
    
    login_manager.login_view = 'auth.login'
    login_manager.login_message_category = 'info'
    
    return app

# Create app instance
app = create_app()

# Initialize Background Scheduler for automated backups
scheduler = BackgroundScheduler()
scheduler.start()

# Cleanup scheduler on app exit
atexit.register(lambda: scheduler.shutdown())

def perform_scheduled_backup():
    """Perform automated backup based on schedule"""
    with app.app_context():
        try:
            print("Starting scheduled backup...")
            
            # Get scheduled backup settings
            scheduled_backup = ScheduledBackupSettings.query.first()
            if not scheduled_backup or not scheduled_backup.enabled:
                print("Scheduled backup is disabled or not configured")
                return
            
            # Create backup directory
            backup_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            
            # Generate backup filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'scheduled_backup_{timestamp}.db'
            backup_path = os.path.join(backup_dir, filename)
            
            # Copy database file
            db_path = 'file_monitor.db'
            if os.path.exists(db_path):
                shutil.copy2(db_path, backup_path)
                
                # Get file size
                size_bytes = os.path.getsize(backup_path)
                size_mb = round(size_bytes / (1024 * 1024), 2)
                
                # Create backup record
                backup_record = DatabaseBackup(
                    filename=filename,
                    type='scheduled_full',
                    size_mb=size_mb,
                    note=f'Automated {scheduled_backup.frequency} backup',
                    created_by_id=1  # System user
                )
                db.session.add(backup_record)
                
                # Update last run time
                scheduled_backup.last_run = datetime.now(timezone.utc)
                scheduled_backup.next_run = calculate_next_backup_time(scheduled_backup)
                
                db.session.commit()
                
                # Clean up old backups based on retention policy
                cleanup_old_backups(scheduled_backup.retention_days)
                
                print(f"Scheduled backup completed: {filename} ({size_mb}MB)")
                
            else:
                print("Database file not found for backup")
                
        except Exception as e:
            print(f"Scheduled backup failed: {str(e)}")
            db.session.rollback()

def calculate_next_backup_time(scheduled_backup):
    """Calculate the next backup time based on frequency"""
    now = datetime.now(timezone.utc)
    time_parts = scheduled_backup.time.split(':')
    hour = int(time_parts[0])
    minute = int(time_parts[1]) if len(time_parts) > 1 else 0
    
    # Create next run time
    next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    
    # If time has passed today, move to next occurrence
    if next_run <= now:
        if scheduled_backup.frequency == 'daily':
            next_run = next_run + timedelta(days=1)
        elif scheduled_backup.frequency == 'weekly':
            next_run = next_run + timedelta(days=7)
        elif scheduled_backup.frequency == 'monthly':
            # Move to next month
            if next_run.month == 12:
                next_run = next_run.replace(year=next_run.year + 1, month=1)
            else:
                next_run = next_run.replace(month=next_run.month + 1)
    
    return next_run

def cleanup_old_backups(retention_days):
    """Clean up old scheduled backups based on retention policy"""
    try:
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=retention_days)
        
        # Find old scheduled backups
        old_backups = DatabaseBackup.query.filter(
            DatabaseBackup.type == 'scheduled_full',
            DatabaseBackup.created_at < cutoff_date
        ).all()
        
        backup_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'backups')
        
        for backup in old_backups:
            # Delete file
            backup_path = os.path.join(backup_dir, backup.filename)
            if os.path.exists(backup_path):
                os.remove(backup_path)
                print(f"Deleted old backup: {backup.filename}")
            
            # Delete database record
            db.session.delete(backup)
        
        db.session.commit()
        
    except Exception as e:
        print(f"Error cleaning up old backups: {str(e)}")
        db.session.rollback()

def update_backup_schedule_job():
    """Update the scheduled backup job based on current settings"""
    with app.app_context():
        try:
            # Remove existing job if it exists
            if scheduler.get_job('scheduled_backup'):
                scheduler.remove_job('scheduled_backup')
            
            # Get current settings
            scheduled_backup = ScheduledBackupSettings.query.first()
            if scheduled_backup and scheduled_backup.enabled:
                time_parts = scheduled_backup.time.split(':')
                hour = int(time_parts[0])
                minute = int(time_parts[1]) if len(time_parts) > 1 else 0
                
                # Create cron trigger based on frequency
                if scheduled_backup.frequency == 'daily':
                    trigger = CronTrigger(hour=hour, minute=minute)
                elif scheduled_backup.frequency == 'weekly':
                    trigger = CronTrigger(day_of_week=0, hour=hour, minute=minute)  # Monday
                elif scheduled_backup.frequency == 'monthly':
                    trigger = CronTrigger(day=1, hour=hour, minute=minute)  # First day of month
                else:
                    return
                
                # Add the job
                scheduler.add_job(
                    func=perform_scheduled_backup,
                    trigger=trigger,
                    id='scheduled_backup',
                    name='Scheduled Database Backup',
                    replace_existing=True
                )
                
                # Calculate and save next run time
                scheduled_backup.next_run = calculate_next_backup_time(scheduled_backup)
                db.session.commit()
                
                print(f"Scheduled backup job updated: {scheduled_backup.frequency} at {scheduled_backup.time}")
                
        except Exception as e:
            print(f"Error updating backup schedule: {str(e)}")

# Add timezone context processor
@app.context_processor
def inject_timezone_functions():
    """Context processor to make timezone functions available in templates"""
    def format_local_time(dt, format='%Y-%m-%d %H:%M:%S'):
        """Format datetime in local timezone"""
        if dt is None:
            return ''
        local_dt = utc_to_local(dt)
        return local_dt.strftime(format) if local_dt else ''
    
    def to_local_time(dt):
        """Convert UTC datetime to local time for display"""
        return utc_to_local(dt)
    
    return {
        'format_local_time': format_local_time,
        'to_local_time': to_local_time,
        'local_now': lambda: datetime.now(get_local_timezone())
    }

def get_local_timezone():
    """Get the configured timezone"""
    import pytz
    return pytz.timezone(app.config.get('TIMEZONE', 'Europe/Brussels'))

def utc_to_local(utc_dt):
    """Convert UTC datetime to local timezone"""
    import pytz
    
    if utc_dt is None:
        return None
    
    # Ensure the datetime is timezone-aware
    if utc_dt.tzinfo is None:
        utc_dt = utc_dt.replace(tzinfo=pytz.UTC)
    
    local_tz = get_local_timezone()
    return utc_dt.astimezone(local_tz)

def local_to_utc(local_dt):
    """Convert local datetime to UTC"""
    import pytz
    
    if local_dt is None:
        return None
    
    local_tz = get_local_timezone()
    
    # If datetime is naive, localize it
    if local_dt.tzinfo is None:
        local_dt = local_tz.localize(local_dt)
    
    return local_dt.astimezone(pytz.UTC)

# Add custom Jinja2 filters
@app.template_filter('basename')
def basename_filter(path):
    """Extract basename from file path"""
    return os.path.basename(path)

# Database Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128))
    role = db.Column(db.String(20), default='operator')
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    is_active = db.Column(db.Boolean, default=True)
    language = db.Column(db.String(2), default='en')
    
    events = db.relationship('Event', backref='user', lazy='dynamic')
    
    def set_password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
    
    def check_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)
    
    @property
    def is_authenticated(self):
        return True
    
    @property
    def is_anonymous(self):
        return False
    
    def get_id(self):
        return str(self.id)

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    keywords = db.Column(db.Text)
    file_patterns = db.Column(db.Text)
    color = db.Column(db.String(7), default='#007bff')
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    
    events = db.relationship('Event', backref='category', lazy='dynamic')
    
    def get_keywords(self):
        return json.loads(self.keywords) if self.keywords else []
    
    def get_patterns(self):
        return json.loads(self.file_patterns) if self.file_patterns else []

class Event(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    file_path = db.Column(db.String(500), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'))
    matched_keyword = db.Column(db.String(100))
    computer_name = db.Column(db.String(100))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    event_type = db.Column(db.String(20))
    file_size = db.Column(db.Integer)

class MonitoredPath(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    path = db.Column(db.String(500), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    is_directory = db.Column(db.Boolean, default=False)
    recursive = db.Column(db.Boolean, default=True)
    description = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    last_modified = db.Column(db.DateTime)
    file_size = db.Column(db.Integer)
    change_count = db.Column(db.Integer, default=0)
    last_change_detected = db.Column(db.DateTime)
    
    user = db.relationship('User', backref='monitored_paths')
    __table_args__ = (db.UniqueConstraint('path', 'user_id', name='_path_user_uc'),)
    
    def increment_change_count(self):
        self.change_count = (self.change_count or 0) + 1
        self.last_change_detected = datetime.now(timezone.utc)

class FileChangeHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    monitored_path_id = db.Column(db.Integer, db.ForeignKey('monitored_path.id'), nullable=False)
    timestamp = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    change_type = db.Column(db.String(20))
    old_size = db.Column(db.Integer)
    new_size = db.Column(db.Integer)
    old_modified = db.Column(db.DateTime)
    new_modified = db.Column(db.DateTime)
    
    monitored_path = db.relationship('MonitoredPath', backref='change_history')

class SystemSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.Text)
    
    @staticmethod
    def get_setting(key, default=None):
        setting = SystemSettings.query.filter_by(key=key).first()
        return setting.value if setting else default
    
    @staticmethod
    def set_setting(key, value):
        setting = SystemSettings.query.filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            setting = SystemSettings(key=key, value=value)
            db.session.add(setting)
        db.session.commit()

class WeeklyWorkHours(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True)
    monday_hours = db.Column(db.Float, default=8.0)
    tuesday_hours = db.Column(db.Float, default=8.0)
    wednesday_hours = db.Column(db.Float, default=8.0)
    thursday_hours = db.Column(db.Float, default=8.0)
    friday_hours = db.Column(db.Float, default=8.0)
    saturday_hours = db.Column(db.Float, default=0.0)
    sunday_hours = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    
    # Efficiency configuration - configurable thresholds for efficiency calculation per user
    efficiency_high_threshold = db.Column(db.Float, default=5.0)  # events per hour for "high" efficiency
    efficiency_medium_threshold = db.Column(db.Float, default=2.0)  # events per hour for "medium" efficiency
    
    user = db.relationship('User', backref=db.backref('work_hours', uselist=False))
    
    def get_hours_for_day(self, day_number):
        """Get hours for a specific day (0=Monday, 6=Sunday)"""
        days = [
            self.monday_hours, self.tuesday_hours, self.wednesday_hours,
            self.thursday_hours, self.friday_hours, self.saturday_hours, self.sunday_hours
        ]
        return days[day_number] if 0 <= day_number <= 6 else 0.0
    
    def get_total_weekly_hours(self):
        """Get total hours for the week"""
        return (self.monday_hours + self.tuesday_hours + self.wednesday_hours + 
                self.thursday_hours + self.friday_hours + self.saturday_hours + self.sunday_hours)
    
    def get_working_days(self):
        """Get number of working days (days with > 0 hours)"""
        hours = [self.monday_hours, self.tuesday_hours, self.wednesday_hours,
                self.thursday_hours, self.friday_hours, self.saturday_hours, self.sunday_hours]
        return sum(1 for h in hours if h > 0)
    
    def get_average_daily_hours(self):
        """Get average hours per working day"""
        working_days = self.get_working_days()
        return self.get_total_weekly_hours() / working_days if working_days > 0 else 0.0
    
    def calculate_efficiency(self, events_per_hour):
        """Calculate efficiency level based on events per hour and configurable thresholds"""
        # Use default values if attributes don't exist (backward compatibility)
        high_threshold = getattr(self, 'efficiency_high_threshold', 5.0) or 5.0
        medium_threshold = getattr(self, 'efficiency_medium_threshold', 2.0) or 2.0
        
        if events_per_hour >= high_threshold:
            return 'high'
        elif events_per_hour >= medium_threshold:
            return 'medium'
        else:
            return 'low'

class CNCAnalysis(db.Model):
    __tablename__ = 'cnc_analysis'
    
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)
    file_path = db.Column(db.Text, nullable=False)
    cycle_time_seconds = db.Column(db.Float)
    machine_time_minutes = db.Column(db.Float)
    tool_changes = db.Column(db.Integer, default=0)
    rapid_moves = db.Column(db.Integer, default=0)
    feed_moves = db.Column(db.Integer, default=0)
    spindle_commands = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    
    # Relationships
    event = db.relationship('Event', backref='cnc_analysis')

class Report(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(200), nullable=False)
    type = db.Column(db.String(50), nullable=False)
    format = db.Column(db.String(10), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    file_path = db.Column(db.String(500))
    
    user = db.relationship('User', backref='reports')

class DatabaseBackup(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(200), nullable=False)
    type = db.Column(db.String(50), nullable=False)  # full, data_only, structure_only
    size_mb = db.Column(db.Float)
    note = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    
    created_by = db.relationship('User', backref='database_backups')

class ScheduledBackupSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    enabled = db.Column(db.Boolean, default=False)
    frequency = db.Column(db.String(20), default='daily')  # daily, weekly, monthly
    time = db.Column(db.String(5), default='02:00')  # HH:MM format
    retention_days = db.Column(db.Integer, default=30)
    last_run = db.Column(db.DateTime)
    next_run = db.Column(db.DateTime)

# Work Calendar Models
class WorkCalendar(db.Model):
    """Yearly work calendar with daily granularity"""
    __tablename__ = 'work_calendar'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    work_hours = db.Column(db.Float, default=8.0)
    day_type = db.Column(db.String(20), default='workday')  # workday, holiday, vacation, sick, maintenance
    notes = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    
    user = db.relationship('User', backref='work_calendar')
    
    # Indexes for performance
    __table_args__ = (
        db.Index('idx_work_calendar_user_date', 'user_id', 'date'),
        db.UniqueConstraint('user_id', 'date', name='unique_user_date'),
    )

class WorkScheduleTemplate(db.Model):
    """Reusable work schedule templates"""
    __tablename__ = 'work_schedule_templates'
    
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(255))
    monday_hours = db.Column(db.Float, default=8.0)
    tuesday_hours = db.Column(db.Float, default=8.0)
    wednesday_hours = db.Column(db.Float, default=8.0)
    thursday_hours = db.Column(db.Float, default=8.0)
    friday_hours = db.Column(db.Float, default=8.0)
    saturday_hours = db.Column(db.Float, default=0.0)
    sunday_hours = db.Column(db.Float, default=0.0)
    
    def get_hours_for_day(self, weekday):
        """Get hours for a specific weekday (0=Monday, 6=Sunday)"""
        days = [
            self.monday_hours, self.tuesday_hours, self.wednesday_hours,
            self.thursday_hours, self.friday_hours, self.saturday_hours, self.sunday_hours
        ]
        return days[weekday] if 0 <= weekday < 7 else 0.0

class HolidayTemplate(db.Model):
    """Predefined holidays and templates"""
    __tablename__ = 'holiday_templates'
    
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    country = db.Column(db.String(2), default='BE')  # BE, NL, etc.
    date = db.Column(db.Date, nullable=False)
    recurring = db.Column(db.Boolean, default=True)
    description = db.Column(db.String(255))

# Work Calendar Helper Functions
def initialize_year_calendar(user_id, year=None):
    """Initialize calendar for entire year with default template"""
    from datetime import date
    if year is None:
        year = datetime.now().year
        
    # Get or create default schedule template
    template = WorkScheduleTemplate.query.filter_by(name='Standard 5-Day Week').first()
    if not template:
        template = WorkScheduleTemplate(
            name='Standard 5-Day Week',
            description='Monday-Friday 8 hours',
            monday_hours=8.0, tuesday_hours=8.0, wednesday_hours=8.0,
            thursday_hours=8.0, friday_hours=8.0,
            saturday_hours=0.0, sunday_hours=0.0
        )
        db.session.add(template)
        db.session.commit()
    
    # Generate all days for the year
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    current_date = start_date
    
    calendar_entries = []
    while current_date <= end_date:
        # Get hours based on day of week
        weekday = current_date.weekday()  # 0=Monday, 6=Sunday
        hours = template.get_hours_for_day(weekday)
        
        # Check if entry already exists
        existing = WorkCalendar.query.filter_by(
            user_id=user_id, 
            date=current_date
        ).first()
        
        if not existing:
            calendar_entry = WorkCalendar(
                user_id=user_id,
                date=current_date,
                work_hours=hours,
                day_type='workday' if hours > 0 else 'weekend'
            )
            calendar_entries.append(calendar_entry)
        
        current_date += timedelta(days=1)
    
    # Bulk insert for performance
    db.session.add_all(calendar_entries)
    
    # Apply holidays
    apply_holidays(user_id, year)
    
    db.session.commit()
    return len(calendar_entries)

def apply_holidays(user_id, year):
    """Apply national holidays to calendar"""
    holidays = get_belgian_holidays(year)
    
    for holiday in holidays:
        calendar_entry = WorkCalendar.query.filter_by(
            user_id=user_id,
            date=holiday['date']
        ).first()
        
        if calendar_entry:
            calendar_entry.work_hours = 0.0
            calendar_entry.day_type = 'holiday'
            calendar_entry.notes = holiday['name']

def get_belgian_holidays(year):
    """Calculate Belgian holidays for given year"""
    from datetime import date
    holidays = []
    
    # Fixed holidays
    fixed_holidays = [
        {'name': 'New Year\'s Day', 'date': date(year, 1, 1)},
        {'name': 'Labour Day', 'date': date(year, 5, 1)},
        {'name': 'Belgian National Day', 'date': date(year, 7, 21)},
        {'name': 'Assumption Day', 'date': date(year, 8, 15)},
        {'name': 'All Saints Day', 'date': date(year, 11, 1)},
        {'name': 'Armistice Day', 'date': date(year, 11, 11)},
        {'name': 'Christmas Day', 'date': date(year, 12, 25)},
    ]
    
    holidays.extend(fixed_holidays)
    
    # Easter-based holidays
    easter_date = calculate_easter(year)
    easter_holidays = [
        {'name': 'Easter Monday', 'date': easter_date + timedelta(days=1)},
        {'name': 'Ascension Day', 'date': easter_date + timedelta(days=39)},
        {'name': 'Whit Monday', 'date': easter_date + timedelta(days=50)},
    ]
    
    holidays.extend(easter_holidays)
    return holidays

def calculate_easter(year):
    """Calculate Easter date for given year"""
    from datetime import date
    # Simplified easter calculation (basic algorithm)
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    n = (h + l - 7 * m + 114) // 31
    p = (h + l - 7 * m + 114) % 31
    return date(year, n, p + 1)

def get_calendar_data(user_id, year=None):
    """Get calendar data for display"""
    from datetime import date
    if year is None:
        year = datetime.now().year
        
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    
    calendar_entries = WorkCalendar.query.filter(
        WorkCalendar.user_id == user_id,
        WorkCalendar.date >= start_date,
        WorkCalendar.date <= end_date
    ).order_by(WorkCalendar.date).all()
    
    # Convert to monthly structure for UI
    monthly_data = {}
    for month in range(1, 13):
        monthly_data[month] = {
            'name': date(year, month, 1).strftime('%B'),
            'days': []
        }
    
    for entry in calendar_entries:
        month = entry.date.month
        monthly_data[month]['days'].append({
            'date': entry.date,
            'day': entry.date.day,
            'weekday': entry.date.weekday(),
            'work_hours': entry.work_hours,
            'day_type': entry.day_type,
            'notes': entry.notes or ''
        })
    
    return monthly_data

def get_work_hours_for_date(user_id, target_date):
    """Get work hours for specific date"""
    calendar_entry = WorkCalendar.query.filter_by(
        user_id=user_id,
        date=target_date
    ).first()
    
    if calendar_entry:
        return calendar_entry.work_hours
    
    # Fallback to weekly pattern if calendar not set up
    work_hours = WeeklyWorkHours.query.filter_by(user_id=user_id).first()
    if work_hours:
        return work_hours.get_hours_for_day(target_date.weekday())
    
    return 8.0  # Default fallback

# Helper function to get user's work hours with error handling
def get_user_work_hours(user_id=None):
    """Get work hours for a user or create default if not exists"""
    if user_id is None:
        user_id = current_user.id if current_user.is_authenticated else None
    
    if user_id:
        try:
            work_hours = WeeklyWorkHours.query.filter_by(user_id=user_id).first()
            if not work_hours:
                # Create default work hours (Monday-Friday 8h, Weekend 0h)
                work_hours = WeeklyWorkHours(
                    user_id=user_id,
                    monday_hours=8.0,
                    tuesday_hours=8.0,
                    wednesday_hours=8.0,
                    thursday_hours=8.0,
                    friday_hours=8.0,
                    saturday_hours=0.0,
                    sunday_hours=0.0,
                    efficiency_high_threshold=5.0,
                    efficiency_medium_threshold=2.0
                )
                db.session.add(work_hours)
                db.session.commit()
            else:
                # Ensure efficiency thresholds exist (backward compatibility)
                if not hasattr(work_hours, 'efficiency_high_threshold') or work_hours.efficiency_high_threshold is None:
                    work_hours.efficiency_high_threshold = 5.0
                if not hasattr(work_hours, 'efficiency_medium_threshold') or work_hours.efficiency_medium_threshold is None:
                    work_hours.efficiency_medium_threshold = 2.0
                db.session.commit()
            
            return work_hours
        except OperationalError as e:
            if "no such column" in str(e):
                print("⚠️ Database schema outdated. Please run the migration script or restart the application.")
                # Return a basic work hours object with default efficiency calculation
                class BasicWorkHours:
                    def __init__(self):
                        self.monday_hours = 8.0
                        self.tuesday_hours = 8.0
                        self.wednesday_hours = 8.0
                        self.thursday_hours = 8.0
                        self.friday_hours = 8.0
                        self.saturday_hours = 0.0
                        self.sunday_hours = 0.0
                        self.efficiency_high_threshold = 5.0
                        self.efficiency_medium_threshold = 2.0
                    
                    def get_hours_for_day(self, day_number):
                        days = [self.monday_hours, self.tuesday_hours, self.wednesday_hours,
                               self.thursday_hours, self.friday_hours, self.saturday_hours, self.sunday_hours]
                        return days[day_number] if 0 <= day_number <= 6 else 0.0
                    
                    def get_total_weekly_hours(self):
                        return 40.0
                    
                    def get_working_days(self):
                        return 5
                    
                    def get_average_daily_hours(self):
                        return 8.0
                    
                    def calculate_efficiency(self, events_per_hour):
                        if events_per_hour >= 5.0:
                            return 'high'
                        elif events_per_hour >= 2.0:
                            return 'medium'
                        else:
                            return 'low'
                
                return BasicWorkHours()
            else:
                raise e
    return None

# Login Manager
@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

@login_manager.unauthorized_handler
def unauthorized():
    # Check if this is an API request
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Authentication required'}), 401
    # For regular web requests, redirect to login
    return redirect(url_for('auth.login'))

# Role-based access decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated_function

# Create Blueprints
auth_bp = Blueprint('auth', __name__)
main_bp = Blueprint('main', __name__)
api_bp = Blueprint('api', __name__, url_prefix='/api')

# API Routes - Essential endpoints for C# client
@api_bp.route('/monitor/status')
@login_required
def monitor_status():
    """Get monitor status for authenticated user"""
    return jsonify({
        'status': 'authenticated',
        'username': current_user.username,
        'role': current_user.role
    })

@api_bp.route('/paths')
@login_required
def get_paths():
    """Get all active monitored paths for the current user"""
    paths = MonitoredPath.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).all()
    
    return jsonify([{
        'id': p.id,
        'path': p.path,
        'is_directory': p.is_directory,
        'recursive': p.recursive,
        'description': p.description
    } for p in paths])

# Updated API endpoints for app.py
# Replace the existing /api/categories and /api/log_event endpoints with these:

@api_bp.route('/categories')
@login_required
def get_categories():
    """Get all categories with keywords and patterns"""
    categories = Category.query.all()
    
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'color': c.color,
        'keywords': c.get_keywords(),
        'file_patterns': c.get_patterns()
    } for c in categories])

@api_bp.route('/monitored_paths')
@login_required
def get_monitored_paths():
    """Get monitored paths for current user"""
    if current_user.role == 'admin':
        # Admin can see all paths
        paths = MonitoredPath.query.filter_by(is_active=True).order_by(MonitoredPath.path).all()
    else:
        # Regular users see only their own paths
        paths = MonitoredPath.query.filter_by(user_id=current_user.id, is_active=True).order_by(MonitoredPath.path).all()
    
    return jsonify([{
        'id': path.id,
        'path': path.path,
        'description': path.description or os.path.basename(path.path),
        'is_directory': path.is_directory
    } for path in paths])

@api_bp.route('/manual_entry', methods=['POST'])
@login_required
def api_manual_entry():
    """API endpoint for manual entry from C# app"""
    try:
        data = request.get_json()
        
        description = data.get('description', '').strip()
        category = data.get('category', '').strip()
        amount = data.get('amount', 1)
        path_id = data.get('path_id')
        
        # Find category by name
        category_obj = Category.query.filter_by(name=category).first()
        if not category_obj:
            return jsonify({'error': 'Category not found'}), 400
        
        # Validate amount
        if not (1 <= amount <= 100):
            return jsonify({'error': 'Amount must be between 1 and 100'}), 400
        
        # Get the monitored path if provided
        path_info = ""
        if path_id:
            monitored_path = MonitoredPath.query.get(path_id)
            if monitored_path and (monitored_path.user_id == current_user.id or current_user.role == 'admin'):
                path_info = monitored_path.path
            else:
                path_info = "Manual Entry"
        else:
            path_info = "Manual Entry"
        
        # Create events
        events_created = 0
        for i in range(amount):
            # Create a unique identifier for each entry if amount > 1
            if amount > 1:
                entry_description = f"{description} (Entry {i+1}/{amount})"
            else:
                entry_description = description
            
            # Format file path with optional monitored path
            if path_info != "Manual Entry":
                file_path = f"{path_info}: {entry_description}"
            else:
                file_path = f"Manual Entry: {entry_description}"
            
            event = Event(
                file_path=file_path,
                category_id=category_obj.id,
                computer_name=socket.gethostname(),
                user_id=current_user.id,
                event_type='manual'
            )
            db.session.add(event)
            events_created += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'events_created': events_created,
            'message': f'Successfully created {events_created} event(s)'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/log_event', methods=['POST'])
@login_required
def log_event():
    """Log a file change event from the monitoring client"""
    try:
        data = request.get_json()
        
        path_id = data.get('path_id')
        change_type = data.get('change_type')
        file_path = data.get('file_path')
        timestamp_str = data.get('timestamp_utc')
        new_size = data.get('new_size')
        computer_name = data.get('computer_name', socket.gethostname())
        
        # New fields from enhanced client
        category_id = data.get('category_id')
        matched_keyword = data.get('matched_keyword')
        
        # Validate required fields
        if not all([path_id, change_type, file_path]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Get the monitored path
        monitored_path = MonitoredPath.query.get(path_id)
        if not monitored_path or monitored_path.user_id != current_user.id:
            return jsonify({'error': 'Invalid path_id'}), 404
        
        # Parse timestamp
        if timestamp_str:
            timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
        else:
            timestamp = datetime.now(timezone.utc)
        
        # If category_id is provided by client, use it
        if category_id:
            category = Category.query.get(category_id)
        else:
            # Otherwise, do server-side matching (fallback for older clients)
            category = None
            
            # Get all categories
            categories = Category.query.all()
            
            # Check each category for a match
            for cat in categories:
                # Check file patterns
                patterns = cat.get_patterns()
                for pattern in patterns:
                    if re.match(pattern, file_path, re.IGNORECASE):
                        category = cat
                        if not matched_keyword:
                            matched_keyword = f"Pattern: {pattern}"
                        break
                
                if category:
                    break
                
                # Check keywords in filename
                keywords = cat.get_keywords()
                filename = os.path.basename(file_path).lower()
                for keyword in keywords:
                    if keyword.lower() in filename:
                        category = cat
                        if not matched_keyword:
                            matched_keyword = keyword
                        break
                
                if category:
                    break
        
        # If no category matched, use "Allerlei" or create it
        if not category:
            category = Category.query.filter_by(name='Allerlei').first()
            if not category:
                category = Category(
                    name='Allerlei',
                    keywords='[]',
                    file_patterns='[]',
                    color='#6c757d'
                )
                db.session.add(category)
                db.session.commit()
        
        # Create event
        event = Event(
            timestamp=timestamp,
            file_path=file_path,
            category_id=category.id if category else None,
            matched_keyword=matched_keyword,
            computer_name=computer_name,
            user_id=current_user.id,
            event_type=change_type,
            file_size=new_size
        )
        db.session.add(event)
        db.session.flush()  # Get the event ID
        
        # Handle CNC analysis if provided
        cnc_analysis_data = data.get('cnc_analysis')
        if cnc_analysis_data:
            try:
                # Map C# field names to database field names
                # MachineTime is the total cycle time (machine ops + movements) in minutes
                machine_time_minutes = cnc_analysis_data.get('MachineTime', 0.0)
                cycle_time_seconds = machine_time_minutes * 60  # Convert minutes to seconds
                
                # Debug output
                machine_time_seconds = machine_time_minutes * 60
                minutes = int(machine_time_seconds // 60)
                seconds = int(machine_time_seconds % 60)
                print(f"[DEBUG] CNC Analysis data received:")
                print(f"  TotalTime: {cnc_analysis_data.get('TotalTime', 0.0)} min")
                print(f"  MachineTime: {cnc_analysis_data.get('MachineTime', 0.0)} min")
                print(f"  Storing cycle_time_seconds: {cycle_time_seconds} sec")
                print(f"  Storing machine_time_minutes: {machine_time_minutes} min")
                print(f"  Converted to human time: {minutes}:{seconds:02d}")
                
                cnc_analysis = CNCAnalysis(
                    event_id=event.id,
                    file_path=cnc_analysis_data.get('Filename', file_path),
                    cycle_time_seconds=cycle_time_seconds,
                    machine_time_minutes=machine_time_minutes,
                    tool_changes=cnc_analysis_data.get('ToolChanges', 0),
                    rapid_moves=0,  # Not provided by C# analyzer yet
                    feed_moves=0,   # Not provided by C# analyzer yet
                    spindle_commands=0  # Not provided by C# analyzer yet
                )
                db.session.add(cnc_analysis)
            except Exception as e:
                print(f"Error storing CNC analysis: {e}")  # Log but don't fail the event
        
        # Update monitored path stats
        monitored_path.increment_change_count()
        if new_size is not None:
            monitored_path.file_size = new_size
        monitored_path.last_modified = timestamp
        
        # Create change history
        history = FileChangeHistory(
            monitored_path_id=monitored_path.id,
            timestamp=timestamp,
            change_type=change_type,
            new_size=new_size,
            new_modified=timestamp
        )
        db.session.add(history)
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'event_id': event.id,
            'category': category.name if category else 'Uncategorized'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/manual_entry', methods=['POST'])
@login_required
def manual_entry_api():
    """API endpoint for manual entry from C# client"""
    try:
        data = request.get_json()
        
        description = data.get('description', '').strip()
        category_name = data.get('category')
        amount = data.get('amount', 1)
        path_id = data.get('path_id')
        cnc_analysis_data = data.get('cnc_analysis')
        
        if not category_name:
            return jsonify({'error': 'Category is required'}), 400
        
        if not (1 <= amount <= 100):
            return jsonify({'error': 'Amount must be between 1 and 100'}), 400
        
        # Find category by name
        category = Category.query.filter_by(name=category_name).first()
        if not category:
            return jsonify({'error': f'Category "{category_name}" not found'}), 404
        
        # Create events
        events_created = 0
        for i in range(amount):
            # Create a unique identifier for each entry if amount > 1
            if amount > 1 and description:
                entry_description = f"{description} (Entry {i+1}/{amount})"
            else:
                entry_description = description if description else f"Manual Entry {i+1}"
            
            event = Event(
                file_path=f"Manual Entry: {entry_description}",
                category_id=category.id,
                matched_keyword=None,
                computer_name=socket.gethostname(),
                user_id=current_user.id,
                event_type='manual'
            )
            db.session.add(event)
            db.session.flush()  # Get the event ID
            
            # Handle CNC analysis if provided (only for first entry to avoid duplicates)
            if cnc_analysis_data and i == 0:
                try:
                    # Map manual entry CNC analysis data to new schema
                    cycle_time_seconds = cnc_analysis_data.get('TotalTime', 0.0) * 60  # Convert minutes to seconds
                    machine_time_minutes = cnc_analysis_data.get('MachineTime', 0.0)
                    
                    cnc_analysis = CNCAnalysis(
                        event_id=event.id,
                        file_path=cnc_analysis_data.get('Filename', ''),
                        cycle_time_seconds=cycle_time_seconds,
                        machine_time_minutes=machine_time_minutes,
                        tool_changes=cnc_analysis_data.get('ToolChanges', 0),
                        rapid_moves=0,  # Not provided by manual entry
                        feed_moves=0,   # Not provided by manual entry
                        spindle_commands=0  # Not provided by manual entry
                    )
                    db.session.add(cnc_analysis)
                except Exception as e:
                    print(f"Error storing CNC analysis: {e}")  # Log but don't fail the event
            
            events_created += 1
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'events_created': events_created,
            'message': f'{events_created} event(s) added successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Routes - Authentication Blueprint
@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password) and user.is_active:
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('main.dashboard'))
        else:
            flash(get_translation('invalid_credentials'), 'danger')
    
    return render_template('login.html')

@auth_bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash(get_translation('logout_success'), 'info')
    return redirect(url_for('auth.login'))

@auth_bp.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        
        if current_user.check_password(current_password):
            current_user.set_password(new_password)
            db.session.commit()
            flash(get_translation('password_changed'), 'success')
            return redirect(url_for('main.dashboard'))
        else:
            flash(get_translation('incorrect_password'), 'danger')
    
    return render_template('change_password.html')

@auth_bp.route('/language', methods=['POST'])
@login_required
def change_language():
    language = request.form.get('language')
    if language in app.config['LANGUAGES']:
        current_user.language = language
        db.session.commit()
        flash(get_translation('language_updated', language), 'success')
    
    return redirect(request.referrer or url_for('main.dashboard'))

# Routes - Main Blueprint
@main_bp.route('/')
@login_required
def dashboard():
    user_filter = request.args.get('user_id', type=int)
    
    # Get work hours for calculations (legacy support)
    work_hours = get_user_work_hours(current_user.id)
    
    # Get current week's work calendar data
    now = datetime.now()
    week_start = now - timedelta(days=now.weekday())
    work_calendar_summary = {
        'total_weekly_hours': 0,
        'working_days': 0,
        'average_daily_hours': 0,
        'today_hours': 0
    }
    
    # Calculate current week summary from work calendar
    for i in range(7):
        day_date = (week_start + timedelta(days=i)).date()
        day_hours = get_work_hours_for_date(current_user.id, day_date)
        
        if day_hours and day_hours > 0:
            work_calendar_summary['total_weekly_hours'] += day_hours
            work_calendar_summary['working_days'] += 1
            
            # If today, store today's hours
            if day_date == now.date():
                work_calendar_summary['today_hours'] = day_hours
    
    # Calculate average
    if work_calendar_summary['working_days'] > 0:
        work_calendar_summary['average_daily_hours'] = work_calendar_summary['total_weekly_hours'] / work_calendar_summary['working_days']
    
    # Get operators list (exclude admin users)
    operators = User.query.filter_by(is_active=True, role='operator').order_by(User.username).all()
    
    # Date range setup - Always show today's data
    local_tz = get_local_timezone()
    now = datetime.now(local_tz)
    now_utc = now.astimezone(pytz.UTC)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Always use today for dashboard
    start_date = today_start
    end_date = now
    date_range = 'today'  # Keep for template compatibility
    
    # Convert to UTC for database queries
    today_start_utc = today_start.astimezone(pytz.UTC)
    start_date_utc = start_date.astimezone(pytz.UTC)
    end_date_utc = end_date.astimezone(pytz.UTC)
    
    # Initialize query objects based on user role
    if current_user.role == 'admin':
        if user_filter:
            events_query = Event.query.filter_by(user_id=user_filter)
            paths_query = MonitoredPath.query.filter_by(user_id=user_filter, is_active=True)
            filtered_user = User.query.get(user_filter)
            # Get the filtered user's work hours for efficiency calculations
            work_hours_for_stats = get_user_work_hours(user_filter)
        else:
            events_query = Event.query
            paths_query = MonitoredPath.query.filter_by(is_active=True)
            filtered_user = None
            work_hours_for_stats = work_hours
    else:
        events_query = Event.query.filter_by(user_id=current_user.id)
        paths_query = MonitoredPath.query.filter_by(user_id=current_user.id, is_active=True)
        filtered_user = current_user
        user_filter = current_user.id
        work_hours_for_stats = work_hours
    
    # Apply date range filter
    date_filtered_events = events_query.filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    
    # KPI Metrics with weekly work hours
    today_events = events_query.filter(Event.timestamp >= today_start_utc).count()
    
    yesterday_start = today_start - timedelta(days=1)
    yesterday_end = today_start
    yesterday_start_utc = yesterday_start.astimezone(pytz.UTC)
    yesterday_end_utc = yesterday_end.astimezone(pytz.UTC)
    yesterday_events = events_query.filter(
        Event.timestamp >= yesterday_start_utc,
        Event.timestamp < yesterday_end_utc
    ).count()
    
    # Calculate hourly average using today's configured work hours
    # Try to get work hours from calendar first, fallback to weekly configuration
    try:
        calendar_today_hours = get_work_hours_for_date(
            user_filter if user_filter else current_user.id, 
            now.date()
        )
        today_work_hours = calendar_today_hours if calendar_today_hours is not None else (
            work_hours_for_stats.get_hours_for_day(now.weekday()) if work_hours_for_stats else 8.0
        )
    except Exception:
        # Fallback to weekly configuration if calendar system is not available
        today_work_hours = work_hours_for_stats.get_hours_for_day(now.weekday()) if work_hours_for_stats else 8.0
    
    if today_work_hours > 0:
        # Calculate how many work hours have passed today
        hours_passed = min(now.hour + now.minute/60, today_work_hours)
        work_hour_events = today_events  # All events today (could be filtered by work time)
        hourly_average = work_hour_events / max(hours_passed, 0.1)
    else:
        hourly_average = 0  # Non-working day
    
    # Active paths and categories
    active_paths = paths_query.count()
    total_files = paths_query.filter_by(is_directory=False).count()
    total_dirs = paths_query.filter_by(is_directory=True).count()
    
    active_categories = db.session.query(Category.id).join(Event).filter(
        Event.timestamp >= start_date_utc
    ).distinct().count()
    total_categories = Category.query.count()
    
    # Recent activity (last 50 events)
    recent_events = date_filtered_events.order_by(Event.timestamp.desc()).limit(50).all()
    
    # Category distribution for pie chart
    category_stats = db.session.query(
        Category.name,
        Category.color,
        func.count(Event.id).label('count')
    ).join(Event).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    
    if user_filter:
        category_stats = category_stats.filter(Event.user_id == user_filter)
    
    category_stats = category_stats.group_by(Category.id).order_by(func.count(Event.id).desc()).all()
    
    category_chart_data = {
        'labels': [stat.name for stat in category_stats],
        'values': [stat.count for stat in category_stats],
        'colors': [stat.color for stat in category_stats]
    } if category_stats else None
    
    # Machine Time vs Work Hours for the date range
    machine_time_stats = db.session.query(
        func.date(Event.timestamp).label('date'),
        func.sum(CNCAnalysis.machine_time_minutes).label('machine_time'),
        func.count(Event.id).label('event_count')
    ).join(
        CNCAnalysis, Event.id == CNCAnalysis.event_id
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc,
        CNCAnalysis.cycle_time_seconds.isnot(None)
    )
    
    if user_filter:
        machine_time_stats = machine_time_stats.filter(Event.user_id == user_filter)
    
    machine_time_stats = machine_time_stats.group_by(func.date(Event.timestamp)).order_by('date').all()
    
    # Fill in missing dates for machine time data
    machine_time_data = {}
    work_hours_data = {}
    current_date = start_date.date()
    
    while current_date <= end_date.date():
        date_key = current_date.strftime('%Y-%m-%d')
        machine_time_data[date_key] = 0
        
        # Calculate work hours for this day
        day_of_week = current_date.weekday()  # 0=Monday, 6=Sunday
        if work_hours:
            work_hours_for_day = work_hours.get_hours_for_day(day_of_week)
            work_hours_data[date_key] = work_hours_for_day
        else:
            work_hours_data[date_key] = 8  # Default 8 hours
        
        current_date += timedelta(days=1)
    
    # Handle both string and date objects from the database
    for stat in machine_time_stats:
        if isinstance(stat.date, str):
            date_key = stat.date
        else:
            date_key = stat.date.strftime('%Y-%m-%d')
        machine_time_data[date_key] = float(stat.machine_time or 0) / 60  # Convert minutes to hours
    
    machine_time_vs_work_hours_data = {
        'labels': list(machine_time_data.keys())[-7:],  # Last 7 days
        'machine_time': list(machine_time_data.values())[-7:],
        'work_hours': list(work_hours_data.values())[-7:]
    }
    
    # Weekly activity pattern using work hours with configurable efficiency
    weekly_activity = []
    week_start = start_date - timedelta(days=start_date.weekday())
    for i in range(7):  # 0=Monday, 6=Sunday
        day_start = week_start + timedelta(days=i)
        day_end = day_start + timedelta(days=1)
        day_start_utc = day_start.astimezone(pytz.UTC)
        day_end_utc = day_end.astimezone(pytz.UTC)
        day_events = events_query.filter(
            Event.timestamp >= day_start_utc,
            Event.timestamp < day_end_utc
        ).count()
        
        # Try to get work hours from calendar first, fallback to weekly configuration
        try:
            calendar_hours = get_work_hours_for_date(
                user_filter if user_filter else current_user.id, 
                day_start.date()
            )
            day_work_hours = calendar_hours if calendar_hours is not None else (
                work_hours_for_stats.get_hours_for_day(i) if work_hours_for_stats else 8.0
            )
        except Exception:
            # Fallback to weekly configuration if calendar system is not available
            day_work_hours = work_hours_for_stats.get_hours_for_day(i) if work_hours_for_stats else 8.0
        events_per_hour = day_events / max(day_work_hours, 0.1) if day_work_hours > 0 else 0
        
        # Calculate efficiency using configurable thresholds for the viewed user
        if work_hours_for_stats:
            efficiency = work_hours_for_stats.calculate_efficiency(events_per_hour)
        else:
            # Default efficiency calculation if no work hours configured
            if events_per_hour >= 5.0:
                efficiency = 'high'
            elif events_per_hour >= 2.0:
                efficiency = 'medium'
            else:
                efficiency = 'low'
        
        weekly_activity.append({
            'day': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'][i],
            'events': day_events,
            'work_hours': day_work_hours,
            'normalized': round(events_per_hour, 2),
            'efficiency': efficiency
        })
    
    # Path distribution - Fixed to properly filter paths by owner
    path_stats_query = db.session.query(
        MonitoredPath.id,
        MonitoredPath.path,
        MonitoredPath.description,
        MonitoredPath.user_id,
        func.count(Event.id).label('count')
    ).join(
        Event,
        and_(
            or_(
                and_(
                    MonitoredPath.is_directory == False,
                    Event.file_path == MonitoredPath.path
                ),
                and_(
                    MonitoredPath.is_directory == True,
                    Event.file_path.like(func.concat(MonitoredPath.path, '%'))
                ),
                # Match manual entries that start with the monitored path followed by ":"
                and_(
                    Event.event_type == 'manual',
                    Event.file_path.like(func.concat(MonitoredPath.path, ':%'))
                )
            ),
            Event.user_id == MonitoredPath.user_id  # Events must belong to path owner
        )
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc,
        MonitoredPath.is_active == True
    )
    
    # Apply strict user filtering on BOTH MonitoredPath and Event
    if current_user.role == 'admin' and user_filter:
        # Admin viewing specific user - show only that user's paths with their events
        path_stats_query = path_stats_query.filter(
            MonitoredPath.user_id == user_filter,
            Event.user_id == user_filter
        )
    elif current_user.role != 'admin':
        # Non-admin users can only see their own paths and events
        path_stats_query = path_stats_query.filter(
            MonitoredPath.user_id == current_user.id,
            Event.user_id == current_user.id
        )
    # Note: If admin without filter, show all paths with their respective user's events
    
    path_stats = path_stats_query.group_by(MonitoredPath.id).order_by(func.count(Event.id).desc()).limit(10).all()
    
    path_distribution_data = {
        'labels': [stat.description or os.path.basename(stat.path) for stat in path_stats],
        'values': [stat.count for stat in path_stats]
    } if path_stats else None
    
    # Hourly timeline for today
    # Get all events for today and process in Python to handle timezone conversion
    hourly_events = Event.query.filter(
        Event.timestamp >= today_start_utc
    )
    
    if user_filter:
        hourly_events = hourly_events.filter(Event.user_id == user_filter)
    
    hourly_events = hourly_events.all()
    
    # Convert timestamps to local time and count by hour
    hourly_data = {h: 0 for h in range(24)}
    for event in hourly_events:
        local_time = utc_to_local(event.timestamp)
        hour = local_time.hour
        hourly_data[hour] += 1
    
    hourly_timeline_data = {
        'labels': [f"{h}:00" for h in range(24)],
        'values': [hourly_data[h] for h in range(24)]
    }
    
    # Top categories today
    top_categories = db.session.query(
        Category.name,
        Category.color,
        func.count(Event.id).label('count')
    ).join(Event).filter(
        Event.timestamp >= today_start_utc
    )
    
    if user_filter:
        top_categories = top_categories.filter(Event.user_id == user_filter)
    
    top_categories = top_categories.group_by(Category.id).order_by(
        func.count(Event.id).desc()
    ).limit(5).all()
    
    # Top changed files
    top_changed_files = paths_query.filter_by(is_directory=False).filter(
        MonitoredPath.change_count > 0
    ).order_by(MonitoredPath.change_count.desc()).limit(5).all()
    
    # User activity today (for admin view)
    user_activity = []
    if current_user.role == 'admin' and not user_filter:
        user_activity = db.session.query(
            User.username,
            func.count(Event.id).label('event_count')
        ).join(Event).filter(
            Event.timestamp >= today_start_utc,
            User.role == 'operator'
        ).group_by(User.id).order_by(
            func.count(Event.id).desc()
        ).limit(5).all()
    
    return render_template('dashboard.html',
                         # Filters
                         operators=operators,
                         user_filter=user_filter,
                         filtered_user=filtered_user,
                         date_range=date_range,
                         # Work hours
                         work_hours=work_hours,
                         work_hours_for_stats=work_hours_for_stats,
                         # KPIs
                         today_events=today_events,
                         yesterday_events=yesterday_events,
                         hourly_average=hourly_average,
                         active_paths=active_paths,
                         total_files=total_files,
                         total_dirs=total_dirs,
                         active_categories=active_categories,
                         total_categories=total_categories,
                         # Recent activity
                         recent_events=recent_events,
                         # Chart data
                         category_chart_data=json.dumps(category_chart_data),
                         machine_time_vs_work_hours_data=json.dumps(machine_time_vs_work_hours_data),
                         path_distribution_data=json.dumps(path_distribution_data),
                         hourly_timeline_data=json.dumps(hourly_timeline_data),
                         weekly_activity=weekly_activity,
                         work_calendar_summary=work_calendar_summary,
                         # Top lists
                         top_categories=top_categories,
                         top_changed_files=top_changed_files,
                         user_activity=user_activity,
                         # Helper
                         datetime=datetime)

@main_bp.route('/events')
@login_required
def events():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    query = Event.query
    
    # Apply filters
    if current_user.role == 'admin':
        user_id = request.args.get('user_id', type=int)
        if user_id:
            query = query.filter_by(user_id=user_id)
    else:
        query = query.filter_by(user_id=current_user.id)
    
    category_id = request.args.get('category', type=int)
    if category_id:
        query = query.filter_by(category_id=category_id)
    
    date_from = request.args.get('date_from')
    if date_from:
        query = query.filter(Event.timestamp >= datetime.fromisoformat(date_from))
    
    date_to = request.args.get('date_to')
    if date_to:
        query = query.filter(Event.timestamp <= datetime.fromisoformat(date_to))
    
    events = query.order_by(Event.timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    categories = Category.query.all()
    users = User.query.all() if current_user.role == 'admin' else None
    
    return render_template('events.html', 
                         events=events, 
                         categories=categories,
                         users=users)

@main_bp.route('/manual_entry', methods=['GET', 'POST'])
@login_required
def manual_entry():
    if request.method == 'POST':
        description = request.form.get('description', '').strip()
        category_id = request.form.get('category_id', type=int)
        matched_keyword = request.form.get('matched_keyword', '').strip()
        amount = request.form.get('amount', type=int)
        monitored_path_id = request.form.get('monitored_path_id', type=int)
        
        if description and category_id and amount and 1 <= amount <= 100:
            # Get the monitored path if selected
            path_info = ""
            if monitored_path_id:
                monitored_path = MonitoredPath.query.get(monitored_path_id)
                if monitored_path and (monitored_path.user_id == current_user.id or current_user.role == 'admin'):
                    path_info = monitored_path.path
                else:
                    path_info = "Manual Entry"
            else:
                path_info = "Manual Entry"
            
            # Create multiple events based on amount
            events_created = 0
            for i in range(amount):
                # Create a unique identifier for each entry if amount > 1
                if amount > 1:
                    entry_description = f"{description} (Entry {i+1}/{amount})"
                else:
                    entry_description = description
                
                # Format file path with optional monitored path
                if path_info != "Manual Entry":
                    file_path = f"{path_info}: {entry_description}"
                else:
                    file_path = f"Manual Entry: {entry_description}"
                
                event = Event(
                    file_path=file_path,
                    category_id=category_id,
                    matched_keyword=matched_keyword if matched_keyword else None,
                    computer_name=socket.gethostname(),
                    user_id=current_user.id,
                    event_type='manual'
                )
                db.session.add(event)
                events_created += 1
            
            db.session.commit()
            
            if events_created == 1:
                flash(get_translation('event_added'), 'success')
            else:
                flash(f"{events_created} {get_translation('events_added')}", 'success')
            
            return redirect(url_for('main.events'))
        else:
            flash('Please provide valid description, category, and amount (1-100)', 'danger')
    
    # Get categories and monitored paths for the form
    categories = Category.query.all()
    
    # Get monitored paths based on user role
    if current_user.role == 'admin':
        monitored_paths = MonitoredPath.query.filter_by(is_active=True).order_by(MonitoredPath.path).all()
    else:
        monitored_paths = MonitoredPath.query.filter_by(user_id=current_user.id, is_active=True).order_by(MonitoredPath.path).all()
    
    return render_template('manual_entry.html', categories=categories, monitored_paths=monitored_paths)

@main_bp.route('/delete_event/<int:event_id>', methods=['POST'])
@login_required
def delete_event(event_id):
    """Delete an event (admin only)"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    try:
        event = Event.query.get_or_404(event_id)
        
        # Delete related CNC analysis records first
        CNCAnalysis.query.filter_by(event_id=event_id).delete()
        
        # Delete the event
        db.session.delete(event)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Event deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

def calculate_work_minutes_in_range(start_date, end_date, work_hours):
    """Calculate total work minutes in a date range"""
    total_minutes = 0
    current = start_date
    
    while current < end_date:
        # Check if it's a work day
        if current.weekday() in [0, 1, 2, 3, 4]:  # Monday to Friday
            # For each work day, add the configured hours
            if hasattr(work_hours, 'get_hours_for_day'):
                day_hours = work_hours.get_hours_for_day(current.weekday())
            else:
                day_hours = 8  # Default 8 hours
            total_minutes += day_hours * 60
        
        current = current + timedelta(days=1)
    
    return total_minutes

@main_bp.route('/statistics')
@login_required
def statistics():
    """Enterprise-grade statistics page with comprehensive CNC data analytics"""
    # Get filters from request
    user_filter = request.args.get('user_id', type=int)
    date_range = request.args.get('range', 'week')
    custom_start = request.args.get('start_date')
    custom_end = request.args.get('end_date')
    
    # Get operators list for filter
    operators = User.query.filter_by(is_active=True, role='operator').order_by(User.username).all()
    
    # Calculate date range
    local_tz = get_local_timezone()
    now = datetime.now(local_tz)
    
    if date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').replace(tzinfo=local_tz)
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=local_tz)
    elif date_range == 'week':
        start_date = now - timedelta(days=7)
        end_date = now
    elif date_range == 'month':
        start_date = now - timedelta(days=30)
        end_date = now
    elif date_range == 'year':
        start_date = now - timedelta(days=365)
        end_date = now
    else:  # today
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    
    # Convert to UTC for queries
    start_date_utc = start_date.astimezone(pytz.UTC)
    end_date_utc = end_date.astimezone(pytz.UTC)
    
    # Build base query
    if current_user.role == 'admin':
        if user_filter:
            events_query = Event.query.filter_by(user_id=user_filter)
            users_query = User.query.filter_by(id=user_filter)
        else:
            events_query = Event.query
            users_query = User.query.filter_by(role='operator')
    else:
        events_query = Event.query.filter_by(user_id=current_user.id)
        users_query = User.query.filter_by(id=current_user.id)
        user_filter = current_user.id
    
    # Apply date filter
    events_query = events_query.filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    
    # 1. Production Overview Statistics
    total_events = events_query.count()
    unique_files = db.session.query(func.count(func.distinct(Event.file_path))).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    ).scalar()
    
    # 2. Machine Utilization (Events per hour of day)
    hourly_distribution = db.session.query(
        func.strftime('%H', Event.timestamp).label('hour'),
        func.count(Event.id).label('count')
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    if user_filter:
        hourly_distribution = hourly_distribution.filter(Event.user_id == user_filter)
    hourly_distribution = hourly_distribution.group_by('hour').all()
    
    # 3. Daily Production Trend
    daily_trend = db.session.query(
        func.date(Event.timestamp).label('date'),
        func.count(Event.id).label('count')
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    if user_filter:
        daily_trend = daily_trend.filter(Event.user_id == user_filter)
    daily_trend = daily_trend.group_by('date').order_by('date').all()
    
    # 4. Category Performance Analysis
    category_stats = db.session.query(
        Category.name,
        Category.color,
        func.count(Event.id).label('count'),
        func.count(func.distinct(func.date(Event.timestamp))).label('active_days')
    ).join(Event).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    if user_filter:
        category_stats = category_stats.filter(Event.user_id == user_filter)
    category_stats = category_stats.group_by(Category.id).order_by(func.count(Event.id).desc()).all()
    
    # 5. Machine Time vs Work Hours Analysis
    machine_time_analysis = db.session.query(
        func.date(Event.timestamp).label('date'),
        func.sum(CNCAnalysis.machine_time_minutes).label('machine_time'),
        func.count(Event.id).label('event_count')
    ).join(
        CNCAnalysis, Event.id == CNCAnalysis.event_id, isouter=True
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    if user_filter:
        machine_time_analysis = machine_time_analysis.filter(Event.user_id == user_filter)
    machine_time_analysis = machine_time_analysis.group_by('date').order_by('date').all()
    
    # 6. Events per Monitored File Analysis
    monitored_path_stats = db.session.query(
        MonitoredPath.description.label('path_description'),
        MonitoredPath.path.label('path'),
        func.count(Event.id).label('event_count'),
        func.count(func.distinct(func.date(Event.timestamp))).label('active_days'),
        func.sum(CNCAnalysis.machine_time_minutes).label('total_machine_time')
    ).join(
        Event, Event.file_path.like(func.concat(MonitoredPath.path, '%'))
    ).join(
        CNCAnalysis, Event.id == CNCAnalysis.event_id, isouter=True
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    if user_filter:
        monitored_path_stats = monitored_path_stats.filter(Event.user_id == user_filter)
    monitored_path_stats = monitored_path_stats.group_by(MonitoredPath.id).order_by(func.count(Event.id).desc()).all()
    
    # 7. Efficiency Metrics Calculation
    work_hours = get_user_work_hours(user_filter if user_filter else current_user.id)
    total_work_minutes = calculate_work_minutes_in_range(start_date, end_date, work_hours)
    events_per_hour = (total_events / (total_work_minutes / 60)) if total_work_minutes > 0 else 0
    
    # Calculate efficiency by day
    efficiency_by_day = []
    for day_data in daily_trend:
        day_date = datetime.strptime(str(day_data.date), '%Y-%m-%d').date()
        weekday = day_date.weekday()  # 0=Monday, 6=Sunday
        day_work_hours = work_hours.get_hours_for_day(weekday)
        
        if day_work_hours > 0:
            day_events_per_hour = day_data.count / day_work_hours
            efficiency_level = work_hours.calculate_efficiency(day_events_per_hour)
        else:
            day_events_per_hour = 0
            efficiency_level = 'off'
            
        efficiency_by_day.append({
            'date': day_data.date,
            'events': day_data.count,
            'work_hours': day_work_hours,
            'events_per_hour': round(day_events_per_hour, 2),
            'efficiency': efficiency_level
        })
    
    # 8. Time-based scaling for charts
    # Determine appropriate time grouping based on date range
    date_diff = (end_date - start_date).days
    if date_diff <= 7:
        time_grouping = 'hour'
        time_format = '%H:00'
    elif date_diff <= 31:
        time_grouping = 'day'  
        time_format = '%m/%d'
    elif date_diff <= 365:
        time_grouping = 'week'
        time_format = 'W%U'
    else:
        time_grouping = 'month'
        time_format = '%Y-%m'
    
    # 5. User Productivity Metrics
    user_stats = db.session.query(
        User.username,
        func.count(Event.id).label('total_events'),
        func.count(func.distinct(func.date(Event.timestamp))).label('active_days'),
        func.count(func.distinct(Event.file_path)).label('unique_files')
    ).join(Event).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc,
        User.role == 'operator'
    ).group_by(User.id).order_by(func.count(Event.id).desc()).all()
    
    # 6. File Change Frequency (Top changed files)
    file_frequency = db.session.query(
        Event.file_path,
        func.count(Event.id).label('change_count'),
        func.max(Event.timestamp).label('last_modified')
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    if user_filter:
        file_frequency = file_frequency.filter(Event.user_id == user_filter)
    file_frequency = file_frequency.group_by(Event.file_path)\
                                   .order_by(func.count(Event.id).desc())\
                                   .limit(20).all()
    
    # 7. Weekly Pattern Analysis
    weekly_pattern = db.session.query(
        func.strftime('%w', Event.timestamp).label('weekday'),
        func.count(Event.id).label('count')
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    if user_filter:
        weekly_pattern = weekly_pattern.filter(Event.user_id == user_filter)
    weekly_pattern = weekly_pattern.group_by('weekday').all()
    
    # 8. Event Type Distribution
    event_types = db.session.query(
        Event.event_type,
        func.count(Event.id).label('count')
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc
    )
    if user_filter:
        event_types = event_types.filter(Event.user_id == user_filter)
    event_types = event_types.group_by(Event.event_type).all()
    
    # 9. Path Activity Heatmap
    path_activity = db.session.query(
        MonitoredPath.path,
        MonitoredPath.description,
        func.count(Event.id).label('event_count')
    ).join(
        Event,
        or_(
            and_(
                MonitoredPath.is_directory == False,
                Event.file_path == MonitoredPath.path
            ),
            and_(
                MonitoredPath.is_directory == True,
                Event.file_path.like(func.concat(MonitoredPath.path, '%'))
            ),
            and_(
                Event.event_type == 'manual',
                Event.file_path.like(func.concat(MonitoredPath.path, ':%'))
            )
        )
    ).filter(
        Event.timestamp >= start_date_utc,
        Event.timestamp <= end_date_utc,
        MonitoredPath.is_active == True
    )
    if user_filter:
        path_activity = path_activity.filter(Event.user_id == user_filter)
    path_activity = path_activity.group_by(MonitoredPath.id)\
                                 .order_by(func.count(Event.id).desc())\
                                 .limit(15).all()
    
    # Calculate total machine time
    total_machine_time = sum(m.machine_time or 0 for m in machine_time_analysis)
    
    # Prepare data for charts with enhanced metrics
    statistics_data = {
        'overview': {
            'total_events': total_events,
            'unique_files': unique_files,
            'events_per_hour': round(events_per_hour, 2),
            'total_work_hours': round(total_work_minutes / 60, 1),
            'total_machine_time': round(total_machine_time / 60, 1),  # Convert to hours
            'efficiency_overall': work_hours.calculate_efficiency(events_per_hour) if events_per_hour > 0 else 'low'
        },
        'time_grouping': time_grouping,
        'date_range_days': date_diff,
        'hourly_distribution': {
            'labels': [f"{i:02d}:00" for i in range(24)],
            'data': [next((h.count for h in hourly_distribution if int(h.hour) == i), 0) for i in range(24)]
        },
        'daily_trend': {
            'labels': [d.date.strftime(time_format) for d in daily_trend],
            'data': [d.count for d in daily_trend],
            'machine_time': [next((m.machine_time or 0 for m in machine_time_analysis if m.date == d.date), 0) for d in daily_trend]
        },
        'category_performance': {
            'labels': [c.name for c in category_stats],
            'data': [c.count for c in category_stats],
            'colors': [c.color for c in category_stats],
            'active_days': [c.active_days for c in category_stats]
        },
        'machine_time_vs_work_hours': {
            'labels': [d['date'].strftime('%m/%d') for d in efficiency_by_day],
            'machine_time': [next((m.machine_time or 0 for m in machine_time_analysis if str(m.date) == str(d['date'])), 0) for d in efficiency_by_day],
            'work_hours': [d['work_hours'] for d in efficiency_by_day],
            'events_per_hour': [d['events_per_hour'] for d in efficiency_by_day],
            'efficiency_levels': [d['efficiency'] for d in efficiency_by_day]
        },
        'monitored_paths_stats': {
            'labels': [p.path_description or p.path[-30:] for p in monitored_path_stats],
            'event_counts': [p.event_count for p in monitored_path_stats],
            'machine_times': [p.total_machine_time or 0 for p in monitored_path_stats],
            'active_days': [p.active_days for p in monitored_path_stats]
        },
        'efficiency_metrics': {
            'daily_efficiency': efficiency_by_day,
            'avg_events_per_hour': round(events_per_hour, 2),
            'high_efficiency_days': len([d for d in efficiency_by_day if d['efficiency'] == 'high']),
            'medium_efficiency_days': len([d for d in efficiency_by_day if d['efficiency'] == 'medium']),
            'low_efficiency_days': len([d for d in efficiency_by_day if d['efficiency'] == 'low']),
            'efficiency_thresholds': {
                'high': work_hours.efficiency_high_threshold,
                'medium': work_hours.efficiency_medium_threshold
            }
        },
        'user_productivity': {
            'labels': [u.username for u in user_stats],
            'events': [u.total_events for u in user_stats],
            'active_days': [u.active_days for u in user_stats],
            'unique_files': [u.unique_files for u in user_stats]
        },
        'file_frequency': [
            {
                'path': os.path.basename(f.file_path),
                'full_path': f.file_path,
                'count': f.change_count,
                'last_modified': utc_to_local(f.last_modified).strftime('%Y-%m-%d %H:%M')
            } for f in file_frequency
        ],
        'weekly_pattern': {
            'labels': ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat'],
            'data': [next((w.count for w in weekly_pattern if w.weekday == str(i)), 0) for i in range(7)]
        },
        'event_types': {
            'labels': [e.event_type for e in event_types],
            'data': [e.count for e in event_types]
        },
        'path_activity': [
            {
                'path': p.description or p.path,
                'count': p.event_count
            } for p in path_activity
        ]
    }
    
    return render_template('statistics.html',
                         operators=operators,
                         user_filter=user_filter,
                         date_range=date_range,
                         start_date=start_date.strftime('%Y-%m-%d'),
                         end_date=end_date.strftime('%Y-%m-%d'),
                         statistics_data=statistics_data)

@main_bp.route('/reports/generate', methods=['POST'])
@login_required
def generate_report():
    report_type = request.form.get('report_type', 'detailed')
    export_format = request.form.get('export_format', 'excel')
    date_range = request.form.get('date_range', 'today')
    
    # Calculate date range
    today = datetime.now(timezone.utc).date()
    if date_range == 'custom':
        date_from = datetime.strptime(request.form.get('date_from'), '%Y-%m-%d').date()
        date_to = datetime.strptime(request.form.get('date_to'), '%Y-%m-%d').date()
    else:
        date_from, date_to = calculate_date_range(date_range)
    
    # User filter
    user_filter = request.form.get('user_filter')
    
    # Generate report based on type
    if report_type == 'dashboard':
        wb = generate_dashboard_report(date_from, date_to, user_filter, request.form)
    elif report_type == 'audit':
        wb = generate_audit_report(date_from, date_to, user_filter, request.form)
    elif report_type == 'summary':
        wb = generate_summary_report(date_from, date_to, user_filter)
    else:  # detailed
        wb = generate_detailed_report(date_from, date_to, user_filter, request.form)
    
    # Save report
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{report_type}_report_{timestamp}.xlsx"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    wb.save(filepath)
    
    # Save report record
    report = Report(
        filename=filename,
        type=report_type,
        format=export_format,
        user_id=current_user.id,
        file_path=filepath
    )
    db.session.add(report)
    db.session.commit()
    
    return send_file(filepath, as_attachment=True, download_name=filename)

@main_bp.route('/report/download/<int:id>')
@login_required
def download_report(id):
    report = Report.query.get_or_404(id)
    
    if current_user.role != 'admin' and report.user_id != current_user.id:
        flash(get_translation('no_permission'), 'danger')
        return redirect(url_for('main.reports'))
    
    if os.path.exists(report.file_path):
        return send_file(report.file_path, as_attachment=True, download_name=report.filename)
    else:
        flash(get_translation('report_not_found'), 'danger')
        return redirect(url_for('main.reports'))

@main_bp.route('/categories')
@login_required
@admin_required
def categories():
    categories = Category.query.all()
    return render_template('categories.html', categories=categories)

@main_bp.route('/category/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_category():
    if request.method == 'POST':
        name = request.form.get('name')
        color = request.form.get('color', '#007bff')
        keywords = request.form.getlist('keywords[]')
        patterns = request.form.getlist('patterns[]')
        
        # Filter out empty values
        keywords = [k.strip() for k in keywords if k.strip()]
        patterns = [p.strip() for p in patterns if p.strip()]
        
        category = Category(
            name=name,
            color=color,
            keywords=json.dumps(keywords),
            file_patterns=json.dumps(patterns)
        )
        
        db.session.add(category)
        db.session.commit()
        
        flash(get_translation('category_added'), 'success')
        return redirect(url_for('main.categories'))
    
    return render_template('category_form.html', category=None)

@main_bp.route('/category/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_category(id):
    category = Category.query.get_or_404(id)
    
    if request.method == 'POST':
        category.name = request.form.get('name')
        category.color = request.form.get('color', '#007bff')
        
        keywords = request.form.getlist('keywords[]')
        patterns = request.form.getlist('patterns[]')
        
        keywords = [k.strip() for k in keywords if k.strip()]
        patterns = [p.strip() for p in patterns if p.strip()]
        
        category.keywords = json.dumps(keywords)
        category.file_patterns = json.dumps(patterns)
        
        db.session.commit()
        flash(get_translation('category_updated'), 'success')
        return redirect(url_for('main.categories'))
    
    return render_template('category_form.html', category=category)

@main_bp.route('/category/delete/<int:id>')
@login_required
@admin_required
def delete_category(id):
    category = Category.query.get_or_404(id)
    db.session.delete(category)
    db.session.commit()
    flash(get_translation('category_deleted'), 'success')
    return redirect(url_for('main.categories'))

@main_bp.route('/users')
@login_required
@admin_required
def users():
    users = User.query.all()
    return render_template('users.html', users=users)

@main_bp.route('/user/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role', 'operator')
        
        user = User(
            username=username,
            email=email,
            role=role
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash(get_translation('user_added'), 'success')
        return redirect(url_for('main.users'))
    
    return render_template('user_form.html')

@main_bp.route('/user/toggle/<int:id>')
@login_required
@admin_required
def toggle_user(id):
    user = User.query.get_or_404(id)
    if user.id != current_user.id:  # Can't disable yourself
        user.is_active = not user.is_active
        db.session.commit()
    return redirect(url_for('main.users'))

@main_bp.route('/user/reset_password/<int:id>', methods=['POST'])
@login_required
@admin_required
def reset_user_password(id):
    user = User.query.get_or_404(id)
    new_password = request.form.get('new_password')
    
    if new_password:
        user.set_password(new_password)
        db.session.commit()
        flash(f"{get_translation('password_reset')} {user.username}", 'success')
    
    return redirect(url_for('main.users'))

@main_bp.route('/user/delete/<int:id>')
@login_required
@admin_required
def delete_user(id):
    user = User.query.get_or_404(id)
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        flash(get_translation('cannot_delete_self'), 'danger')
        return redirect(url_for('main.users'))
    
    username = user.username
    
    # Delete the user
    db.session.delete(user)
    db.session.commit()
    
    flash(f"{get_translation('user_deleted')} {username}", 'success')
    return redirect(url_for('main.users'))

@main_bp.route('/settings')
@login_required
def settings():
    if current_user.role == 'admin':
        paths = MonitoredPath.query.all()
        users = User.query.filter_by(is_active=True).all()
        # Get all operators with their work hours for efficiency configuration
        operators = User.query.filter_by(role='operator', is_active=True).all()
        # Ensure each operator has work hours
        for operator in operators:
            operator.work_hours = get_user_work_hours(operator.id)
    else:
        paths = MonitoredPath.query.filter_by(user_id=current_user.id).all()
        users = None
        operators = None
    
    # Get weekly work hours for the current user
    work_hours = get_user_work_hours(current_user.id)
    
    return render_template('settings.html', 
                         paths=paths, 
                         users=users,
                         operators=operators,
                         work_hours=work_hours)

@main_bp.route('/settings/work_calendar')
@login_required
def work_calendar():
    """Work calendar management page"""
    try:
        # Get year from query parameter, default to current year
        year = request.args.get('year', datetime.now().year, type=int)
        
        # Initialize calendar if it doesn't exist
        initialize_year_calendar(current_user.id, year)
        
        # Get calendar data for display
        calendar_data = get_calendar_data(current_user.id, year)
        
        # Calculate calendar statistics
        from datetime import date
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        
        # Get WorkCalendar entries
        calendar_entries = WorkCalendar.query.filter(
            WorkCalendar.user_id == current_user.id,
            WorkCalendar.date >= start_date,
            WorkCalendar.date <= end_date
        ).all()
        
        # Calculate statistics
        total_work_days = sum(1 for entry in calendar_entries if entry.work_hours > 0)
        total_work_hours = sum(entry.work_hours for entry in calendar_entries)
        holidays = sum(1 for entry in calendar_entries if entry.day_type == 'holiday')
        vacation_days = sum(1 for entry in calendar_entries if entry.day_type == 'vacation')
        avg_weekly_hours = total_work_hours / 52 if total_work_hours > 0 else 0
        
        calendar_stats = {
            'total_work_days': total_work_days,
            'total_work_hours': round(total_work_hours, 1),
            'holidays': holidays,
            'vacation_days': vacation_days,
            'avg_weekly_hours': round(avg_weekly_hours, 1)
        }
        
        return render_template('work_calendar.html',
                             calendar_data=calendar_data,
                             calendar_year=year,
                             calendar_stats=calendar_stats)
                             
    except Exception as e:
        flash(f'Error loading work calendar: {str(e)}', 'error')
        return redirect(url_for('main.settings'))

@main_bp.route('/settings/work_calendar/update_day', methods=['POST'])
@login_required
def update_calendar_day():
    """Update a single day in the work calendar"""
    try:
        data = request.get_json()
        
        from datetime import datetime
        
        # Parse the date
        date_str = data.get('date')
        work_hours = float(data.get('work_hours', 0))
        day_type = data.get('day_type', 'workday')
        notes = data.get('notes', '')
        
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Find existing entry or create new one
        calendar_entry = WorkCalendar.query.filter_by(
            user_id=current_user.id,
            date=target_date
        ).first()
        
        if not calendar_entry:
            calendar_entry = WorkCalendar(
                user_id=current_user.id,
                date=target_date
            )
            db.session.add(calendar_entry)
        
        # Update the entry
        calendar_entry.work_hours = work_hours
        calendar_entry.day_type = day_type
        calendar_entry.notes = notes
        
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/settings/work_calendar/apply_template', methods=['POST'])
@login_required
def apply_calendar_template():
    """Apply a work schedule template to the entire year"""
    try:
        data = request.get_json()
        template_name = data.get('template')
        year = data.get('year', datetime.now().year)
        
        # Apply the template
        if template_name == 'standard':
            # Standard 5-day week - reinitialize the calendar
            initialize_year_calendar(current_user.id, year)
            updated_count = 365  # Approximate
        else:
            return jsonify({'success': False, 'error': 'Unknown template'}), 400
        
        return jsonify({'success': True, 'updated': updated_count})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/settings/work_calendar/apply_holidays', methods=['POST'])
@login_required
def apply_calendar_holidays():
    """Apply national holidays to the work calendar"""
    try:
        data = request.get_json()
        year = data.get('year', datetime.now().year)
        
        # Apply holidays
        apply_holidays(current_user.id, year)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/settings/work_calendar/bulk_apply', methods=['POST'])
@login_required
def bulk_apply_work_hours():
    """Apply bulk work hours to calendar period"""
    try:
        data = request.get_json()
        template = data.get('template', {})
        from_month = data.get('from_month', 1)
        to_month = data.get('to_month', 12)
        year = data.get('year', datetime.now().year)
        preserve_holidays = data.get('preserve_holidays', True)
        preserve_vacations = data.get('preserve_vacations', True)
        
        # Validate template
        if not template or not isinstance(template, dict):
            return jsonify({'success': False, 'error': 'Invalid template'}), 400
        
        # Map weekdays to template
        weekday_map = {
            0: template.get('monday', 0),    # Monday
            1: template.get('tuesday', 0),   # Tuesday
            2: template.get('wednesday', 0), # Wednesday
            3: template.get('thursday', 0),  # Thursday
            4: template.get('friday', 0),    # Friday
            5: template.get('saturday', 0),  # Saturday
            6: template.get('sunday', 0)     # Sunday
        }
        
        # Get date range
        from datetime import date
        start_date = date(year, from_month, 1)
        if to_month == 12:
            end_date = date(year, 12, 31)
        else:
            # Get last day of to_month
            from calendar import monthrange
            end_date = date(year, to_month, monthrange(year, to_month)[1])
        
        # Apply template to each day in range
        days_updated = 0
        current_date = start_date
        
        while current_date <= end_date:
            # Get or create calendar entry
            calendar_entry = WorkCalendar.query.filter_by(
                user_id=current_user.id,
                date=current_date
            ).first()
            
            if not calendar_entry:
                calendar_entry = WorkCalendar(
                    user_id=current_user.id,
                    date=current_date
                )
                db.session.add(calendar_entry)
            
            # Check if we should preserve this entry
            should_preserve = False
            if preserve_holidays and calendar_entry.day_type == 'holiday':
                should_preserve = True
            elif preserve_vacations and calendar_entry.day_type == 'vacation':
                should_preserve = True
            
            if not should_preserve:
                # Apply template based on weekday
                weekday = current_date.weekday()
                hours = weekday_map.get(weekday, 0)
                
                calendar_entry.work_hours = hours
                calendar_entry.day_type = 'workday' if hours > 0 else 'weekend'
                calendar_entry.updated_at = datetime.now(timezone.utc)
                
                days_updated += 1
            
            # Move to next day
            current_date += timedelta(days=1)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'days_updated': days_updated,
            'message': f'Applied template to {days_updated} days'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@main_bp.route('/settings/work_hours', methods=['POST'])
@login_required
def update_work_hours():
    try:
        # Get or create work hours for the current user
        work_hours = get_user_work_hours(current_user.id)
        
        # Update hours for each day
        work_hours.monday_hours = float(request.form.get('monday_hours', 8.0))
        work_hours.tuesday_hours = float(request.form.get('tuesday_hours', 8.0))
        work_hours.wednesday_hours = float(request.form.get('wednesday_hours', 8.0))
        work_hours.thursday_hours = float(request.form.get('thursday_hours', 8.0))
        work_hours.friday_hours = float(request.form.get('friday_hours', 8.0))
        work_hours.saturday_hours = float(request.form.get('saturday_hours', 0.0))
        work_hours.sunday_hours = float(request.form.get('sunday_hours', 0.0))
        
        work_hours.updated_at = datetime.now(timezone.utc)
        
        # Validate hours (0-24 for each day)
        hours = [work_hours.monday_hours, work_hours.tuesday_hours, work_hours.wednesday_hours,
                work_hours.thursday_hours, work_hours.friday_hours, work_hours.saturday_hours, work_hours.sunday_hours]
        
        if any(h < 0 or h > 24 for h in hours):
            flash(get_translation('invalid_work_hours'), 'danger')
            return redirect(url_for('main.settings'))
        
        db.session.commit()
        flash(get_translation('work_hours_updated'), 'success')
        
    except ValueError:
        flash(get_translation('invalid_work_hours'), 'danger')
    
    return redirect(url_for('main.settings'))

@main_bp.route('/settings/operator_efficiency', methods=['POST'])
@login_required
@admin_required
def update_operator_efficiency():
    try:
        operator_id = request.form.get('operator_id', type=int)
        efficiency_high = request.form.get('efficiency_high_threshold', type=float)
        efficiency_medium = request.form.get('efficiency_medium_threshold', type=float)
        
        if not operator_id:
            flash('Invalid operator ID', 'danger')
            return redirect(url_for('main.settings'))
        
        # Verify the user is an operator
        operator = User.query.get(operator_id)
        if not operator or operator.role != 'operator':
            flash('User not found or not an operator', 'danger')
            return redirect(url_for('main.settings'))
        
        # Get or create work hours for the operator
        work_hours = get_user_work_hours(operator_id)
        
        # Update efficiency thresholds
        if efficiency_high:
            work_hours.efficiency_high_threshold = efficiency_high
        if efficiency_medium:
            work_hours.efficiency_medium_threshold = efficiency_medium
        
        # Validate efficiency thresholds
        if work_hours.efficiency_high_threshold <= work_hours.efficiency_medium_threshold:
            flash('High efficiency threshold must be greater than medium threshold', 'danger')
            return redirect(url_for('main.settings'))
        
        work_hours.updated_at = datetime.now(timezone.utc)
        db.session.commit()
        
        flash(f'Efficiency thresholds updated for {operator.username}', 'success')
        
    except ValueError:
        flash('Invalid efficiency values', 'danger')
    except Exception as e:
        flash(f'Error updating efficiency thresholds: {str(e)}', 'danger')
    
    return redirect(url_for('main.settings'))

@main_bp.route('/path/add', methods=['POST'])
@login_required
def add_monitored_path():
    try:
        path = request.form.get('path')
        path_type = request.form.get('path_type', 'file')
        description = request.form.get('description', '').strip()
        recursive = request.form.get('recursive') == 'on'
        
        # Check if this is an API request
        is_api_request = (request.content_type and 'application/json' in request.content_type) or \
                         request.headers.get('X-Client-Type') == 'FileMonitorTray'
        
        if not path:
            error_msg = 'Path is required'
            if is_api_request:
                return jsonify({'error': error_msg}), 400
            flash(error_msg, 'danger')
            return redirect(url_for('main.settings'))
        
        if current_user.role == 'admin':
            user_id = request.form.get('user_id', type=int) or current_user.id
        else:
            user_id = current_user.id
        
        # Skip path existence check if the request comes from the C# client
        is_client_request = request.headers.get('X-Client-Type') == 'FileMonitorTray' or \
                         (request.content_type and 'application/json' in request.content_type)
        
        if not is_client_request and not os.path.exists(path):
            error_msg = 'Path does not exist'
            if is_api_request:
                return jsonify({'error': error_msg}), 400
            flash(error_msg, 'danger')
            return redirect(url_for('main.settings'))
        
        existing = MonitoredPath.query.filter_by(path=path, user_id=user_id).first()
        if existing:
            error_msg = 'This path is already being monitored by this user'
            if is_api_request:
                return jsonify({'error': error_msg}), 400
            flash(error_msg, 'warning')
            return redirect(url_for('main.settings'))
        
        is_directory = path_type == 'directory'
        
        # Verify path type matches actual filesystem (only if not a client request)
        if not is_client_request:
            actual_is_dir = os.path.isdir(path)
            if is_directory != actual_is_dir:
                error_msg = f'Path type mismatch: {path} is {"a directory" if actual_is_dir else "a file"}'
                if is_api_request:
                    return jsonify({'error': error_msg}), 400
                flash(error_msg, 'danger')
                return redirect(url_for('main.settings'))
        
        last_modified = None
        file_size = None
        if not is_directory and not is_client_request:
            try:
                stat = os.stat(path)
                last_modified = datetime.fromtimestamp(stat.st_mtime)
                file_size = stat.st_size
            except:
                pass
        
        if not description and not is_directory:
            description = os.path.splitext(os.path.basename(path))[0].replace('_', ' ').replace('-', ' ').title()
        
        monitored_path = MonitoredPath(
            path=path, 
            user_id=user_id, 
            is_directory=is_directory,
            recursive=recursive if is_directory else False,
            description=description if description else None,
            last_modified=last_modified,
            file_size=file_size,
            change_count=0
        )
        db.session.add(monitored_path)
        db.session.commit()
        
        path_type = "directory" if is_directory else "file"
        success_msg = f'{path_type.capitalize()} added successfully'
        
        # Return JSON for API requests, redirect for web requests
        if is_api_request:
            return jsonify({'status': 'success', 'message': success_msg}), 200
        
        flash(success_msg, 'success')
        return redirect(url_for('main.settings'))
        
    except Exception as e:
        db.session.rollback()
        error_msg = f'Error adding monitored path: {str(e)}'
        
        # Check if this is an API request (need to redefine here since we're in except block)
        is_api_request = (request.content_type and 'application/json' in request.content_type) or \
                         request.headers.get('X-Client-Type') == 'FileMonitorTray'
        
        if is_api_request:
            return jsonify({'error': error_msg}), 500
        
        flash(error_msg, 'danger')
        return redirect(url_for('main.settings'))

@main_bp.route('/path/toggle/<int:id>')
@login_required
def toggle_monitored_path(id):
    path = MonitoredPath.query.get_or_404(id)
    
    if current_user.role == 'admin' or path.user_id == current_user.id:
        path.is_active = not path.is_active
        db.session.commit()
    
    return redirect(url_for('main.settings'))

@main_bp.route('/path/delete/<int:id>')
@login_required
def delete_monitored_path(id):
    path = MonitoredPath.query.get_or_404(id)
    
    if current_user.role == 'admin' or path.user_id == current_user.id:
        # First, delete all related FileChangeHistory records
        FileChangeHistory.query.filter_by(monitored_path_id=id).delete()
        
        # Then delete the monitored path
        db.session.delete(path)
        db.session.commit()
        
        flash('Monitored path deleted successfully', 'success')
    else:
        flash('You do not have permission to delete this path', 'error')
    
    return redirect(url_for('main.settings'))

@main_bp.route('/database_control')
@login_required
@admin_required
def database_control():
    # Calculate database statistics
    db_path = 'file_monitor.db'
    db_size_bytes = os.path.getsize(db_path) if os.path.exists(db_path) else 0
    db_size_mb = round(db_size_bytes / (1024 * 1024), 2)
    
    # Get oldest event
    oldest_event = Event.query.order_by(Event.timestamp.asc()).first()
    oldest_event_days = 0
    if oldest_event:
        # Ensure both datetimes are timezone-aware
        if oldest_event.timestamp.tzinfo is None:
            # If timestamp is naive, assume it's UTC
            oldest_timestamp_aware = oldest_event.timestamp.replace(tzinfo=timezone.utc)
        else:
            oldest_timestamp_aware = oldest_event.timestamp
        oldest_event_days = (datetime.now(timezone.utc) - oldest_timestamp_aware).days
    
    stats = {
        'total_events': Event.query.count(),
        'total_users': User.query.count(),
        'total_categories': Category.query.count(),
        'total_paths': MonitoredPath.query.count(),
        'db_size': f"{db_size_mb} MB",
        'oldest_event_days': oldest_event_days
    }
    
    # Get recent backups
    recent_backups = DatabaseBackup.query.order_by(DatabaseBackup.created_at.desc()).limit(10).all()
    
    # Get scheduled backup settings
    scheduled_backup = ScheduledBackupSettings.query.first()
    if not scheduled_backup:
        scheduled_backup = ScheduledBackupSettings()
        db.session.add(scheduled_backup)
        db.session.commit()
    
    return render_template('database_control.html',
                         stats=stats,
                         recent_backups=recent_backups,
                         scheduled_backup=scheduled_backup)

@main_bp.route('/database/backup', methods=['POST'])
@login_required
@admin_required
def backup_database():
    backup_type = request.form.get('backup_type', 'full')
    compress = request.form.get('compress') == 'on'
    note = request.form.get('backup_note', '').strip()
    
    # Create backup directory
    backup_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if backup_type == 'full':
        # Full database backup
        db_path = 'file_monitor.db'
        backup_filename = f"backup_full_{timestamp}.db"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Copy database file
        shutil.copy2(db_path, backup_path)
        
        if compress:
            # Compress the backup
            zip_filename = f"backup_full_{timestamp}.zip"
            zip_path = os.path.join(backup_dir, zip_filename)
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(backup_path, backup_filename)
            
            # Remove uncompressed file
            os.remove(backup_path)
            backup_filename = zip_filename
            backup_path = zip_path
        
        # Get file size
        size_mb = os.path.getsize(backup_path) / (1024 * 1024)
        
        # Save backup record
        backup_record = DatabaseBackup(
            filename=backup_filename,
            type=backup_type,
            size_mb=round(size_mb, 2),
            note=note,
            created_by_id=current_user.id
        )
        db.session.add(backup_record)
        db.session.commit()
        
        flash(get_translation('backup_created_successfully'), 'success')
        
        # Send file for download
        return send_file(backup_path, as_attachment=True, download_name=backup_filename)
    
    elif backup_type == 'data_only':
        # Export data to SQL format
        backup_filename = f"backup_data_{timestamp}.sql"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Generate SQL export
        with open(backup_path, 'w', encoding='utf-8') as f:
            # Export events
            f.write("-- Events Table\n")
            events = Event.query.all()
            for event in events:
                f.write(f"INSERT INTO event (timestamp, file_path, category_id, matched_keyword, computer_name, user_id, event_type, file_size) VALUES ('{event.timestamp}', '{event.file_path}', {event.category_id or 'NULL'}, '{event.matched_keyword or ''}', '{event.computer_name}', {event.user_id or 'NULL'}, '{event.event_type}', {event.file_size or 'NULL'});\n")
            
            # Export categories
            f.write("\n-- Categories Table\n")
            categories = Category.query.all()
            for cat in categories:
                f.write(f"INSERT INTO category (name, keywords, file_patterns, color) VALUES ('{cat.name}', '{cat.keywords}', '{cat.file_patterns}', '{cat.color}');\n")
        
        if compress:
            zip_filename = f"backup_data_{timestamp}.zip"
            zip_path = os.path.join(backup_dir, zip_filename)
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(backup_path, backup_filename)
            
            os.remove(backup_path)
            backup_filename = zip_filename
            backup_path = zip_path
        
        size_mb = os.path.getsize(backup_path) / (1024 * 1024)
        
        backup_record = DatabaseBackup(
            filename=backup_filename,
            type=backup_type,
            size_mb=round(size_mb, 2),
            note=note,
            created_by_id=current_user.id
        )
        db.session.add(backup_record)
        db.session.commit()
        
        return send_file(backup_path, as_attachment=True, download_name=backup_filename)
    
    elif backup_type == 'structure_only':
        # Export database structure only (schema)
        backup_filename = f"backup_structure_{timestamp}.sql"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Generate SQL schema export
        with open(backup_path, 'w', encoding='utf-8') as f:
            f.write("-- Database Structure Export\n")
            f.write("-- Generated on: {}\n\n".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            
            # Get schema from SQLAlchemy metadata
            from sqlalchemy import create_engine, MetaData
            engine = create_engine(app.config['SQLALCHEMY_DATABASE_URI'])
            metadata = MetaData()
            metadata.reflect(bind=engine)
            
            # Write CREATE TABLE statements
            for table in metadata.sorted_tables:
                f.write(f"-- Table: {table.name}\n")
                f.write(f"DROP TABLE IF EXISTS {table.name};\n")
                f.write(f"{str(table.create(engine)).compile(engine)};\n\n")
        
        if compress:
            zip_filename = f"backup_structure_{timestamp}.zip"
            zip_path = os.path.join(backup_dir, zip_filename)
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.write(backup_path, backup_filename)
            
            os.remove(backup_path)
            backup_filename = zip_filename
            backup_path = zip_path
        
        size_mb = os.path.getsize(backup_path) / (1024 * 1024)
        
        backup_record = DatabaseBackup(
            filename=backup_filename,
            type=backup_type,
            size_mb=round(size_mb, 2),
            note=note,
            created_by_id=current_user.id
        )
        db.session.add(backup_record)
        db.session.commit()
        
        flash(get_translation('backup_created_successfully'), 'success')
        return send_file(backup_path, as_attachment=True, download_name=backup_filename)
    
    return redirect(url_for('main.database_control'))

# Add these missing routes to your app.py file after the backup_database route (around line 1650)

@main_bp.route('/database/restore', methods=['POST'])
@login_required
@admin_required
def restore_database():
    """Restore database from uploaded backup file"""
    try:
        # Check if a file was uploaded
        if 'backup_file' not in request.files:
            flash(get_translation('no_file_selected'), 'error')
            return redirect(url_for('main.database_control'))
        
        file = request.files['backup_file']
        
        # Check if file is selected
        if file.filename == '':
            flash(get_translation('no_file_selected'), 'error')
            return redirect(url_for('main.database_control'))
        
        # Validate file extension
        allowed_extensions = {'.db', '.zip', '.sql'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            flash(get_translation('invalid_file_format'), 'error')
            return redirect(url_for('main.database_control'))
        
        # Save uploaded file temporarily
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, file.filename)
        file.save(temp_path)
        
        # Create safety backup of current database
        current_db_path = 'file_monitor.db'
        safety_backup_path = f"{current_db_path}.safety_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shutil.copy2(current_db_path, safety_backup_path)
        
        try:
            # Close all database connections
            db.session.remove()
            db.engine.dispose()
            
            if file_ext == '.zip':
                # Extract and restore from zip
                with zipfile.ZipFile(temp_path, 'r') as zip_ref:
                    # Find the .db file in the zip
                    db_files = [f for f in zip_ref.namelist() if f.endswith('.db')]
                    if not db_files:
                        raise ValueError("No database file found in zip")
                    
                    # Extract to temp location
                    zip_ref.extract(db_files[0], temp_dir)
                    extracted_db = os.path.join(temp_dir, db_files[0])
                    
                    # Replace current database
                    shutil.copy2(extracted_db, current_db_path)
                    
            elif file_ext == '.db':
                # Direct database file replacement
                shutil.copy2(temp_path, current_db_path)
                
            elif file_ext == '.sql':
                # SQL file restoration (not implemented yet)
                raise NotImplementedError("SQL file restoration not yet implemented")
            
            # Clean up safety backup if successful
            os.remove(safety_backup_path)
            
            # Clean up temp files
            shutil.rmtree(temp_dir)
            
            flash(get_translation('restore_successful'), 'success')
            
            # Force app restart or reconnect to database
            return redirect(url_for('main.database_control'))
            
        except Exception as e:
            # Restore from safety backup if something went wrong
            if os.path.exists(safety_backup_path):
                shutil.copy2(safety_backup_path, current_db_path)
                os.remove(safety_backup_path)
            
            # Clean up temp files
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            
            flash(f"{get_translation('restore_failed')}: {str(e)}", 'error')
            return redirect(url_for('main.database_control'))
            
    except Exception as e:
        flash(f"{get_translation('restore_failed')}: {str(e)}", 'error')
        return redirect(url_for('main.database_control'))

@main_bp.route('/database/cleanup_events', methods=['POST'])
@login_required
@admin_required
def cleanup_events():
    """Clean up old events"""
    cleanup_period = request.form.get('cleanup_period', type=int)
    
    if cleanup_period:
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=cleanup_period)
        
        # Get events to delete
        events_to_delete = Event.query.filter(Event.timestamp < cutoff_date).all()
        event_ids = [event.id for event in events_to_delete]
        
        # Delete related CNC analysis records first
        CNCAnalysis.query.filter(CNCAnalysis.event_id.in_(event_ids)).delete(synchronize_session=False)
        
        # Delete events
        deleted_count = Event.query.filter(Event.timestamp < cutoff_date).delete()
        db.session.commit()
        
        flash(f'Deleted {deleted_count} events older than {cleanup_period} days', 'success')
    
    return redirect(url_for('main.database_control'))

@main_bp.route('/database/optimize', methods=['POST'])
@login_required
@admin_required
def optimize_database():
    """Optimize database"""
    try:
        # Run VACUUM on SQLite database
        db.session.execute(text('VACUUM'))
        db.session.commit()
        flash('Database optimized successfully', 'success')
    except Exception as e:
        flash(f'Error optimizing database: {str(e)}', 'danger')
    
    return redirect(url_for('main.database_control'))

@main_bp.route('/database/update_backup_schedule', methods=['POST'])
@login_required
@admin_required
def update_backup_schedule():
    """Update scheduled backup settings"""
    try:
        # Get form data
        enabled = request.form.get('enable_scheduled') == 'on'
        frequency = request.form.get('backup_frequency', 'daily')
        backup_time = request.form.get('backup_time', '02:00')
        retention_days = int(request.form.get('retention_days', 30))
        
        # Get or create scheduled backup settings
        scheduled_backup = ScheduledBackupSettings.query.first()
        if not scheduled_backup:
            scheduled_backup = ScheduledBackupSettings()
            db.session.add(scheduled_backup)
        
        # Update settings
        scheduled_backup.enabled = enabled
        scheduled_backup.frequency = frequency
        scheduled_backup.time = backup_time
        scheduled_backup.retention_days = retention_days
        
        # Calculate next run time if enabled
        if enabled:
            scheduled_backup.next_run = calculate_next_backup_time(scheduled_backup)
        
        db.session.commit()
        
        # Update the scheduler job
        update_backup_schedule_job()
        
        flash(get_translation('backup_schedule_updated'), 'success')
    except Exception as e:
        print(f"Error updating backup schedule: {str(e)}")
        flash(get_translation('error_updating_schedule'), 'error')
    
    return redirect(url_for('main.database_control'))

# API routes for database control
@api_bp.route('/database/cleanup_count')
@login_required
@admin_required
def cleanup_count():
    """Get count of events that would be deleted"""
    days = request.args.get('days', type=int, default=30)
    cutoff_date = datetime.now(timezone.utc) - timedelta(days=days)
    count = Event.query.filter(Event.timestamp < cutoff_date).count()
    return jsonify({'count': count})

@api_bp.route('/database/reset', methods=['POST'])
@login_required
@admin_required
def reset_database():
    """Reset database (delete all events)"""
    try:
        # Delete CNC analysis records first
        CNCAnalysis.query.delete()
        
        # Delete all events
        Event.query.delete()
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@api_bp.route('/database/delete_all', methods=['POST'])
@login_required
@admin_required
def delete_all_data():
    """Delete all data - extremely dangerous"""
    # This is a placeholder - implement with extreme caution
    return jsonify({'success': False, 'error': 'Not implemented for safety'})

# Add these missing routes to app.py after the existing database routes (around line 1650)

@main_bp.route('/database/export', methods=['POST'])
@login_required
@admin_required
def export_database():
    """Export database to various formats"""
    export_format = request.form.get('export_format', 'csv')
    
    # Get selected data types
    export_events = request.form.get('export_events') == 'on'
    export_users = request.form.get('export_users') == 'on'
    export_categories = request.form.get('export_categories') == 'on'
    export_paths = request.form.get('export_paths') == 'on'
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if export_format == 'csv':
        # Create a ZIP file with multiple CSVs
        zip_filename = f"export_{timestamp}.zip"
        zip_path = os.path.join(tempfile.gettempdir(), zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Export events
            if export_events:
                csv_content = StringIO()
                writer = csv.writer(csv_content)
                writer.writerow(['ID', 'Timestamp', 'File Path', 'Category', 'Keyword', 'Computer', 'User', 'Type', 'Size'])
                
                events = Event.query.all()
                for event in events:
                    writer.writerow([
                        event.id,
                        event.timestamp.isoformat(),
                        event.file_path,
                        event.category.name if event.category else '',
                        event.matched_keyword or '',
                        event.computer_name,
                        event.user.username if event.user else '',
                        event.event_type,
                        event.file_size or ''
                    ])
                
                zipf.writestr('events.csv', csv_content.getvalue())
            
            # Export users
            if export_users:
                csv_content = StringIO()
                writer = csv.writer(csv_content)
                writer.writerow(['ID', 'Username', 'Email', 'Role', 'Active', 'Created At'])
                
                users = User.query.all()
                for user in users:
                    writer.writerow([
                        user.id,
                        user.username,
                        user.email,
                        user.role,
                        user.is_active,
                        user.created_at.isoformat()
                    ])
                
                zipf.writestr('users.csv', csv_content.getvalue())
            
            # Export categories
            if export_categories:
                csv_content = StringIO()
                writer = csv.writer(csv_content)
                writer.writerow(['ID', 'Name', 'Keywords', 'Patterns', 'Color'])
                
                categories = Category.query.all()
                for cat in categories:
                    writer.writerow([
                        cat.id,
                        cat.name,
                        ', '.join(cat.get_keywords()),
                        ', '.join(cat.get_patterns()),
                        cat.color
                    ])
                
                zipf.writestr('categories.csv', csv_content.getvalue())
            
            # Export monitored paths
            if export_paths:
                csv_content = StringIO()
                writer = csv.writer(csv_content)
                writer.writerow(['ID', 'Path', 'User', 'Type', 'Active', 'Description'])
                
                paths = MonitoredPath.query.all()
                for path in paths:
                    writer.writerow([
                        path.id,
                        path.path,
                        path.user.username,
                        'Directory' if path.is_directory else 'File',
                        path.is_active,
                        path.description or ''
                    ])
                
                zipf.writestr('monitored_paths.csv', csv_content.getvalue())
        
        return send_file(zip_path, as_attachment=True, download_name=zip_filename, mimetype='application/zip')
    
    elif export_format == 'json':
        # Export as JSON
        data = {}
        
        if export_events:
            data['events'] = [{
                'id': e.id,
                'timestamp': e.timestamp.isoformat(),
                'file_path': e.file_path,
                'category': e.category.name if e.category else None,
                'matched_keyword': e.matched_keyword,
                'computer_name': e.computer_name,
                'user': e.user.username if e.user else None,
                'event_type': e.event_type,
                'file_size': e.file_size
            } for e in Event.query.all()]
        
        if export_users:
            data['users'] = [{
                'id': u.id,
                'username': u.username,
                'email': u.email,
                'role': u.role,
                'is_active': u.is_active,
                'created_at': u.created_at.isoformat()
            } for u in User.query.all()]
        
        if export_categories:
            data['categories'] = [{
                'id': c.id,
                'name': c.name,
                'keywords': c.get_keywords(),
                'patterns': c.get_patterns(),
                'color': c.color
            } for c in Category.query.all()]
        
        if export_paths:
            data['monitored_paths'] = [{
                'id': p.id,
                'path': p.path,
                'user': p.user.username,
                'is_directory': p.is_directory,
                'is_active': p.is_active,
                'description': p.description
            } for p in MonitoredPath.query.all()]
        
        json_content = json.dumps(data, indent=2)
        json_filename = f"export_{timestamp}.json"
        
        return send_file(
            BytesIO(json_content.encode()),
            as_attachment=True,
            download_name=json_filename,
            mimetype='application/json'
        )
    
    elif export_format == 'sql':
        # Export as SQL dump
        sql_content = StringIO()
        sql_content.write("-- Database Export\n")
        sql_content.write(f"-- Generated on {datetime.now()}\n\n")
        
        if export_events:
            sql_content.write("-- Events Table\n")
            events = Event.query.all()
            for event in events:
                sql_content.write(
                    f"INSERT INTO event (id, timestamp, file_path, category_id, matched_keyword, "
                    f"computer_name, user_id, event_type, file_size) VALUES ("
                    f"{event.id}, '{event.timestamp}', '{event.file_path}', "
                    f"{event.category_id or 'NULL'}, '{event.matched_keyword or ''}', "
                    f"'{event.computer_name}', {event.user_id or 'NULL'}, "
                    f"'{event.event_type}', {event.file_size or 'NULL'});\n"
                )
            sql_content.write("\n")
        
        if export_categories:
            sql_content.write("-- Categories Table\n")
            categories = Category.query.all()
            for cat in categories:
                sql_content.write(
                    f"INSERT INTO category (id, name, keywords, file_patterns, color) VALUES ("
                    f"{cat.id}, '{cat.name}', '{cat.keywords}', '{cat.file_patterns}', '{cat.color}');\n"
                )
            sql_content.write("\n")
        
        sql_filename = f"export_{timestamp}.sql"
        
        return send_file(
            BytesIO(sql_content.getvalue().encode()),
            as_attachment=True,
            download_name=sql_filename,
            mimetype='text/plain'
        )
    
    # Default: redirect back
    return redirect(url_for('main.database_control'))

@main_bp.route('/database/backup/<int:id>', methods=['DELETE'])
@login_required
@admin_required
def delete_backup(id):
    """Delete a database backup"""
    backup = DatabaseBackup.query.get_or_404(id)
    
    try:
        # Delete the file
        backup_path = os.path.join(app.config['UPLOAD_FOLDER'], 'backups', backup.filename)
        if os.path.exists(backup_path):
            os.remove(backup_path)
        
        # Delete the record
        db.session.delete(backup)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/database/backup/<int:id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_backup(id):
    """API endpoint to delete a database backup"""
    return delete_backup(id)

@api_bp.route('/database/restore/<filename>', methods=['POST'])
@login_required
@admin_required
def api_restore_database(filename):
    """API endpoint to restore database from backup"""
    try:
        backup_path = os.path.join(app.config['UPLOAD_FOLDER'], 'backups', filename)
        
        if not os.path.exists(backup_path):
            return jsonify({'success': False, 'error': 'Backup file not found'}), 404
        
        # For SQLite, we can simply replace the database file
        # First, make a safety backup of current database
        current_db_path = 'file_monitor.db'
        safety_backup_path = f"{current_db_path}.safety_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shutil.copy2(current_db_path, safety_backup_path)
        
        try:
            # Close all database connections
            db.session.close_all()
            db.engine.dispose()
            
            # Extract if it's a zip file
            if filename.endswith('.zip'):
                with zipfile.ZipFile(backup_path, 'r') as zip_ref:
                    # Find the .db file in the zip
                    db_files = [f for f in zip_ref.namelist() if f.endswith('.db')]
                    if db_files:
                        # Extract to temp location
                        temp_db = os.path.join(tempfile.gettempdir(), 'restore_temp.db')
                        zip_ref.extract(db_files[0], tempfile.gettempdir())
                        extracted_path = os.path.join(tempfile.gettempdir(), db_files[0])
                        shutil.move(extracted_path, temp_db)
                        
                        # Replace current database
                        shutil.copy2(temp_db, current_db_path)
                        os.remove(temp_db)
            else:
                # Direct database file
                shutil.copy2(backup_path, current_db_path)
            
            # Remove safety backup if restore was successful
            os.remove(safety_backup_path)
            
            return jsonify({'success': True, 'message': 'Database restored successfully'})
            
        except Exception as e:
            # Restore from safety backup if something went wrong
            shutil.copy2(safety_backup_path, current_db_path)
            os.remove(safety_backup_path)
            raise e
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/report/<int:id>', methods=['DELETE'])
@login_required
def delete_report(id):
    """Delete a report"""
    report = Report.query.get_or_404(id)
    
    if current_user.role != 'admin' and report.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        # Delete file if exists
        if os.path.exists(report.file_path):
            os.remove(report.file_path)
        
        db.session.delete(report)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@api_bp.route('/report/preview', methods=['POST'])
@login_required
def preview_report():
    """Generate report preview"""
    # Placeholder implementation
    return jsonify({
        'html': '<div class="alert alert-info">Report preview not yet implemented</div>'
    })

@api_bp.route('/validate_path', methods=['POST'])
@login_required
def validate_path():
    """Validate if a path exists and is accessible"""
    data = request.get_json()
    path = data.get('path', '')
    
    result = {
        'exists': os.path.exists(path),
        'is_directory': os.path.isdir(path) if os.path.exists(path) else False,
        'readable': os.access(path, os.R_OK) if os.path.exists(path) else False
    }
    
    return jsonify(result)

@api_bp.route('/monitor/start', methods=['POST'])
@login_required
@admin_required
def start_monitor():
    """Start monitor (placeholder)"""
    return jsonify({'status': 'Monitor is handled by client application'})

@api_bp.route('/monitor/status')
@login_required
def monitor_status_api():
    """Get detailed monitor status"""
    paths_query = MonitoredPath.query.filter_by(is_active=True)
    
    if current_user.role != 'admin':
        paths_query = paths_query.filter_by(user_id=current_user.id)
    
    total_paths = paths_query.count()
    total_files = paths_query.filter_by(is_directory=False).count()
    total_dirs = paths_query.filter_by(is_directory=True).count()
    
    return jsonify({
        'running': False,  # Monitor runs on client
        'paths': total_paths,
        'files': total_files,
        'directories': total_dirs
    })
# Helper functions for report generation
def calculate_date_range(range_type):
    """Calculate date range based on range type"""
    today = datetime.now(timezone.utc).date()
    
    if range_type == 'today':
        return today, today
    elif range_type == 'yesterday':
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday
    elif range_type == 'week':
        start = today - timedelta(days=7)
        return start, today
    elif range_type == 'month':
        start = today.replace(day=1)
        return start, today
    elif range_type == 'year':
        start = today.replace(month=1, day=1)
        return start, today
    
    return today, today

def generate_dashboard_report(date_from, date_to, user_filter, form_data):
    """Generate dashboard-style report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard Report"
    
    # Add headers and basic implementation
    ws['A1'] = 'Dashboard Report'
    ws['A2'] = f'Period: {date_from} to {date_to}'
    
    # Add more implementation as needed
    return wb

def generate_audit_report(date_from, date_to, user_filter, form_data):
    """Generate audit report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Report"
    
    # Add headers and basic implementation
    ws['A1'] = 'Audit Report'
    ws['A2'] = f'Period: {date_from} to {date_to}'
    
    # Add more implementation as needed
    return wb

def generate_summary_report(date_from, date_to, user_filter):
    """Generate summary report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary Report"
    
    # Add headers and basic implementation
    ws['A1'] = 'Summary Report'
    ws['A2'] = f'Period: {date_from} to {date_to}'
    
    # Add more implementation as needed
    return wb

def generate_detailed_report(date_from, date_to, user_filter, form_data):
    """Generate detailed report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed Report"
    
    # Add headers and basic implementation
    ws['A1'] = 'Detailed Report'
    ws['A2'] = f'Period: {date_from} to {date_to}'
    
    # Add more implementation as needed
    return wb

def create_default_categories():
    """Create some default categories for demo purposes"""
    with app.app_context():
        if Category.query.count() == 0:
            default_categories = [
                {
                    'name': 'Documents',
                    'keywords': ['document', 'doc', 'report', 'text'],
                    'file_patterns': ['.*\\.doc.*', '.*\\.pdf', '.*\\.txt'],
                    'color': '#007bff'
                },
                {
                    'name': 'Images',
                    'keywords': ['image', 'photo', 'picture'],
                    'file_patterns': ['.*\\.jpg', '.*\\.png', '.*\\.gif', '.*\\.bmp'],
                    'color': '#28a745'
                },
                {
                    'name': 'Data Files',
                    'keywords': ['data', 'csv', 'excel', 'spreadsheet'],
                    'file_patterns': ['.*\\.csv', '.*\\.xlsx?', '.*\\.json'],
                    'color': '#ffc107'
                },
                {
                    'name': 'Log Files',
                    'keywords': ['log', 'error', 'debug', 'trace'],
                    'file_patterns': ['.*\\.log', '.*\\.trace'],
                    'color': '#dc3545'
                },
                {
                    'name': 'Allerlei',
                    'keywords': [],
                    'file_patterns': [],
                    'color': '#6c757d'
                }
            ]
            
            for cat_data in default_categories:
                category = Category(
                    name=cat_data['name'],
                    keywords=json.dumps(cat_data['keywords']),
                    file_patterns=json.dumps(cat_data['file_patterns']),
                    color=cat_data['color']
                )
                db.session.add(category)
            
            db.session.commit()
            print("Default categories created.")

# Register blueprints
app.register_blueprint(auth_bp)
app.register_blueprint(main_bp)
app.register_blueprint(api_bp)

# Initialize database function
def initialize_database():
    """Initialize the database"""
    with app.app_context():
        try:
            # Check and migrate database first
            if not check_and_migrate_database():
                print("❌ Database migration failed. Please check the error messages.")
                return False
            
            # Create all database tables
            db.create_all()
            print("Database tables created successfully.")
            
            # Create backup directories
            os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'backups'), exist_ok=True)
            os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'backups', 'temp'), exist_ok=True)
            
            # Create default admin user if it doesn't exist
            if not User.query.filter_by(role='admin').first():
                admin = User(
                    username='admin',
                    email='admin@example.com',
                    role='admin'
                )
                admin.set_password('admin123')
                db.session.add(admin)
                db.session.commit()
                print("Default admin user created: admin/admin123")
            else:
                print("Admin user already exists.")
            
            # File monitoring is now handled by client applications
            print("Note: File monitoring is handled by client applications.")
            
            return True
            
        except Exception as e:
            print(f"Error during database initialization: {e}")
            return False

# Main execution
if __name__ == '__main__':
    try:
        print("Enterprise File Monitor starting...")
        if not initialize_database():
            print("❌ Failed to initialize database. Exiting.")
            exit(1)
        
        create_default_categories()
        
        # Initialize scheduled backup job if enabled
        with app.app_context():
            try:
                update_backup_schedule_job()
                print("✓ Scheduled backup job initialized")
            except Exception as e:
                print(f"Warning: Failed to initialize scheduled backup job: {e}")
        
        if not os.environ.get('WERKZEUG_RUN_MAIN'):
            print("="*50)
            print("Enterprise File Monitor is running!")
            print("URL: http://localhost:5002")
            print("Default login: admin / admin123")
            print("Note: File monitoring is handled by client applications")
            print("="*50)
        
        app.run(debug=False, host='0.0.0.0', port=5002, use_reloader=False)
        
    except KeyboardInterrupt:
        print("\nShutting down gracefully...")
    except Exception as e:
        print(f"Error starting application: {e}")
        import traceback
        traceback.print_exc()