import datetime # Temporary import for early debug
print(f"[{datetime.datetime.now()}] DEBUG: Script execution started (top of file).")
import os
import time
import json
import atexit
import traceback
import gc
import threading
import signal
import logging
from file_lock import FileLock
from datetime import date, timedelta # datetime already imported
from collections import defaultdict
print(f"[{datetime.datetime.now()}] DEBUG: Standard library imports complete.")
import openpyxl
from openpyxl.chart import BarChart, PieChart, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import Series
from openpyxl.chart.reference import Reference
from openpyxl.styles import Font
print(f"[{datetime.datetime.now()}] DEBUG: Openpyxl imports complete.")

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] [%(module)s.%(funcName)s:%(lineno)d] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

# --- Global Variables & Configuration ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
try:
    with open(os.path.join(BASE_DIR, 'logger_config.json'), 'r') as f:
        CONFIG = json.load(f)
except Exception as e_config:
    print(f"[{datetime.datetime.now()}] CRITICAL_ERROR: Failed to load or parse 'logger_config.json'. Error: {e_config}")
    print(traceback.format_exc())
    # Exit immediately if config fails to load, as the script cannot run.
    # We use os._exit(1) for an immediate exit without further cleanup, 
    # as standard sys.exit() might be affected if other parts of the script are broken.
    os._exit(1) # type: ignore

REPORT_GENERATION_LOCK = threading.Lock()
SEEN_ENTRIES_FILE = os.path.join(BASE_DIR, CONFIG['seen_entries_file'])
CATEGORIES_FILE = os.path.join(BASE_DIR, CONFIG['categories_file'])
EVENT_HISTORY_FILE = os.path.join(BASE_DIR, 'events_history.json')
EXCEL_REPORT_DIR = os.path.join(BASE_DIR, CONFIG['excel_report_dir'])
SCAN_INTERVAL = CONFIG['scan_interval_seconds']
SEEN_ENTRIES_LOCK_FILE = SEEN_ENTRIES_FILE + ".lock"

PROGRAM_START_TIME = datetime.datetime.now()
RUNTIME_CATEGORY_TOTALS = defaultdict(int)
RUNTIME_DAILY_ITEM_COUNTS = defaultdict(lambda: defaultdict(int))
RUNTIME_SEEN_ENTRIES = {}
print(f"[{datetime.datetime.now()}] DEBUG: Imports and initial global variables complete.")

# --- GUI Data Provider Functions ---
def get_program_start_time():
    return PROGRAM_START_TIME

def get_runtime_category_totals():
    return dict(RUNTIME_CATEGORY_TOTALS)

def get_runtime_daily_item_counts():
    return dict(RUNTIME_DAILY_ITEM_COUNTS)

# --- File and State Management ---
with open(CATEGORIES_FILE, 'r') as f:
    CATEGORIES = json.load(f)

def categorize(path):
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read().lower()
            sorted_categories = sorted(CATEGORIES.items(), key=lambda x: -len(x[0]))
            for prefix, category in sorted_categories:
                if prefix.lower() in content:
                    return category
    except Exception as e:
        print(f"[ERROR] Could not read file {path} for categorization: {e}")
    return 'Allerlei'

def load_seen():
    try:
        with FileLock(SEEN_ENTRIES_LOCK_FILE, timeout=10):
            if not os.path.exists(SEEN_ENTRIES_FILE):
                return {'entries': {}, 'save_events': []}
            with open(SEEN_ENTRIES_FILE, 'r') as f:
                content = f.read()
                return json.loads(content) if content else {'entries': {}, 'save_events': []}
    except Exception as e:
        print(f"[ERROR] Failed to load seen entries: {e}. Starting fresh.")
        return {'entries': {}, 'save_events': []}

def save_seen(data):
    try:
        with FileLock(SEEN_ENTRIES_LOCK_FILE, timeout=10):
            with open(SEEN_ENTRIES_FILE, 'w') as f:
                json.dump(data, f, indent=4)
    except Exception as e:
        print(f"[ERROR] Failed to save seen entries: {e}")

def clear_seen_entries_but_keep_last_on_exit():
    """
    Atexit handler to clear processed manual events from the state file.
    This is a safeguard to ensure that on a clean shutdown, the manual events queue is empty
    for the next run. It preserves the file mtime entries.
    """
    logging.info("atexit: Performing cleanup of seen entries file.")
    try:
        with FileLock(SEEN_ENTRIES_LOCK_FILE, timeout=5):
            if not os.path.exists(SEEN_ENTRIES_FILE):
                return

            data = None
            try:
                with open(SEEN_ENTRIES_FILE, 'r') as f:
                    data = json.load(f)
            except (json.JSONDecodeError, FileNotFoundError):
                logging.warning("atexit: Could not read or decode seen entries file. Cannot clear events.")
                return 

            if data and 'save_events' in data and data['save_events']:
                logging.info(f"atexit: Clearing {len(data['save_events'])} manual events.")
                data['save_events'] = []
                with open(SEEN_ENTRIES_FILE, 'w') as f:
                    json.dump(data, f, indent=4)

    except Exception as e:
        logging.error(f"atexit: Exception during cleanup: {e}", exc_info=True)


def load_event_history():
    if os.path.exists(EVENT_HISTORY_FILE):
        with open(EVENT_HISTORY_FILE, 'r') as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                return []
    return []

def append_event_history(new_events):
    if not new_events:
        return
    try:
        history = load_event_history()
        history.extend(new_events)
        with open(EVENT_HISTORY_FILE, 'w') as f:
            json.dump(history, f, indent=4)
    except Exception as e:
        print(f"[ERROR] Failed to append to event history: {e}")

def process_runtime_manual_events(event_history):
    manual_events_processed = False
    seen_data = load_seen()
    manual_events_to_process = seen_data.get('save_events', [])

    if not manual_events_to_process:
        return False

    newly_processed_manual_events = []
    for event_data in manual_events_to_process:
        try:
            category = event_data.get('category', 'Allerlei')
            event_dt = datetime.datetime.fromisoformat(event_data['timestamp'])
            RUNTIME_CATEGORY_TOTALS[category] += 1
            RUNTIME_DAILY_ITEM_COUNTS[event_dt.date()][event_dt.hour] += 1
            event_history.append(event_data)
            newly_processed_manual_events.append(event_data)
            manual_events_processed = True
        except Exception as e:
            print(f"[ERROR] Failed to process manual event: {e}")

    if newly_processed_manual_events:
        append_event_history(newly_processed_manual_events)
        seen_data['save_events'] = []
        save_seen(seen_data)
        print(f"[INFO] Processed and cleared {len(newly_processed_manual_events)} manual events.")

    return manual_events_processed

# --- Report Generation ---
def generate_weekly_report(events):
    with REPORT_GENERATION_LOCK:
        now = datetime.datetime.now()
        year, week_num, _ = now.isocalendar()
        week_events = [e for e in events if datetime.datetime.fromisoformat(e['timestamp']).isocalendar()[:2] == (year, week_num)]
        if not week_events: return

        report_dir = os.path.join(EXCEL_REPORT_DIR, 'Week Raporten')
        os.makedirs(report_dir, exist_ok=True)
        report_path = os.path.join(report_dir, f"Week_{week_num}_{year}.xlsx")
        # Add basic content to prove generation
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Week {week_num}"
        ws['A1'] = f"Week Rapport for Week {week_num}, {year}"
        ws['A2'] = f"Totaal Events: {len(week_events)}"
        wb.save(report_path)
        print(f"[INFO] Weekly report generated: {report_path}")
        return str(year)


# --- New function for weekly report ---

def generate_monthly_report(events):
    """
    Generates a monthly report for the current month.
    The report includes a summary of events per category and a bar chart.
    """
    with REPORT_GENERATION_LOCK:
        now_dt = datetime.datetime.now()
        year = now_dt.year
        month = now_dt.month

        # --- 1. Filter events for the current month ---
        month_events = [
            event for event in events
            if datetime.datetime.fromisoformat(event['timestamp']).year == year and \
               datetime.datetime.fromisoformat(event['timestamp']).month == month
        ]

        if not month_events:
            print(f"[{datetime.datetime.now()}] [MONTHLY_REPORT] No events found for {year}-{month:02d}. Report not generated.")
            return

        # --- 2. Prepare folder and filename ---
        report_dir = os.path.join(EXCEL_REPORT_DIR, 'Maand Raporten')
        try:
            os.makedirs(report_dir, exist_ok=True)
        except OSError as e:
            print(f"[{datetime.datetime.now()}] [MONTHLY_REPORT_ERROR] Could not create report directory {report_dir}: {e}")
            return
        
        report_path = os.path.join(report_dir, f"Maand_{year}-{month:02d}.xlsx")

        # --- 3. Aggregate data by category ---
        category_counts = defaultdict(int)
        for event in month_events:
            category = event.get('category', 'Allerlei') # Default category
            if event['type'] == 'new':
                category_counts[category] += 1
            elif event['type'] == 'duplicate':
                entries = event.get('entries', [])
                if isinstance(entries, list):
                    category_counts[category] += len(entries)
                else:
                    category_counts[category] += 1 # Count as one if not a list

        # --- 4. Create Workbook and Worksheet ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Maand {year}-{month:02d}"

        # --- 5. Create Data Table ---
        ws.append(["Categorie", "Aantal", "Percentage"])
        total_events = sum(category_counts.values())
        
        sorted_categories = sorted(category_counts.items(), key=lambda item: (-item[1], item[0]))

        for category, count in sorted_categories:
            percentage = (count / total_events) * 100 if total_events > 0 else 0
            ws.append([category, count, percentage])
            ws.cell(row=ws.max_row, column=3).number_format = '0.00"%"'

        # --- 6. Create Bar Chart for Category Totals ---
        if ws.max_row > 1:
            chart = BarChart()
            chart.title = f"Meldingen per Categorie (Maand {year}-{month:02d})"
            chart.x_axis.title = "Categorie"
            chart.y_axis.title = "Aantal Meldingen"
            chart.legend = None

            data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)

            chart.style = 10
            chart.width = 20
            chart.height = 12

            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True

            ws.add_chart(chart, "E2")

        # --- 7. Save the Workbook ---
        try:
            print(f"[{datetime.datetime.now()}] [MONTHLY_REPORT] Attempting to save: {report_path}")
            wb.save(report_path)
            print(f"[{datetime.datetime.now()}] [MONTHLY_REPORT_SUCCESS] Monthly report saved: {report_path}")
        except Exception as e:
            print(f"[{datetime.datetime.now()}] [MONTHLY_REPORT_ERROR] Failed to save monthly report {report_path}: {e}")
            traceback.print_exc()
        finally:
            if 'wb' in locals() and wb is not None:
                try:
                    wb.close()
                except Exception as e_close:
                    print(f"[{datetime.datetime.now()}] [MONTHLY_REPORT_ERROR] Error closing workbook: {e_close}")
            gc.collect()

def generate_weekly_report(events):
    """
    Generates a weekly report for the current ISO week.
    Includes: Items per Category (Bar & Pie), Items per Worked Hour per Day (Bar), Usage per Scan File (Pie).
    """
    with REPORT_GENERATION_LOCK:
        now_dt = datetime.datetime.now()
        year, week_num, _ = now_dt.isocalendar()

        week_events = [
            event for event in events
            if datetime.datetime.fromisoformat(event['timestamp']).isocalendar()[:2] == (year, week_num)
        ]

        if not week_events:
            print(f"[{datetime.datetime.now()}] [WEEKLY_REPORT] No events found for week {week_num}, {year}. Report not generated.")
            return

        report_dir = os.path.join(EXCEL_REPORT_DIR, 'Week Raporten')
        os.makedirs(report_dir, exist_ok=True)
        report_path = os.path.join(report_dir, f"Week_{week_num}_{year}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Week {week_num}"
        
        current_row = 1
        ws.cell(row=current_row, column=1, value=f"Rapportage Week {week_num}, {year}").font = Font(bold=True, size=14)
        current_row += 2

        # --- 1. Items per Category --- 
        ws.cell(row=current_row, column=1, value="Items per Categorie").font = Font(bold=True)
        current_row += 1
        cat_header = ["Categorie", "Aantal", "Percentage"]
        for col_idx, h_val in enumerate(cat_header, 1):
            ws.cell(row=current_row, column=col_idx, value=h_val).font = Font(bold=True)
        current_row += 1
        
        category_counts = defaultdict(int)
        for event in week_events:
            category = event.get('category', 'Allerlei')
            if event['type'] == 'new':
                category_counts[category] += 1
            elif event['type'] == 'duplicate':
                entries = event.get('entries', [])
                category_counts[category] += len(entries) if isinstance(entries, list) else 1
        
        total_items_in_week = sum(category_counts.values())
        start_data_row_cat, end_data_row_cat = 0, 0

        if total_items_in_week > 0:
            start_data_row_cat = current_row
            sorted_categories = sorted(category_counts.items(), key=lambda item: (-item[1], item[0]))
            for category, count in sorted_categories:
                percentage = (count / total_items_in_week) * 100
                ws.cell(row=current_row, column=1, value=category)
                ws.cell(row=current_row, column=2, value=count)
                ws.cell(row=current_row, column=3, value=percentage).number_format = '0.00"%"'
                current_row += 1
            end_data_row_cat = current_row - 1

        current_row += 1 # Spacer

        # --- 2. Items per Day per Worked Hour --- 
        ws.cell(row=current_row, column=1, value="Items per Gewerkt Uur per Dag").font = Font(bold=True)
        current_row += 1
        daily_header = ["Dag", "Totaal Items", "Gewerkte Uren", "Items/Gew.Uur"]
        for col_idx, h_val in enumerate(daily_header, 1):
            ws.cell(row=current_row, column=col_idx, value=h_val).font = Font(bold=True)
        current_row += 1

        daily_item_counts = defaultdict(int)
        for event in week_events:
            event_dt = datetime.datetime.fromisoformat(event['timestamp'])
            day_of_week = event_dt.weekday()
            if event['type'] == 'new':
                daily_item_counts[day_of_week] += 1
            elif event['type'] == 'duplicate':
                entries = event.get('entries', [])
                daily_item_counts[day_of_week] += len(entries) if isinstance(entries, list) else 1

        workday_hours_config = CONFIG.get('workday_hours', {})
        days_of_week_names = ["Maandag", "Dinsdag", "Woensdag", "Donderdag", "Vrijdag", "Zaterdag", "Zondag"]
        start_data_row_daily, end_data_row_daily = current_row, 0

        for i in range(7):
            day_name = days_of_week_names[i]
            items_on_day = daily_item_counts[i]
            worked_hours_on_day = float(workday_hours_config.get(str(i), 0.0))
            items_per_worked_hour = (items_on_day / worked_hours_on_day) if worked_hours_on_day > 0 else 0
            
            ws.cell(row=current_row, column=1, value=day_name)
            ws.cell(row=current_row, column=2, value=items_on_day)
            ws.cell(row=current_row, column=3, value=worked_hours_on_day)
            ws.cell(row=current_row, column=4, value=items_per_worked_hour).number_format = '0.00'
            current_row += 1
        end_data_row_daily = current_row - 1
        current_row += 1 # Spacer

        # --- 3. Scan Files Usage --- 
        ws.cell(row=current_row, column=1, value="Gebruik per Scan Bestand").font = Font(bold=True)
        current_row += 1
        scan_header = ["Scan Bestand", "Aantal Items", "Percentage"]
        for col_idx, h_val in enumerate(scan_header, 1):
            ws.cell(row=current_row, column=col_idx, value=h_val).font = Font(bold=True)
        current_row += 1

        scan_files_config = CONFIG.get('scan_files', [])
        scan_file_item_counts = defaultdict(int)
        for event in week_events:
            event_entries_paths = []
            if event['type'] == 'new' and event.get('entry'):
                event_entries_paths.append(event['entry'])
            elif event['type'] == 'duplicate' and isinstance(event.get('entries'), list):
                event_entries_paths.extend(event['entries'])
            
            for path in event_entries_paths:
                if path in scan_files_config:
                    scan_file_item_counts[os.path.basename(path)] += 1 # Use basename as key for counts
        
        total_items_from_scan_files = sum(scan_file_item_counts.values())
        start_data_row_scan, end_data_row_scan = 0, 0

        if total_items_from_scan_files > 0:
            start_data_row_scan = current_row
            sorted_scan_files = sorted(scan_file_item_counts.items(), key=lambda item: (-item[1], item[0]))
            for file_name, count in sorted_scan_files:
                percentage = (count / total_items_from_scan_files) * 100
                ws.cell(row=current_row, column=1, value=file_name)
                ws.cell(row=current_row, column=2, value=count)
                ws.cell(row=current_row, column=3, value=percentage).number_format = '0.00"%"'
                current_row += 1
            end_data_row_scan = current_row - 1

        # --- Add Charts --- 
        if total_items_in_week > 0 and start_data_row_cat <= end_data_row_cat:
            # Original Category Bar Chart
            bar_cat_orig = BarChart()
            bar_cat_orig.title = f"Items per Categorie (Week {week_num})"
            bar_cat_orig.x_axis.title = "Categorie"
            bar_cat_orig.y_axis.title = "Aantal Items"
            data_ref_cat_bar = Reference(ws, min_col=2, min_row=start_data_row_cat, max_row=end_data_row_cat)
            cats_ref_cat_bar = Reference(ws, min_col=1, min_row=start_data_row_cat, max_row=end_data_row_cat)
            bar_cat_orig.add_data(data_ref_cat_bar, titles_from_data=False)
            bar_cat_orig.set_categories(cats_ref_cat_bar)
            bar_cat_orig.legend = None
            bar_cat_orig.dataLabels = DataLabelList(); bar_cat_orig.dataLabels.showVal = True
            bar_cat_orig.width = 20; bar_cat_orig.height = 12
            ws.add_chart(bar_cat_orig, "E2")

            # New Category Pie Chart
            pie_cat = PieChart()
            pie_cat.title = f"Percentage per Categorie (Week {week_num})"
            labels_ref_cat_pie = Reference(ws, min_col=1, min_row=start_data_row_cat, max_row=end_data_row_cat)
            data_ref_cat_pie = Reference(ws, min_col=2, min_row=start_data_row_cat, max_row=end_data_row_cat)
            pie_cat.add_data(data_ref_cat_pie, titles_from_data=False)
            pie_cat.set_categories(labels_ref_cat_pie)
            pie_cat.dataLabels = DataLabelList(); pie_cat.dataLabels.showPercent = True; pie_cat.dataLabels.showVal = False
            pie_cat.width = 15; pie_cat.height = 10 # Slightly smaller
            ws.add_chart(pie_cat, "E18")

        if start_data_row_daily <= end_data_row_daily: # Check if daily data exists
            # Items per Day per Worked Hour Bar Chart
            bar_daily = BarChart()
            bar_daily.title = "Items per Gewerkt Uur per Dag"
            bar_daily.x_axis.title = "Dag van de Week"
            bar_daily.y_axis.title = "Items / Gewerkt Uur"
            labels_ref_daily = Reference(ws, min_col=1, min_row=start_data_row_daily, max_row=end_data_row_daily)
            data_ref_daily = Reference(ws, min_col=4, min_row=start_data_row_daily, max_row=end_data_row_daily)
            bar_daily.add_data(data_ref_daily, titles_from_data=False)
            bar_daily.set_categories(labels_ref_daily)
            bar_daily.legend = None
            bar_daily.dataLabels = DataLabelList(); bar_daily.dataLabels.showVal = True
            bar_daily.width = 20; bar_daily.height = 12
            ws.add_chart(bar_daily, "M2")

        if total_items_from_scan_files > 0 and start_data_row_scan <= end_data_row_scan:
            # Scan Files Usage Pie Chart
            pie_scan = PieChart()
            pie_scan.title = "Percentage Gebruik per Scan Bestand"
            labels_ref_scan = Reference(ws, min_col=1, min_row=start_data_row_scan, max_row=end_data_row_scan)
            data_ref_scan = Reference(ws, min_col=2, min_row=start_data_row_scan, max_row=end_data_row_scan)
            pie_scan.add_data(data_ref_scan, titles_from_data=False)
            pie_scan.set_categories(labels_ref_scan)
            pie_scan.dataLabels = DataLabelList(); pie_scan.dataLabels.showPercent = True; pie_scan.dataLabels.showVal = False
            pie_scan.width = 15; pie_scan.height = 10
            ws.add_chart(pie_scan, "M18")

        # --- Save Workbook --- 
        try:
            print(f"[{datetime.datetime.now()}] [WEEKLY_REPORT] Attempting to save: {report_path}")
            wb.save(report_path)
            print(f"[{datetime.datetime.now()}] [WEEKLY_REPORT_SUCCESS] Weekly report saved: {report_path}")
        except Exception as e:
            print(f"[{datetime.datetime.now()}] [WEEKLY_REPORT_ERROR] Failed to save weekly report {report_path}: {e}")
            traceback.print_exc()
        finally:
            if 'wb' in locals() and wb is not None:
                try: wb.close()
                except Exception as e_close: print(f"[{datetime.datetime.now()}] [WEEKLY_REPORT_ERROR] Error closing workbook: {e_close}")
            gc.collect()
 # Suggest garbage collection

def generate_yearly_report(events):
    """
    Generates a yearly report for the current year.
    The report includes a summary of events per category and a bar chart.
    """
    with REPORT_GENERATION_LOCK:
        now_dt = datetime.datetime.now()
        current_year = now_dt.year

        # --- 1. Filter events for the current year ---
        year_events = [
            event for event in events
            if datetime.datetime.fromisoformat(event['timestamp']).year == current_year
        ]

        if not year_events:
            print(f"[{datetime.datetime.now()}] [YEARLY_REPORT] No events found for {current_year}. Report not generated.")
            return

        # --- 2. Prepare folder and filename ---
        report_dir = os.path.join(EXCEL_REPORT_DIR, 'Jaar Raporten')
        try:
            os.makedirs(report_dir, exist_ok=True)
        except OSError as e:
            print(f"[{datetime.datetime.now()}] [YEARLY_REPORT_ERROR] Could not create report directory {report_dir}: {e}")
            return
        
        report_path = os.path.join(report_dir, f"Jaar_{current_year}.xlsx")

        # --- 3. Aggregate data by category ---
        category_counts = defaultdict(int)
        for event in year_events:
            category = event.get('category', 'Allerlei') # Default category
            if event['type'] == 'new':
                category_counts[category] += 1
            elif event['type'] == 'duplicate':
                entries = event.get('entries', [])
                if isinstance(entries, list):
                    category_counts[category] += len(entries)
                else:
                    category_counts[category] += 1 # Count as one if not a list

        # --- 4. Create Workbook and Worksheet ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Jaar {current_year}"

        # --- 5. Create Data Table ---
        ws.append(["Categorie", "Aantal", "Percentage"])
        total_events = sum(category_counts.values())
        
        sorted_categories = sorted(category_counts.items(), key=lambda item: (-item[1], item[0]))

        for category, count in sorted_categories:
            percentage = (count / total_events) * 100 if total_events > 0 else 0
            ws.append([category, count, percentage])
            ws.cell(row=ws.max_row, column=3).number_format = '0.00"%"'

        # --- 6. Create Bar Chart for Category Totals ---
        if ws.max_row > 1:
            chart = BarChart()
            chart.title = f"Meldingen per Categorie (Jaar {current_year})"
            chart.x_axis.title = "Categorie"
            chart.y_axis.title = "Aantal Meldingen"
            chart.legend = None

            data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)

            chart.style = 10
            chart.width = 20
            chart.height = 12

            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True

            ws.add_chart(chart, "E2")

        # --- 7. Save the Workbook ---
        try:
            print(f"[{datetime.datetime.now()}] [YEARLY_REPORT] Attempting to save: {report_path}")
            wb.save(report_path)
            print(f"[{datetime.datetime.now()}] [YEARLY_REPORT_SUCCESS] Yearly report saved: {report_path}")
        except Exception as e:
            print(f"[{datetime.datetime.now()}] [YEARLY_REPORT_ERROR] Failed to save yearly report {report_path}: {e}")
            traceback.print_exc()
        finally:
            if 'wb' in locals() and wb is not None:
                try:
                    wb.close()
                except Exception as e_close:
                    print(f"[{datetime.datetime.now()}] [YEARLY_REPORT_ERROR] Error closing workbook: {e_close}")
            gc.collect()

def update_reports(events):
    try:
        print(f"[{datetime.datetime.now()}] [DEBUG] update_reports called with {len(events)} events.")
        generate_weekly_report(events)
        generate_monthly_report(events)
        generate_yearly_report(events)
    except Exception as e:
        print(f"[{datetime.datetime.now()}] [ERROR] Exception in update_reports: {e}")
        traceback.print_exc()

def main(shutdown_event):
    global RUNTIME_SEEN_ENTRIES, RUNTIME_CATEGORY_TOTALS, RUNTIME_DAILY_ITEM_COUNTS
    print(f"[{datetime.datetime.now()}] [DEBUG_MAIN_START] main() started. id(RUNTIME_DAILY_ITEM_COUNTS): {id(RUNTIME_DAILY_ITEM_COUNTS)}")
    """Main function for the Hops File Logger."""
    print(f'[{datetime.datetime.now()}] [STARTUP] Hops File Logger starting...')

    # --- Initial State Loading ---
    event_history = load_event_history()
    seen_data = load_seen()
    RUNTIME_SEEN_ENTRIES = seen_data.get('entries', {})

    # Initialize runtime stats from historical events
    for event in event_history:
        try:
            event_dt = datetime.datetime.fromisoformat(event['timestamp'])
            event_date = event_dt.date()
            category = event.get('category', 'Allerlei')
            # --- DEBUG PRINTS START (INIT LOOP) ---
            # print(f"[{datetime.datetime.now()}] [DEBUG_HFL_INIT] Before update. event_date: {event_date}, hour: {event_dt.hour}")
            # print(f"[{datetime.datetime.now()}] [DEBUG_HFL_INIT] RUNTIME_DAILY_ITEM_COUNTS[{event_date}] current type: {type(RUNTIME_DAILY_ITEM_COUNTS.get(event_date))}, value: {RUNTIME_DAILY_ITEM_COUNTS.get(event_date)}")
            # --- DEBUG PRINTS END (INIT LOOP) ---
            RUNTIME_DAILY_ITEM_COUNTS[event_date][event_dt.hour] += 1
            RUNTIME_CATEGORY_TOTALS[category] += 1
            # --- DEBUG PRINTS START (INIT LOOP AFTER) ---
            # print(f"[{datetime.datetime.now()}] [DEBUG_HFL_INIT] After update. RUNTIME_DAILY_ITEM_COUNTS[{event_date}] new type: {type(RUNTIME_DAILY_ITEM_COUNTS.get(event_date))}, value: {RUNTIME_DAILY_ITEM_COUNTS.get(event_date)}")
            # print(f"[{datetime.datetime.now()}] [DEBUG_HFL_INIT] RUNTIME_DAILY_ITEM_COUNTS[{event_date}][{event_dt.hour}] new value: {RUNTIME_DAILY_ITEM_COUNTS.get(event_date, {}).get(event_dt.hour)}")
            # --- DEBUG PRINTS END (INIT LOOP AFTER) ---
        except (ValueError, TypeError):
            continue # Skip events with bad timestamps or structure
    print(f"[{datetime.datetime.now()}] [STARTUP] Initialized runtime stats from {len(event_history)} historical events.")

    # Process any manual events that were saved while the logger was offline
    print(f"[{datetime.datetime.now()}] [STARTUP] Processing any pending manual events from last session...")
    manual_events_processed = process_runtime_manual_events(event_history)

    # --- Main Loop ---
    print(f'[{datetime.datetime.now()}] [MAIN_LOOP] Starting main loop. Press Ctrl+C to exit.')
    while not shutdown_event.is_set():
        try:
            new_entries_found = False
            scan_files = CONFIG.get('scan_files', [])

            for path in scan_files:
                if not os.path.exists(path):
                    continue
                
                try:
                    mtime = os.path.getmtime(path)
                    if path not in RUNTIME_SEEN_ENTRIES or mtime > RUNTIME_SEEN_ENTRIES[path]:
                        print(f"[{datetime.datetime.now()}] [NEW_ENTRY] New entry detected in: {path}")
                        category = categorize(path)
                        event_time = datetime.datetime.now()
                        event_date = event_time.date()
                        
                        new_event = {
                            'timestamp': event_time.isoformat(),
                            'type': 'new',
                            'entry': f'file_update_{os.path.basename(path)}',
                            'source_file': path,
                            'category': category
                        }
                        
                        event_history.append(new_event)
                        append_event_history([new_event])
                        
                        RUNTIME_SEEN_ENTRIES[path] = mtime
                        RUNTIME_CATEGORY_TOTALS[category] += 1
                        new_entries_found = True
                        # --- DEBUG PRINTS START (SCAN LOOP) ---
                        print(f"[{datetime.datetime.now()}] [DEBUG_HFL_SCAN] Before update. event_date: {event_date}, hour: {event_time.hour}")
                        print(f"[{datetime.datetime.now()}] [DEBUG_HFL_SCAN] RUNTIME_DAILY_ITEM_COUNTS[{event_date}] current type: {type(RUNTIME_DAILY_ITEM_COUNTS.get(event_date))}, value: {RUNTIME_DAILY_ITEM_COUNTS.get(event_date)}")
                        # --- DEBUG PRINTS END (SCAN LOOP) ---
                        RUNTIME_DAILY_ITEM_COUNTS[event_date][event_time.hour] += 1
                        # --- DEBUG PRINTS START (SCAN LOOP AFTER) ---
                        print(f"[{datetime.datetime.now()}] [DEBUG_HFL_SCAN] After update. RUNTIME_DAILY_ITEM_COUNTS[{event_date}] new type: {type(RUNTIME_DAILY_ITEM_COUNTS.get(event_date))}, value: {RUNTIME_DAILY_ITEM_COUNTS.get(event_date)}")
                        print(f"[{datetime.datetime.now()}] [DEBUG_HFL_SCAN] RUNTIME_DAILY_ITEM_COUNTS[{event_date}][{event_time.hour}] new value: {RUNTIME_DAILY_ITEM_COUNTS.get(event_date, {}).get(event_time.hour)}")
                        # --- DEBUG PRINTS END (SCAN LOOP AFTER) ---
                        new_entries_found = True

                except OSError as e:
                    print(f"[{datetime.datetime.now()}] [FILE_ERROR] Error accessing file {path}: {e}")

            if new_entries_found:
                print(f"[{datetime.datetime.now()}] [REPORTS] New entries found, updating reports...")
                update_reports(event_history)

            # Process any manual events that have been added
            manual_events_found = process_runtime_manual_events(event_history)
            if manual_events_found:
                new_entries_found = True

            # Save the updated seen entries to disk
            save_seen(RUNTIME_SEEN_ENTRIES)

            # If any new events were found in this cycle, update all reports
            if new_entries_found:
                print(f"[{datetime.datetime.now()}] [REPORT_TRIGGER] New events detected, updating all reports...")
                update_reports(event_history)

        except Exception as e:
            print(f'[{datetime.datetime.now()}] [ERROR] An error occurred in the main loop: {e}')
            traceback.print_exc() # Print full traceback for debugging
        
        # Wait for the next cycle
        shutdown_event.wait(5) # Wait a bit before retrying

    print(f'[{datetime.datetime.now()}] [SHUTDOWN] Main loop exited. Finalizing...')
    print(f"[{datetime.datetime.now()}] [SHUTDOWN_STATE] RUNTIME_SEEN_ENTRIES before main exits: {RUNTIME_SEEN_ENTRIES}")
    # The atexit handler will save the final state.

print(f"[{datetime.datetime.now()}] DEBUG: About to enter __main__ block.")
if __name__ == '__main__':
    # This is the main entry point for the logger application.

    # atexit.register is called early to ensure cleanup runs even if initialization fails.
    atexit.register(clear_seen_entries_but_keep_last_on_exit)

    # The shutdown_event is a global threading event used to signal graceful shutdown.
    # The signal_handler will set it, and the main loop will check it.

    def setup_signal_handling(shutdown_event_param):
        """
        A handler for system signals (like Ctrl+C) to ensure graceful shutdown.
        It sets the shutdown_event, which the main loop checks.
        """
        def signal_handler(signum, frame):
            print(f"[{datetime.datetime.now()}] [INFO] Shutdown signal ({signal.Signals(signum).name}) received. Initiating graceful shutdown...")
            shutdown_event_param.set() # Use the event passed to the outer function

        # Register signal handlers for SIGINT (Ctrl+C) and SIGTERM.
        # These will call signal_handler, which sets shutdown_event_param.
        signal.signal(signal.SIGINT, signal_handler)
        signal.signal(signal.SIGTERM, signal_handler)

    # Create the shutdown event that will be used by main and signal handling
    shutdown_event = threading.Event()

    setup_signal_handling(shutdown_event) # Pass the event

    try:
        # Start the main application logic.
        main(shutdown_event) # Pass the event
    except Exception as e:
        # Catch any critical, unhandled exceptions from the main function.
        print(f"[{datetime.datetime.now()}] [CRITICAL] An unhandled exception occurred in main: {e}")
        print(traceback.format_exc())
    finally:
        # This block will run regardless of whether an exception occurred.
        print(f"[{datetime.datetime.now()}] [INFO] Application shutdown complete.")
