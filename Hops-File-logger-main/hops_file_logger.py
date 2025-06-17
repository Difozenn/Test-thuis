import os
import time
import json
import atexit

from datetime import datetime, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.chart.series import SeriesLabel

# Load config
with open('logger_config.json', 'r') as f:
    CONFIG = json.load(f)

SEEN_ENTRIES_FILE = CONFIG['seen_entries_file']
CATEGORIES_FILE = CONFIG['categories_file']
EXCEL_REPORT_DIR = CONFIG['excel_report_dir'].replace('reports', 'Realtime Raport')
SCAN_INTERVAL = CONFIG['scan_interval_seconds']

# Files to scan are now loaded from config.json


def clear_seen_entries_but_keep_last_on_exit():
    seen_file = SEEN_ENTRIES_FILE
    scan_files = CONFIG['scan_files']
    try:
        with open(seen_file, 'r') as f:
            seen_entries = json.load(f)
    except Exception:
        seen_entries = {}

    new_seen_entries = {}
    for path in scan_files:
        entries = seen_entries.get(path, [])
        if isinstance(entries, list) and entries:
            new_seen_entries[path] = [entries[-1]]
        elif entries:
            new_seen_entries[path] = [entries]
        else:
            new_seen_entries[path] = []

    with open(seen_file, 'w') as f:
        json.dump(new_seen_entries, f, indent=4)

# Register the cleanup function to run on program exit
atexit.register(clear_seen_entries_but_keep_last_on_exit)

# Ensure report directory exists
os.makedirs(EXCEL_REPORT_DIR, exist_ok=True)

# Load categories
with open(CATEGORIES_FILE, 'r') as f:
    CATEGORIES = json.load(f)

def categorize(path):
    # Only categorize by file content
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            print(f'[DEBUG] Content length: {len(content)}')
            sorted_categories = sorted(CATEGORIES.items(), key=lambda x: -len(x[0]))
            for prefix, category in sorted_categories:
                print(f'[DEBUG] Checking prefix: {repr(prefix)}')
                print(f'[DEBUG] prefix in content: {prefix.lower() in content.lower()}')
                if prefix.lower() in content.lower():
                    print(f'[DEBUG] Match found: {repr(prefix)} in {path}')
                    return category
    except Exception as e:
        print(f'[ERROR] Could not read file {path} for categorization: {e}')
    return 'Allerlei'

def load_seen():
    if os.path.exists(SEEN_ENTRIES_FILE):
        with open(SEEN_ENTRIES_FILE, 'r') as f:
            return json.load(f)
    else:
        return {'entries': {}, 'last_mod_time': None}  # No save_events, history is now separate

def save_seen(data):
    with open(SEEN_ENTRIES_FILE, 'w') as f:
        json.dump(data, f)


def now():
    return datetime.now()

def period_key(dt, period):
    if period == 'hour':
        return dt.strftime('%Y-%m-%d %H:00')
    elif period == 'week':
        return f"{dt.year}-W{dt.isocalendar()[1]:02d}"
    elif period == 'month':
        return dt.strftime('%Y-%m')
    elif period == 'year':
        return str(dt.year)

from openpyxl.chart import BarChart, Reference

# --- New function for weekly report ---

def generate_monthly_report(events):
    """
    Genereer een maandrapport als 'Maand_<YYYY-MM>.xlsx' in de 'Maand' map.
    """
    now_dt = datetime.now()
    year = now_dt.year
    month = now_dt.month
    # Filter events for current month
    maand_events = []
    for event in events:
        dt = datetime.fromisoformat(event['timestamp'])
        if dt.year == year and dt.month == month:
            maand_events.append(event)
    # Aggregate by category
    cat_counts = defaultdict(int)
    for event in maand_events:
        # For all 'new' events, always use the provided 'category' field (works for both file and manual entries)
        if event['type'] == 'new':
            cat = event.get('category', 'Allerlei')
            cat_counts[cat] += 1
        # For 'duplicate' events, categorize each entry (assumes entry is a file path)
        elif event['type'] == 'duplicate':
            for entry in event.get('entries', []):
                cat = categorize(entry)
                cat_counts[cat] += 1
    # Prepare folder and filename
    maand_folder = os.path.join(EXCEL_REPORT_DIR, 'Maand Raporten')
    os.makedirs(maand_folder, exist_ok=True)
    filename = f"Maand_{year}-{month:02d}.xlsx"
    report_path = os.path.join(maand_folder, filename)
    # Write Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Maand {year}-{month:02d}"
    ws.append(['Categorie', 'Aantal', 'Percentage', 'Aantal (Percentage)'])
    total = sum(cat_counts.values())
    for cat, count in cat_counts.items():
        perc = (count / total) * 100 if total > 0 else 0
        ws.append([cat, count, perc])
    # Add a bar chart if there is data
    if ws.max_row > 1:
        chart = BarChart()
        chart.title = f'Aantallen per categorie (Maand)'
        chart.x_axis.title = 'Categorie'
        chart.y_axis.title = 'Aantal'
        data_ref = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row, max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.series[0].title = SeriesLabel()
        chart.width = 21  # centimeters (A4 width)
        chart.height = 14  # centimeters (half A4 height, landscape)
        from openpyxl.chart.label import DataLabelList
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showPercent = True
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        chart.dataLabels.showLegendKey = False
        ws.add_chart(chart, f'E2')
        # Add a pie chart for percentages only, placed at E29 (second chart position)
        from openpyxl.chart import PieChart
        pie = PieChart()
        pie.title = "Verdeling per categorie (%)"
        data_ref_pie = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row, max_col=3)  # Percentage column
        cats_ref_pie = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        pie.add_data(data_ref_pie, titles_from_data=True)
        pie.set_categories(cats_ref_pie)
        pie.width = 21
        pie.height = 14
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = False
        pie.dataLabels.showCatName = False
        ws.add_chart(pie, "E29")  # Place pie chart at standard second chart position

    wb.save(report_path)

def generate_weekly_report(events):
    """
    Generate a weekly report (Monday-Friday) as 'Week_#<number>.xlsx' in 'Week Raport' folder.
    """
    # Get current ISO week and year
    now_dt = datetime.now()
    week_num = now_dt.isocalendar()[1]
    year = now_dt.isocalendar()[0]
    # Determine Monday and Friday of current week
    monday = now_dt - timedelta(days=now_dt.weekday())
    friday = monday + timedelta(days=4)
    # Filter events for current week, only Monday-Friday
    week_events = []
    for event in events:
        dt = datetime.fromisoformat(event['timestamp'])
        iso_year, iso_week, iso_weekday = dt.isocalendar()
        if iso_year == year and iso_week == week_num and 1 <= iso_weekday <= 5:
            week_events.append(event)
    # Aggregate by category
    cat_counts = defaultdict(int)
    for event in week_events:
        if event['type'] == 'new':
            cat = event['category']
            cat_counts[cat] += 1
        elif event['type'] == 'duplicate':
            for entry in event.get('entries', []):
                cat = categorize(entry)
                cat_counts[cat] += 1
    # Prepare folder and filename
    week_folder = os.path.join(EXCEL_REPORT_DIR, 'Week Raporten')
    os.makedirs(week_folder, exist_ok=True)
    filename = f"Week_{week_num}.xlsx"
    report_path = os.path.join(week_folder, filename)
    # Write Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Week {week_num}"
    ws.append(['Category', 'Count', 'Percentage'])
    total = sum(cat_counts.values())
    for cat, count in cat_counts.items():
        perc = (count / total) * 100 if total > 0 else 0
        ws.append([cat, count, perc])
    # Add a bar chart if there is data
    if ws.max_row > 1:
        chart = BarChart()
        chart.title = f'Aantallen per categorie (Ma-Vr)'
        chart.x_axis.title = 'Categorie'
        chart.y_axis.title = 'Aantal'
        data_ref = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row, max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.series[0].title = SeriesLabel()
        from openpyxl.chart.label import DataLabelList
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showPercent = True
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        chart.dataLabels.showLegendKey = False
        # Set chart size for A4 fit
        chart.width = 21  # A4 width in cm
        chart.height = 14  # A4 height in cm
        ws.add_chart(chart, f'E2')  # Top chart
        # Pie chart will be added after both bar charts (see below)
    # --- Second Bar Chart: Items per Workday per Hour ---
    # Prepare data: {day: [count, hours]}
    day_counts = {i: 0 for i in range(1, 6)}  # 1=Mon, ..., 5=Fri
    for event in week_events:
        dt = datetime.fromisoformat(event['timestamp'])
        iso_weekday = dt.isocalendar()[2]
        if 1 <= iso_weekday <= 5:
            # For all 'new' events, always use the provided 'category' field (manual entries included)
            if event['type'] == 'new':
                day_counts[iso_weekday] += 1
            elif event['type'] == 'duplicate':
                day_counts[iso_weekday] += len(event.get('entries', []))
    # Hours worked per day
    hours_per_day = {1: 8, 2: 8, 3: 8, 4: 8, 5: 7}
    # Prepare rows for the chart
    ws.append([])  # Blank row
    ws.append(["Werkdag", "Geregistreerd", "Gewerkte uren", "Per uur"])
    day_names = {1: "Maandag", 2: "Dinsdag", 3: "Woensdag", 4: "Donderdag", 5: "Vrijdag"}
    start_row = ws.max_row + 1
    per_hour_values = []
    for i in range(1, 6):
        per_hour = day_counts[i] / hours_per_day[i] if hours_per_day[i] else 0
        ws.append([day_names[i], day_counts[i], hours_per_day[i], per_hour])
        per_hour_values.append(per_hour)
    # Add average row
    gemiddelde = sum(per_hour_values) / len(per_hour_values) if per_hour_values else 0
    ws.append(["Gemiddelde", "", "", gemiddelde])
    end_row = ws.max_row
    # Add bar chart for items per hour per day
    if end_row > start_row:
        chart2 = BarChart()
        chart2.title = "Aantal items per dag, per gewerkt uur (Ma-Vr)"
        chart2.x_axis.title = "Werkdag"
        chart2.y_axis.title = "Aantal per uur"
        data_ref2 = Reference(ws, min_col=4, min_row=start_row-1, max_row=end_row-1, max_col=4)
        cats_ref2 = Reference(ws, min_col=1, min_row=start_row, max_row=end_row-1)
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cats_ref2)
        # Add average as a line
        from openpyxl.chart.series import Series
        from openpyxl.chart.line_chart import LineChart
        line = LineChart()
        line.title = "Gemiddelde"
        avg_ref = Reference(ws, min_col=4, min_row=end_row, max_row=end_row)
        line.add_data(avg_ref, titles_from_data=False)
        # line.y_axis.axId = 200  # Removed to prevent secondary Y-axis
        chart2 += line
        # Set chart size for A4 fit
        chart2.width = 21  # A4 width in cm
        chart2.height = 14  # A4 height in cm
        # Show value (amount) on top of each bar
        from openpyxl.chart.label import DataLabelList
        chart2.dataLabels = DataLabelList()
        chart2.dataLabels.showVal = True
        chart2.dataLabels.showCatName = False
        chart2.dataLabels.showSerName = False
        ws.add_chart(chart2, f'E29')  # Place second chart at E29
        # Add a pie chart for percentages only, placed below both bar charts
        from openpyxl.chart import PieChart
        pie = PieChart()
        pie.title = "Verdeling per categorie (%)"
        # Only use the C1 (category) summary rows; these are at the top, before the first blank row
        # Find the first blank row after the summary
        summary_end = 2
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value is None or row[0].value == "":
                break
            summary_end += 1
        # summary_end is now the first row after the summary
        data_ref_pie = Reference(ws, min_col=3, min_row=1, max_row=summary_end-1, max_col=3)  # Percentage column
        cats_ref_pie = Reference(ws, min_col=1, min_row=2, max_row=summary_end-1)
        pie.add_data(data_ref_pie, titles_from_data=True)
        pie.set_categories(cats_ref_pie)
        pie.width = 21
        pie.height = 14
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = False
        pie.dataLabels.showCatName = False
        ws.add_chart(pie, "E56")  # Place third chart at E56

    wb.save(report_path)

def update_reports(events):
    try:
        print(f"[DEBUG] update_reports called with {len(events)} events.")
        generate_weekly_report(events)
        generate_monthly_report(events)
        generate_yearly_report(events)
    except Exception as e:
        print(f"[ERROR] Exception in update_reports: {e}")


# --- New function for yearly report ---
def generate_yearly_report(events):
    """
    Generate a yearly report Excel file in 'Jaar Raporten' folder as '<year>.xlsx'.
    Sheet 1: 'Weekly Report' - Bar chart: x=week numbers, y=total items per week.
    Sheet 2: 'Monthly Report' - Bar chart: x=months, y=total items per month.
    """
    now_dt = datetime.now()
    year = now_dt.year
    jaar_folder = os.path.join(EXCEL_REPORT_DIR, 'Jaar Raporten')
    os.makedirs(jaar_folder, exist_ok=True)
    filename = f"{year}.xlsx"
    report_path = os.path.join(jaar_folder, filename)

    # Filter events for current year
    jaar_events = [e for e in events if datetime.fromisoformat(e['timestamp']).year == year]

    # --- Weekly aggregation ---
    week_totals = defaultdict(int)  # week_num -> total
    for event in jaar_events:
        dt = datetime.fromisoformat(event['timestamp'])
        if event['type'] == 'new':
            week_num = dt.isocalendar()[1]
            week_totals[week_num] += 1
        elif event['type'] == 'duplicate':
            week_num = dt.isocalendar()[1]
            week_totals[week_num] += len(event.get('entries', []))

    # --- Monthly aggregation ---
    month_totals = defaultdict(int)  # month_num -> total
    for event in jaar_events:
        dt = datetime.fromisoformat(event['timestamp'])
        if event['type'] == 'new':
            month_totals[dt.month] += 1
        elif event['type'] == 'duplicate':
            month_totals[dt.month] += len(event.get('entries', []))

    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Weekly Report Sheet ---
    ws_week = wb.create_sheet('Wekelijks rapport')
    ws_week.append(['Week', 'Total'])
    for week in range(1, 54):
        ws_week.append([week, week_totals.get(week, 0)])
    if ws_week.max_row > 1:
        chart = BarChart()
        chart.title = 'Aantal items per week'
        chart.x_axis.title = 'Week'
        chart.y_axis.title = 'Aantal'
        data_ref = Reference(ws_week, min_col=2, min_row=2, max_row=ws_week.max_row, max_col=2)
        cats_ref = Reference(ws_week, min_col=1, min_row=2, max_row=ws_week.max_row)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.width = 21  # A4 width in cm
        chart.height = 14  # A4 height in cm
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        ws_week.add_chart(chart, 'E2')  # First chart


    # --- Monthly Report Sheet ---
    ws_month = wb.create_sheet('Maandelijks rapport')
    ws_month.append(['Month', 'Total'])
    for month in range(1, 13):
        ws_month.append([month, month_totals.get(month, 0)])
    if ws_month.max_row > 1:
        chart = BarChart()
        chart.title = 'Aantal items per maand'
        chart.x_axis.title = 'Maand'
        chart.y_axis.title = 'Aantal'
        data_ref = Reference(ws_month, min_col=2, min_row=2, max_row=ws_month.max_row, max_col=2)
        cats_ref = Reference(ws_month, min_col=1, min_row=2, max_row=ws_month.max_row)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.width = 21  # A4 width in cm
        chart.height = 14  # A4 height in cm
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        ws_month.add_chart(chart, 'E2')  # First chart


    # --- Yearly Report Sheet ---
    ws_year = wb.create_sheet('Jaaroverzicht')
    ws_year.append(['Category', 'Total', 'Percentage'])
    # Aggregate by category
    cat_totals = defaultdict(int)
    for event in jaar_events:
        if event['type'] == 'new':
            cat = event.get('category', 'Allerlei')
            cat_totals[cat] += 1
        elif event['type'] == 'duplicate':
            for entry in event.get('entries', []):
                cat = categorize(entry)
                cat_totals[cat] += 1
    total_year = sum(cat_totals.values())
    for cat, count in cat_totals.items():
        perc = (count / total_year) * 100 if total_year > 0 else 0
        ws_year.append([cat, count, perc])
    # Bar chart: totals by category
    if ws_year.max_row > 1:
        from openpyxl.chart import BarChart, Reference, PieChart
        from openpyxl.chart.label import DataLabelList
        chart = BarChart()
        chart.title = 'Aantal items per categorie (Jaar)'
        chart.x_axis.title = 'Categorie'
        chart.y_axis.title = 'Aantal'
        data_ref = Reference(ws_year, min_col=2, min_row=2, max_row=ws_year.max_row, max_col=2)
        cats_ref = Reference(ws_year, min_col=1, min_row=2, max_row=ws_year.max_row)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.width = 18
        chart.height = 10
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showCatName = False
        chart.dataLabels.showSerName = False
        chart.width = 21  # A4 width in cm
        chart.height = 14  # A4 height in cm
        ws_year.add_chart(chart, 'E2')  # First chart

        # Pie chart: percentages by category
        pie = PieChart()
        pie.title = "Verdeling per categorie (%)"
        data_ref_pie = Reference(ws_year, min_col=3, min_row=1, max_row=ws_year.max_row, max_col=3)
        cats_ref_pie = Reference(ws_year, min_col=1, min_row=2, max_row=ws_year.max_row)
        pie.add_data(data_ref_pie, titles_from_data=True)
        pie.set_categories(cats_ref_pie)
        pie.width = 21
        pie.height = 14
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = False
        pie.dataLabels.showCatName = False
        ws_year.add_chart(pie, "E56")  # Place third chart at E56
    wb.save(report_path)
    print(f"[DEBUG] Saved yearly report to {report_path}")


def archive_old_events():
    """
    Move events from previous years in events_history.json to archive files (events_history_<year>.json).
    Retain only current year's events in events_history.json.
    """
    path = os.path.join(os.path.dirname(__file__), 'events_history.json')
    if not os.path.exists(path):
        return
    try:
        with open(path, 'r', encoding='utf-8') as f:
            events = json.load(f)
        if not events:
            return
        current_year = datetime.now().year
        current_year_events = []
        archive_map = {}
        for event in events:
            try:
                event_year = datetime.fromisoformat(event['timestamp']).year
            except Exception:
                continue
            if event_year == current_year:
                current_year_events.append(event)
            else:
                archive_map.setdefault(event_year, []).append(event)
        # Write archive files
        for year, year_events in archive_map.items():
            archive_path = os.path.join(os.path.dirname(__file__), f'events_history_{year}.json')
            if os.path.exists(archive_path):
                # Append to existing archive
                with open(archive_path, 'r', encoding='utf-8') as f:
                    old = json.load(f)
                year_events = old + year_events
            with open(archive_path, 'w', encoding='utf-8') as f:
                json.dump(year_events, f, indent=2)
        # Write only current year's events back
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(current_year_events, f, indent=2)
    except Exception as e:
        print(f"[ERROR] Could not archive old events: {e}")

def load_event_history():
    archive_old_events()
    path = os.path.join(os.path.dirname(__file__), 'events_history.json')
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            try:
                return json.load(f)
            except Exception as e:
                print(f"[ERROR] Could not load events_history.json: {e}")
                return []
    return []

def append_event_history(new_events):
    path = os.path.join(os.path.dirname(__file__), 'events_history.json')
    try:
        events = []
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                events = json.load(f)
        events.extend(new_events)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(events, f, indent=2)
    except Exception as e:
        print(f"[ERROR] Could not append to events_history.json: {e}")

def main():
    seen = load_seen()
    entries = seen.get('entries', {})
    # Load persistent event history
    event_history = load_event_history()
    print('Starting logger...')
    while True:
        try:
            files = CONFIG['scan_files']
            new_entries = []
            for f in files:
                abs_f = os.path.abspath(f)
                mtime = os.path.getmtime(abs_f)
                last_seen = entries.get(abs_f)
                if last_seen is None or last_seen != mtime:
                    dt = now().isoformat()
                    cat = categorize(abs_f)
                    event = {'timestamp': dt, 'type': 'new', 'entry': abs_f, 'category': cat, 'mtime': mtime}
                    new_entries.append(event)
                    entries[abs_f] = mtime
            if not new_entries:
                print(f"[DEBUG] No new entries found. Reports not updated.")
                time.sleep(SCAN_INTERVAL)
                continue
            else:
                # Append new events to persistent event history
                append_event_history(new_entries)
                event_history.extend(new_entries)
                save_seen({'entries': entries})  # Only track entries, not save_events
                print(f"[DEBUG] Calling update_reports with {len(event_history)} events...")
                update_reports(event_history)
                time.sleep(SCAN_INTERVAL)
        except Exception as e:
            print(f'[ERROR] Logger main loop: {e}')
            time.sleep(SCAN_INTERVAL)

if __name__ == '__main__':
    main()
