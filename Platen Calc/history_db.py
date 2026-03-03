#!/usr/bin/env python3
"""
History DB - SQLite snapshot storage for bestelberekening exports.
Saves a snapshot on each export so the next export can show comparison columns.
"""

import csv
import sqlite3
from datetime import datetime
from pathlib import Path


def _to_iso(date_str):
    """Convert DD_MM_YYYY to YYYY-MM-DD for correct chronological sorting."""
    parts = date_str.split('_')
    if len(parts) == 3:
        return f"{parts[2]}-{parts[1]}-{parts[0]}"
    return date_str


class HistoryDB:
    """Minimal SQLite database for storing bestelberekening snapshots."""

    def __init__(self, db_path=None):
        if db_path is None:
            db_path = Path(__file__).parent / "bestelberekening_history.db"
        self.db_path = str(db_path)
        self.conn = sqlite3.connect(self.db_path)
        self.conn.execute("PRAGMA journal_mode=WAL")
        self._create_tables()

    def _create_tables(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                snapshot_date TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
            );

            CREATE TABLE IF NOT EXISTS snapshot_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                snapshot_id INTEGER NOT NULL,
                material TEXT NOT NULL,
                artikel_nummer TEXT,
                materiaal_id TEXT,
                stock_m2 REAL NOT NULL DEFAULT 0,
                saldo_m2 REAL NOT NULL DEFAULT 0,
                bruto_m2 REAL NOT NULL DEFAULT 0,
                FOREIGN KEY (snapshot_id) REFERENCES snapshots(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_snapshot_rows_snapshot
                ON snapshot_rows(snapshot_id);

            CREATE TABLE IF NOT EXISTS handmagazijn_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                snapshot_date TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
            );

            CREATE TABLE IF NOT EXISTS handmagazijn_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                snapshot_id INTEGER NOT NULL,
                material TEXT NOT NULL,
                m2 REAL NOT NULL DEFAULT 0,
                FOREIGN KEY (snapshot_id) REFERENCES handmagazijn_snapshots(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_handmagazijn_rows_snapshot
                ON handmagazijn_rows(snapshot_id);
        """)
        # Migrate existing DBs: add bruto_m2 column if missing
        try:
            self.conn.execute("ALTER TABLE snapshot_rows ADD COLUMN bruto_m2 REAL NOT NULL DEFAULT 0")
            self.conn.commit()
        except Exception:
            pass  # Column already exists
        self.conn.commit()

    def save_snapshot(self, date_str, results):
        """Save (or replace) a snapshot for the given date.

        Args:
            date_str: Date string like "11_02_2026" (DD_MM_YYYY).
            results: List of dicts with keys: material, artikel_nummer,
                     materiaal_id, stock, bestellen (saldo).
        """
        iso_date = _to_iso(date_str)
        cur = self.conn.cursor()

        # Delete existing snapshot for this date (upsert)
        cur.execute("SELECT id FROM snapshots WHERE snapshot_date = ?", (iso_date,))
        row = cur.fetchone()
        if row:
            cur.execute("DELETE FROM snapshot_rows WHERE snapshot_id = ?", (row[0],))
            cur.execute("DELETE FROM snapshots WHERE id = ?", (row[0],))

        cur.execute(
            "INSERT INTO snapshots (snapshot_date) VALUES (?)",
            (iso_date,)
        )
        snapshot_id = cur.lastrowid

        for r in results:
            cur.execute(
                """INSERT INTO snapshot_rows
                   (snapshot_id, material, artikel_nummer, materiaal_id, stock_m2, saldo_m2, bruto_m2)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    snapshot_id,
                    r['material'],
                    r.get('artikel_nummer', ''),
                    r.get('materiaal_id', ''),
                    r.get('stock', 0),
                    r.get('bestellen', 0),  # saldo
                    r.get('bruto', 0),
                )
            )

        self.conn.commit()

    def get_previous_snapshot(self):
        """Return the most recent snapshot as {material: {'stock': x, 'saldo': y}}.

        Returns None if no snapshots exist.
        """
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id FROM snapshots ORDER BY snapshot_date DESC LIMIT 1"
        )
        row = cur.fetchone()
        if not row:
            return None

        snapshot_id = row[0]
        cur.execute(
            "SELECT material, stock_m2, saldo_m2, bruto_m2 FROM snapshot_rows WHERE snapshot_id = ?",
            (snapshot_id,)
        )

        data = {}
        for material, stock, saldo, bruto in cur.fetchall():
            data[material] = {'stock': stock, 'saldo': saldo, 'bruto': bruto}
        return data

    def get_snapshot_count(self):
        """Return the total number of stored snapshots."""
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM snapshots")
        return cur.fetchone()[0]

    def import_xlsx(self, filepath):
        """Import a bestelberekening_DD_MM_YYYY.xlsx file into the database.

        Reads Materiaal (col A), Stock (col G), and Saldo (col I).
        Extracts the date from the filename.
        Returns the date string on success, None if skipped.
        """
        import re
        import openpyxl

        basename = Path(filepath).name
        match = re.search(r'bestelberekening_(\d{2}_\d{2}_\d{4})', basename)
        if not match:
            return None

        date_str = match.group(1)
        iso_date = _to_iso(date_str)

        # Skip if already in DB
        cur = self.conn.cursor()
        cur.execute("SELECT id FROM snapshots WHERE snapshot_date = ?", (iso_date,))
        if cur.fetchone():
            return None

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb["Bestelberekening"] if "Bestelberekening" in wb.sheetnames else wb.worksheets[0]

        results = []
        col_map = None
        for row in ws.iter_rows(min_col=1, max_col=9, values_only=True):
            if col_map is None:
                # Detect column layout from header row
                headers = [str(c).strip() if c else '' for c in row]
                col_map = {}
                for idx, h in enumerate(headers):
                    hl = h.lower()
                    if 'materiaal' in hl and 'id' not in hl:
                        col_map['material'] = idx
                    elif 'artikel' in hl:
                        col_map['artikel'] = idx
                    elif 'bruto' in hl:
                        col_map['bruto'] = idx
                    elif 'stock' in hl:
                        col_map['stock'] = idx
                    elif 'saldo' in hl:
                        col_map['saldo'] = idx
                continue

            material_idx = col_map.get('material', 0)
            artikel_idx = col_map.get('artikel')
            bruto_idx = col_map.get('bruto')
            stock_idx = col_map.get('stock')
            saldo_idx = col_map.get('saldo')

            material = row[material_idx] if material_idx is not None and material_idx < len(row) else None
            stock = row[stock_idx] if stock_idx is not None and stock_idx < len(row) else None
            saldo = row[saldo_idx] if saldo_idx is not None and saldo_idx < len(row) else None
            bruto = row[bruto_idx] if bruto_idx is not None and bruto_idx < len(row) else None
            artikel = row[artikel_idx] if artikel_idx is not None and artikel_idx < len(row) else ''

            if material and stock is not None and saldo is not None:
                try:
                    results.append({
                        'material': str(material),
                        'artikel_nummer': str(artikel or ''),
                        'materiaal_id': '',
                        'stock': float(stock),
                        'bestellen': float(saldo),
                        'bruto': float(bruto) if bruto is not None else 0.0,
                    })
                except (ValueError, TypeError):
                    continue

        wb.close()

        if results:
            self.save_snapshot(date_str, results)
            return date_str
        return None

    def delete_all_snapshots(self):
        """Delete all snapshots from the database."""
        self.conn.execute("DELETE FROM snapshot_rows")
        self.conn.execute("DELETE FROM snapshots")
        self.conn.commit()

    def get_pivot_data(self):
        """Return raw snapshot data as a pivot structure.

        Returns None if no snapshots exist.
        Otherwise returns:
            {
                'dates': ['2025-11-13', '2025-12-02', ...],  # ISO, sorted
                'materials': {
                    'HSP 18mm MELxMEL wit': {
                        '2025-11-13': {'stock': 488.8, 'saldo': 460.3},
                        ...
                    },
                }
            }
        """
        cur = self.conn.cursor()

        cur.execute("SELECT COUNT(*) FROM snapshots")
        if cur.fetchone()[0] == 0:
            return None

        # Single JOIN query for all data
        cur.execute("""
            SELECT s.snapshot_date, r.material, r.artikel_nummer, r.stock_m2, r.saldo_m2
            FROM snapshot_rows r
            JOIN snapshots s ON s.id = r.snapshot_id
            ORDER BY s.snapshot_date, r.material
        """)

        dates_set = set()
        materials = {}
        artikel_nrs = {}

        for snap_date, material, artikel_nr, stock, saldo in cur.fetchall():
            dates_set.add(snap_date)
            if material not in materials:
                materials[material] = {}
            materials[material][snap_date] = {'stock': stock, 'saldo': saldo}
            if artikel_nr:
                artikel_nrs[material] = artikel_nr

        return {
            'dates': sorted(dates_set),
            'materials': materials,
            'artikel_nrs': artikel_nrs,
        }

    # ── Handmagazijn methods ──────────────────────────────────────────

    def save_handmagazijn_snapshot(self, date_str, data):
        """Save (or replace) a handmagazijn snapshot for the given date.

        Args:
            date_str: Date string like "11_02_2026" (DD_MM_YYYY).
            data: Dict of {material: m2_total}.
        """
        iso_date = _to_iso(date_str)
        cur = self.conn.cursor()

        # Delete existing snapshot for this date (upsert)
        cur.execute("SELECT id FROM handmagazijn_snapshots WHERE snapshot_date = ?", (iso_date,))
        row = cur.fetchone()
        if row:
            cur.execute("DELETE FROM handmagazijn_rows WHERE snapshot_id = ?", (row[0],))
            cur.execute("DELETE FROM handmagazijn_snapshots WHERE id = ?", (row[0],))

        cur.execute(
            "INSERT INTO handmagazijn_snapshots (snapshot_date) VALUES (?)",
            (iso_date,)
        )
        snapshot_id = cur.lastrowid

        for material, m2 in data.items():
            cur.execute(
                "INSERT INTO handmagazijn_rows (snapshot_id, material, m2) VALUES (?, ?, ?)",
                (snapshot_id, material, m2)
            )

        self.conn.commit()

    def get_handmagazijn_pivot(self):
        """Return handmagazijn data as a pivot structure.

        Returns None if no snapshots exist.
        Otherwise returns:
            {
                'dates': ['2025-11-13', ...],  # ISO, sorted
                'materials': {
                    'HSP 18mm MELxMEL wit': {
                        '2025-11-13': 23.29,
                        ...
                    },
                }
            }
        """
        cur = self.conn.cursor()

        cur.execute("SELECT COUNT(*) FROM handmagazijn_snapshots")
        if cur.fetchone()[0] == 0:
            return None

        cur.execute("""
            SELECT s.snapshot_date, r.material, r.m2
            FROM handmagazijn_rows r
            JOIN handmagazijn_snapshots s ON s.id = r.snapshot_id
            ORDER BY s.snapshot_date, r.material
        """)

        dates_set = set()
        materials = {}

        for snap_date, material, m2 in cur.fetchall():
            dates_set.add(snap_date)
            if material not in materials:
                materials[material] = {}
            materials[material][snap_date] = m2

        return {
            'dates': sorted(dates_set),
            'materials': materials,
        }

    def import_handmagazijn_csv(self, filepath):
        """Import a single DD_MM_YYYY_autofit.Csv file.

        Parses the CSV, filters rows where Referentie nummer is between
        10,000 and 100,000, aggregates m² by Materiaal, and saves snapshot.

        Returns the date string on success, 'EXISTS' if already imported,
        None if no valid data or failed.
        """
        import re

        basename = Path(filepath).name
        match = re.search(r'(\d{2}_\d{2}_\d{4})_autofit', basename, re.IGNORECASE)
        if not match:
            return None

        date_str = match.group(1)
        iso_date = _to_iso(date_str)

        # Skip if already in DB
        cur = self.conn.cursor()
        cur.execute("SELECT id FROM handmagazijn_snapshots WHERE snapshot_date = ?", (iso_date,))
        if cur.fetchone():
            return 'EXISTS'

        aggregated = {}  # {material: total_m2}

        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f, delimiter=';')
                header = None
                for row in reader:
                    if header is None:
                        header = row
                        continue
                    if len(row) < 10:
                        continue

                    # Column indices (0-based):
                    # 0=(row nr), 1=Materiaal, 2=Dikte, 3=Nerf, 4=Lengte,
                    # 5=Breedte, 6=Aantal platen, 7=Referentie, 8=Bemerking, 9=Referentie nummer
                    try:
                        ref_str = row[9].strip()
                        if not ref_str:
                            continue
                        ref_num = float(ref_str)
                        if ref_num < 10000 or ref_num >= 100000:
                            continue

                        materiaal = row[1].strip()
                        lengte = float(row[4].strip().replace(',', '.'))
                        breedte = float(row[5].strip().replace(',', '.'))
                        aantal = float(row[6].strip().replace(',', '.'))

                        m2 = (lengte * breedte * aantal) / 1_000_000
                        aggregated[materiaal] = aggregated.get(materiaal, 0) + m2
                    except (ValueError, IndexError):
                        continue
        except Exception:
            return None

        if aggregated:
            # Round values for cleaner display
            aggregated = {k: round(v, 2) for k, v in aggregated.items()}
            self.save_handmagazijn_snapshot(date_str, aggregated)
            return date_str
        return None

    def backfill_handmagazijn_zeros(self):
        """Fill in m2=0 rows for materials missing from snapshots.

        Ensures every material that appears in any snapshot has a row
        in every snapshot (with m2=0 if it was absent).
        """
        cur = self.conn.cursor()

        # Get all unique materials
        cur.execute("SELECT DISTINCT material FROM handmagazijn_rows")
        all_materials = {row[0] for row in cur.fetchall()}
        if not all_materials:
            return

        # Get all snapshots
        cur.execute("SELECT id FROM handmagazijn_snapshots")
        snapshot_ids = [row[0] for row in cur.fetchall()]

        inserted = 0
        for sid in snapshot_ids:
            cur.execute("SELECT material FROM handmagazijn_rows WHERE snapshot_id = ?", (sid,))
            existing = {row[0] for row in cur.fetchall()}
            missing = all_materials - existing
            for mat in missing:
                cur.execute(
                    "INSERT INTO handmagazijn_rows (snapshot_id, material, m2) VALUES (?, ?, 0)",
                    (sid, mat)
                )
                inserted += 1

        if inserted > 0:
            self.conn.commit()
        return inserted

    def get_handmagazijn_snapshot_count(self):
        """Return the total number of stored handmagazijn snapshots."""
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM handmagazijn_snapshots")
        return cur.fetchone()[0]

    def close(self):
        """Close the database connection."""
        if self.conn:
            self.conn.close()
            self.conn = None
