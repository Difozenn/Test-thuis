#!/usr/bin/env python3
"""
Script to inspect the structure of an Excel file
"""

import sys

try:
    import openpyxl
    import xlrd
    print("✓ openpyxl and xlrd are available")
except ImportError as e:
    print(f"Missing library: {e}")
    print("\nPlease install required libraries:")
    print("pip install openpyxl xlrd")
    sys.exit(1)

def inspect_xls_file(file_path):
    """Inspect an .xls file using xlrd"""
    print(f"\n{'='*60}")
    print(f"Inspecting: {file_path}")
    print('='*60)

    try:
        # Try to open with xlrd (for older .xls files)
        workbook = xlrd.open_workbook(file_path, formatting_info=False)
        print(f"\n✓ Successfully opened with xlrd")
        print(f"Number of sheets: {workbook.nsheets}")
        print("\nSheet names:")

        for idx, sheet_name in enumerate(workbook.sheet_names()):
            print(f"  {idx + 1}. {sheet_name}")

        # Look for the 6_Rendement sheet
        target_sheet = "6_Rendement"
        if target_sheet in workbook.sheet_names():
            print(f"\n{'='*60}")
            print(f"Found '{target_sheet}' sheet - Inspecting data...")
            print('='*60)

            sheet = workbook.sheet_by_name(target_sheet)
            print(f"Rows: {sheet.nrows}, Columns: {sheet.ncols}")

            # Print first 15 rows to understand structure
            print("\nFirst 15 rows:")
            for row_idx in range(min(15, sheet.nrows)):
                row_data = []
                for col_idx in range(min(10, sheet.ncols)):  # First 10 columns
                    cell = sheet.cell(row_idx, col_idx)
                    row_data.append(str(cell.value)[:20])  # Limit cell display to 20 chars
                print(f"Row {row_idx}: {row_data}")

            # Focus on columns A and B starting from row 7
            print(f"\n{'='*60}")
            print("Data from row 6 onwards (0-indexed row 6 = Excel row 7)")
            print("Column A (Materiaal) | Column B (Netto m²)")
            print('='*60)

            for row_idx in range(6, min(30, sheet.nrows)):  # Start from row 6 (Excel row 7)
                col_a = sheet.cell(row_idx, 0).value  # Column A
                col_b = sheet.cell(row_idx, 1).value  # Column B
                print(f"Row {row_idx} (Excel {row_idx+1}): {col_a} | {col_b}")

        else:
            print(f"\n⚠ Sheet '{target_sheet}' not found!")

    except Exception as e:
        print(f"✗ Error reading file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    file_path = "Stuklijsten/S03548_0703_MO04584_Halkast_(1-1)_Ukuqonda-Meylemans.xls"
    inspect_xls_file(file_path)
