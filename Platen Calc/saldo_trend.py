#!/usr/bin/env python3
"""
Saldo Trend - Weekly saldo trend report from historical bestelberekening exports.
Reads bestelberekening_DD_MM_YYYY.xlsx files and generates a trend analysis.
"""

import glob
import os
import re
from datetime import datetime
from collections import defaultdict

import openpyxl
import xlsxwriter


def parse_date_from_filename(filename):
    """Extract date from bestelberekening_DD_MM_YYYY.xlsx filename."""
    match = re.search(r'bestelberekening_(\d{2})_(\d{2})_(\d{4})', filename)
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        return datetime(year, month, day)
    return None


def read_saldo_from_file(filepath):
    """Read Materiaal (col A) and Saldo (col I) from a bestelberekening xlsx file.

    Returns dict {material_name: saldo_value}.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Try sheet name "Bestelberekening", fall back to first sheet
    if "Bestelberekening" in wb.sheetnames:
        ws = wb["Bestelberekening"]
    else:
        ws = wb.worksheets[0]

    saldo_data = {}
    first_row = True
    for row in ws.iter_rows(min_col=1, max_col=9, values_only=True):
        # Skip the header row
        if first_row:
            first_row = False
            continue

        material = row[0]  # Column A
        saldo = row[8]     # Column I (Saldo)

        if material and saldo is not None:
            try:
                saldo_data[str(material)] = float(saldo)
            except (ValueError, TypeError):
                continue

    wb.close()
    return saldo_data


def generate_saldo_trend(folder_path, output_path=None):
    """Generate a saldo trend Excel report from historical bestelberekening files.

    Args:
        folder_path: Folder containing bestelberekening_*.xlsx files
        output_path: Where to save the output. Defaults to folder_path/saldo_trend.xlsx

    Returns:
        Path to the generated file, or None on failure.
    """
    if output_path is None:
        output_path = os.path.join(folder_path, "saldo_trend.xlsx")

    # 1. Scan for matching files
    pattern = os.path.join(folder_path, "bestelberekening_*.xlsx")
    files = glob.glob(pattern)

    if not files:
        raise FileNotFoundError(
            f"Geen bestelberekening_*.xlsx bestanden gevonden in:\n{folder_path}"
        )

    # 2. Read each file and parse date
    file_entries = []  # [(date, filepath)]
    for filepath in files:
        basename = os.path.basename(filepath)
        # Skip the trend output file itself
        if basename == "saldo_trend.xlsx":
            continue
        date = parse_date_from_filename(basename)
        if date:
            file_entries.append((date, filepath))

    if not file_entries:
        raise FileNotFoundError(
            "Geen geldige bestelberekening_DD_MM_YYYY.xlsx bestanden gevonden."
        )

    # Sort by date
    file_entries.sort(key=lambda x: x[0])

    # 3. Group by ISO week, keep latest file per week
    week_files = {}  # {(year, week): (date, filepath)}
    for date, filepath in file_entries:
        iso = date.isocalendar()
        key = (iso[0], iso[1])  # (year, week)
        # Keep the latest file per week
        if key not in week_files or date > week_files[key][0]:
            week_files[key] = (date, filepath)

    # Sort weeks chronologically
    sorted_weeks = sorted(week_files.keys())

    # 4. Read saldo data per week
    # {material: {(year, week): saldo}}
    pivot = defaultdict(dict)
    for week_key in sorted_weeks:
        date, filepath = week_files[week_key]
        saldo_data = read_saldo_from_file(filepath)
        for material, saldo in saldo_data.items():
            pivot[material][week_key] = saldo

    # Sort materials alphabetically
    sorted_materials = sorted(pivot.keys())

    # Format week labels: "Wk 05 '25"
    week_labels = []
    for year, week in sorted_weeks:
        short_year = str(year)[-2:]
        week_labels.append(f"Wk {week:02d} '{short_year}")

    # 5. Generate Excel with xlsxwriter
    workbook = xlsxwriter.Workbook(output_path)

    # --- Sheet 1: Trend Data ---
    ws_data = workbook.add_worksheet("Trend Data")

    # Formats
    header_fmt = workbook.add_format({
        'bold': True,
        'bg_color': '#1a73e8',
        'font_color': 'white',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
    })
    material_fmt = workbook.add_format({
        'bold': True,
        'border': 1,
    })
    num_fmt = workbook.add_format({
        'num_format': '0.00',
        'border': 1,
        'align': 'right',
    })
    red_fmt = workbook.add_format({
        'num_format': '0.00',
        'bg_color': '#ffebee',
        'font_color': '#c62828',
        'border': 1,
        'align': 'right',
    })
    green_fmt = workbook.add_format({
        'num_format': '0.00',
        'bg_color': '#e8f5e9',
        'font_color': '#2e7d32',
        'border': 1,
        'align': 'right',
    })

    # Header row
    ws_data.write(0, 0, "Materiaal", header_fmt)
    for col_idx, label in enumerate(week_labels, 1):
        ws_data.write(0, col_idx, label, header_fmt)

    # Data rows
    for row_idx, material in enumerate(sorted_materials, 1):
        ws_data.write(row_idx, 0, material, material_fmt)
        for col_idx, week_key in enumerate(sorted_weeks, 1):
            value = pivot[material].get(week_key)
            if value is not None:
                ws_data.write_number(row_idx, col_idx, value, num_fmt)
            else:
                ws_data.write_blank(row_idx, col_idx, None, num_fmt)

    # Column widths
    ws_data.set_column(0, 0, 35)  # Material name
    ws_data.set_column(1, len(sorted_weeks), 12)  # Week columns

    # Conditional formatting on data cells
    if sorted_materials and sorted_weeks:
        last_row = len(sorted_materials)
        last_col = len(sorted_weeks)
        ws_data.conditional_format(1, 1, last_row, last_col, {
            'type': 'cell',
            'criteria': '<',
            'value': 0,
            'format': red_fmt,
        })
        ws_data.conditional_format(1, 1, last_row, last_col, {
            'type': 'cell',
            'criteria': '>',
            'value': 0,
            'format': green_fmt,
        })

    # Freeze top row and first column
    ws_data.freeze_panes(1, 1)

    # --- Sheet 2: Grafiek (Chart) ---
    ws_chart = workbook.add_worksheet("Grafiek")

    chart = workbook.add_chart({'type': 'line'})
    chart.set_title({'name': 'Saldo Trend per Materiaal'})
    chart.set_x_axis({
        'name': 'Week',
        'categories': ['Trend Data', 0, 1, 0, len(sorted_weeks)],
    })
    chart.set_y_axis({
        'name': 'Saldo (m\u00b2)',
    })
    chart.set_size({'width': 1200, 'height': 600})
    chart.set_legend({'position': 'bottom'})

    # Add a series per material
    for row_idx, material in enumerate(sorted_materials, 1):
        chart.add_series({
            'name': ['Trend Data', row_idx, 0],
            'categories': ['Trend Data', 0, 1, 0, len(sorted_weeks)],
            'values': ['Trend Data', row_idx, 1, row_idx, len(sorted_weeks)],
            'line': {'width': 2},
        })

    ws_chart.insert_chart('A1', chart)

    workbook.close()
    return output_path
